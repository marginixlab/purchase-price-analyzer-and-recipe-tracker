print("RUNNING MAIN FROM:", __file__)

import io
import json
import logging
import random
import re
import shutil
import contextvars
import os
import uuid
import base64
import hashlib
import hmac
import secrets
import threading
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from time import monotonic, perf_counter
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request as UrlRequest, urlopen

import pandas as pd
from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from jinja2 import TemplateError, TemplateNotFound
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s"
)
logger = logging.getLogger("price_analyzer")


def log_perf(label: str, started_at: float) -> float:
    elapsed_ms = (perf_counter() - started_at) * 1000
    logger.info("[PERF] %s: %.1fms", label, elapsed_ms)
    return elapsed_ms


def log_perf_details(label: str, **details: Any) -> None:
    logger.info("[PERF] %s: %s", label, json.dumps(make_json_safe(details), ensure_ascii=False, separators=(",", ":")))


def env_flag(name: str, default: bool) -> bool:
    raw_value = os.getenv(name)
    if raw_value is None:
        return default
    return raw_value.strip().lower() in {"1", "true", "yes", "on"}


def env_int(name: str, default: int) -> int:
    raw_value = os.getenv(name)
    if raw_value is None:
        return default
    try:
        parsed_value = int(raw_value)
    except ValueError:
        logger.warning("Invalid integer for %s; falling back to default.", name)
        return default
    return parsed_value if parsed_value > 0 else default

BASE_DIR = Path(__file__).resolve().parent
PERSIST_DIR = Path(os.getenv("PERSIST_DIR", "/var/data")).expanduser()
PERSIST_DIR.mkdir(parents=True, exist_ok=True)
STATIC_DIR = BASE_DIR / "static"
TEMPLATES_DIR = BASE_DIR / "templates"
DATA_DIR = BASE_DIR / "data"
USER_DATA_DIR = DATA_DIR / "users"
RECIPES_DATA_DIR = PERSIST_DIR / "recipes"
INDEX_TEMPLATE = "index.html"
LATEST_RESULTS_PATH = BASE_DIR / "latest_results.csv"
ANALYSIS_HISTORY_PATH = BASE_DIR / "analysis_history.json"
CODES_PATH = BASE_DIR / "codes.json"
RECIPES_PATH = BASE_DIR / "recipes.json"
QUOTE_COMPARISONS_PATH = BASE_DIR / "quote_comparisons.json"
QUOTE_COMPARE_UPLOAD_CACHE_DIR = BASE_DIR / ".quote_compare_uploads"
QUOTE_COMPARE_SESSION_CACHE_DIR = BASE_DIR / ".quote_compare_sessions"
DEMO_RECIPES_SESSION_DIR = BASE_DIR / ".demo_recipe_sessions"
GUIDE_KNOWLEDGE_PATH = BASE_DIR / "guide_knowledge.json"
LATEST_ANALYSIS_CACHE: dict[str, Any] = {
    "signature": None,
    "context": None
}
QUOTE_COMPARE_STORE_CACHE: dict[str, Any] = {
    "signature": None,
    "store": None
}
ANALYSIS_HISTORY_CACHE: dict[str, Any] = {
    "signature": None,
    "store": None
}
CURRENT_UPLOAD_ANALYSIS_FRAME_CACHE: dict[str, Any] = {
    "signature": None,
    "frame": None
}
RECIPES_STORE_CACHE: dict[str, Any] = {
    "signature": None,
    "store": None
}
RECIPE_ANALYSIS_CACHE: dict[str, Any] = {
    "signature": None,
    "frame": None,
    "product_catalog": None,
    "pricing_lookup": None
}
LATEST_ANALYSIS_UPLOAD_META_CACHE: dict[str, Any] = {
    "signature": None,
    "meta": None
}
RECIPES_BOOTSTRAP_RESPONSE_CACHE: dict[str, Any] = {
    "signature": None,
    "response_json": None
}
CURRENT_STORAGE_USER_ID: contextvars.ContextVar[int | None] = contextvars.ContextVar(
    "current_storage_user_id",
    default=None
)
RECIPE_BOOTSTRAP_METRICS_CONTEXT: contextvars.ContextVar[dict[str, Any] | None] = contextvars.ContextVar(
    "recipe_bootstrap_metrics_context",
    default=None
)
OPTIONAL_UNIT_COLUMNS = ["Purchase Unit", "Unit"]
OPTIONAL_QUANTITY_COLUMNS = ["Quantity", "Qty"]
OPTIONAL_UNIT_PRICE_COLUMNS = ["Unit Price", "Price"]
READABLE_CODE_ALPHABET = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
REQUIRED_ANALYSIS_FIELDS = [
    "Product Name",
    "Supplier",
    "Unit",
    "Quantity",
    "Unit Price",
    "Date"
]
REQUIRED_FIELD_SYNONYMS = {
    "Product Name": [
        "product",
        "product name",
        "item",
        "item name",
        "description",
        "item description",
        "product description",
        "material",
        "article"
    ],
    "Supplier": [
        "supplier",
        "supplier name",
        "vendor",
        "vendor name",
        "company",
        "company name",
        "seller"
    ],
    "Unit": [
        "unit",
        "purchase unit",
        "uom",
        "unit of measure",
        "measure",
        "pack size",
        "size"
    ],
    "Quantity": [
        "quantity",
        "qty",
        "amount",
        "ordered qty",
        "ordered quantity",
        "order quantity",
        "purchase quantity",
        "units ordered"
    ],
    "Unit Price": [
        "unit price",
        "price",
        "purchase price",
        "cost",
        "unit cost",
        "item cost",
        "price per unit",
        "invoice price"
    ],
    "Date": [
        "date",
        "invoice date",
        "purchase date",
        "transaction date",
        "posting date",
        "order date",
        "document date"
    ]
}
QUOTE_COMPARE_REQUIRED_FIELDS = [
    "Product Name",
    "Supplier",
    "Unit",
    "Quantity",
    "Unit Price",
    "Date"
]
QUOTE_COMPARE_OPTIONAL_FIELDS = [
    "Total Price",
    "Currency",
    "Delivery Time",
    "Payment Terms",
    "Valid Until",
    "Notes"
]
QUOTE_COMPARE_FIELD_SYNONYMS = {
    "Supplier": [
        "supplier",
        "supplier name",
        "vendor",
        "vendor name",
        "company",
        "seller"
    ],
    "Product Name": [
        "product",
        "product name",
        "item",
        "item name",
        "description",
        "product description"
    ],
    "Unit": [
        "unit",
        "uom",
        "purchase unit",
        "unit of measure",
        "measure"
    ],
    "Quantity": [
        "quantity",
        "qty",
        "amount",
        "order quantity",
        "units"
    ],
    "Unit Price": [
        "unit price",
        "price",
        "quote price",
        "quoted price",
        "cost",
        "unit cost"
    ],
    "Total Price": [
        "total",
        "total price",
        "line total",
        "quote total",
        "extended price"
    ],
    "Currency": [
        "currency",
        "currency code"
    ],
    "Delivery Time": [
        "delivery",
        "delivery time",
        "delivery days",
        "lead time"
    ],
    "Payment Terms": [
        "payment",
        "payment term",
        "terms",
        "payment terms"
    ],
    "Valid Until": [
        "valid until",
        "validity",
        "expiry",
        "expiry date",
        "expiration date"
    ],
    "Notes": [
        "notes",
        "comment",
        "comments",
        "remark",
        "remarks"
    ],
    "Date": [
        "date",
        "quote date",
        "offer date",
        "pricing date",
        "submitted date"
    ]
}
RECIPE_PRICING_MODES = {
    "latest_price": "Latest Price",
    "average_price": "Average Price"
}
RECIPE_PRICING_GOAL_TYPES = {
    "food_cost_pct": "Target Food Cost %",
    "gross_margin_pct": "Target Gross Margin %",
    "markup_pct": "Markup %"
}
QUOTE_COMPARE_DEFAULT_WEIGHTS = {
    "price": 0.5,
    "delivery": 0.25,
    "payment": 0.25
}
ANALYSIS_SCOPE_OPTIONS = [
    {"value": "current_upload", "label": "Current File"}
]
ANALYSIS_DEDUPE_TEXT_COLUMNS = {
    "Supplier": "supplier",
    "Product Name": "product",
    "Unit": "unit"
}
APP_ENV = (
    os.getenv("APP_ENV")
    or os.getenv("ENV")
    or os.getenv("FASTAPI_ENV")
    or "development"
).strip().lower()
IS_PRODUCTION = APP_ENV in {"prod", "production"}
DEBUG_MODE = env_flag("DEBUG", not IS_PRODUCTION)
MAX_UPLOAD_SIZE_BYTES = env_int("MAX_UPLOAD_SIZE_BYTES", 10 * 1024 * 1024)
PASSWORD_RESET_TOKEN_TTL_SECONDS = env_int("PASSWORD_RESET_TOKEN_TTL_SECONDS", 60 * 60)
RATE_LIMIT_STORE: dict[str, list[float]] = {}
RATE_LIMIT_LOCK = threading.Lock()
UPLOAD_ROUTE_PATHS = {
    "/quote-compare/upload/inspect",
    "/quote-compare/upload/confirm"
}
RATE_LIMIT_RULES = {
    "login": {"limit": env_int("RATE_LIMIT_LOGIN_MAX_ATTEMPTS", 10), "window_seconds": env_int("RATE_LIMIT_LOGIN_WINDOW_SECONDS", 60)},
    "forgot_password": {"limit": env_int("RATE_LIMIT_FORGOT_PASSWORD_MAX_ATTEMPTS", 5), "window_seconds": env_int("RATE_LIMIT_FORGOT_PASSWORD_WINDOW_SECONDS", 300)},
    "reset_password": {"limit": env_int("RATE_LIMIT_RESET_PASSWORD_MAX_ATTEMPTS", 5), "window_seconds": env_int("RATE_LIMIT_RESET_PASSWORD_WINDOW_SECONDS", 300)},
    "admin_license_generate": {"limit": env_int("RATE_LIMIT_ADMIN_LICENSE_GENERATE_MAX_ATTEMPTS", 10), "window_seconds": env_int("RATE_LIMIT_ADMIN_LICENSE_GENERATE_WINDOW_SECONDS", 60)}
}


def format_upload_size_limit(limit_bytes: int) -> str:
    if limit_bytes % (1024 * 1024) == 0:
        return f"{limit_bytes // (1024 * 1024)} MB"
    return f"{round(limit_bytes / (1024 * 1024), 1)} MB"


def get_upload_size_limit_message(limit_bytes: int = MAX_UPLOAD_SIZE_BYTES) -> str:
    return f"File is too large. Maximum upload size is {format_upload_size_limit(limit_bytes)}."


def get_request_client_identifier(request: Request) -> str:
    forwarded_for = str(request.headers.get("x-forwarded-for") or "").split(",")[0].strip()
    real_ip = str(request.headers.get("x-real-ip") or "").strip()
    client_host = getattr(request.client, "host", "") or ""
    return forwarded_for or real_ip or client_host or "unknown"


def check_rate_limit(bucket: str, identifier: str, *, limit: int, window_seconds: int) -> int | None:
    now = monotonic()
    bucket_key = f"{bucket}:{identifier}"
    window_start = now - window_seconds
    with RATE_LIMIT_LOCK:
        attempts = RATE_LIMIT_STORE.get(bucket_key, [])
        attempts = [attempt for attempt in attempts if attempt > window_start]
        if len(attempts) >= limit:
            retry_after = max(1, int(min(attempts) + window_seconds - now))
            RATE_LIMIT_STORE[bucket_key] = attempts
            return retry_after
        attempts.append(now)
        RATE_LIMIT_STORE[bucket_key] = attempts
    return None


def build_rate_limited_json_response(message: str, retry_after: int) -> JSONResponse:
    return JSONResponse(
        content={"success": False, "message": message, "detail": message},
        status_code=429,
        headers={"Retry-After": str(retry_after)}
    )


def check_named_rate_limit(
    request: Request,
    rule_name: str,
    *,
    identifier_suffix: str = ""
) -> int | None:
    rule = RATE_LIMIT_RULES[rule_name]
    client_identifier = get_request_client_identifier(request)
    identifier = f"{client_identifier}|{identifier_suffix}" if identifier_suffix else client_identifier
    return check_rate_limit(
        rule_name,
        identifier,
        limit=int(rule["limit"]),
        window_seconds=int(rule["window_seconds"])
    )


def ensure_upload_size_within_limit(
    file: UploadFile,
    *,
    limit_bytes: int = MAX_UPLOAD_SIZE_BYTES
) -> int:
    filename = file.filename or "uploaded file"
    try:
        file.file.seek(0, os.SEEK_END)
        size_bytes = int(file.file.tell())
        file.file.seek(0)
    except Exception as exc:
        logger.exception("Failed to inspect upload size: %s", filename)
        raise ValueError("The uploaded file could not be prepared for reading.") from exc

    if size_bytes <= 0:
        raise ValueError("The uploaded file is empty.")
    if size_bytes > limit_bytes:
        raise ValueError(get_upload_size_limit_message(limit_bytes))
    return size_bytes


def build_upload_size_error_response() -> JSONResponse:
    return JSONResponse(
        content={"success": False, "message": get_upload_size_limit_message()},
        status_code=413
    )
ANALYSIS_DEDUPE_NUMBER_COLUMNS = {
    "Quantity": 6,
    "Unit Price": 4
}


def ensure_app_paths() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    USER_DATA_DIR.mkdir(parents=True, exist_ok=True)
    RECIPES_DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not STATIC_DIR.is_dir():
        raise RuntimeError(f"Static directory not found: {STATIC_DIR}")
    if not TEMPLATES_DIR.is_dir():
        raise RuntimeError(f"Templates directory not found: {TEMPLATES_DIR}")


def get_current_user_id(request: Request | None) -> int | None:
    if request is None:
        return None

    raw_user_id = getattr(getattr(request, "state", None), "auth_user_id", None)
    try:
        return int(raw_user_id) if raw_user_id is not None else None
    except (TypeError, ValueError):
        return None


def get_storage_user_id(user_id: int | str | None = None) -> int | None:
    if user_id is not None:
        try:
            return int(user_id)
        except (TypeError, ValueError):
            return None
    return CURRENT_STORAGE_USER_ID.get()


def get_user_storage_root(user_id: int | str | None) -> Path:
    normalized_user_id = str(user_id or "").strip() or "anonymous"
    user_root = USER_DATA_DIR / normalized_user_id
    user_root.mkdir(parents=True, exist_ok=True)
    return user_root


def get_user_latest_results_path(user_id: int | str | None) -> Path:
    return get_user_storage_root(user_id) / "latest_results.csv"


def get_current_latest_results_path(user_id: int | str | None = None) -> Path:
    return get_user_latest_results_path(get_storage_user_id(user_id))


def get_user_recipes_path(user_id: int | str | None) -> Path:
    normalized_user_id = str(user_id or "").strip() or "anonymous"
    return RECIPES_DATA_DIR / f"{normalized_user_id}.json"


def get_user_quote_comparisons_path(user_id: int | str | None) -> Path:
    return get_user_storage_root(user_id) / "quote_comparisons.json"


def get_user_analysis_history_path(user_id: int | str | None) -> Path:
    return get_user_storage_root(user_id) / "analysis_history.json"


def get_user_upload_cache_dir(user_id: int | str | None) -> Path:
    upload_cache_dir = get_user_storage_root(user_id) / "quote_compare_uploads"
    upload_cache_dir.mkdir(parents=True, exist_ok=True)
    return upload_cache_dir


def get_user_session_cache_dir(user_id: int | str | None) -> Path:
    session_cache_dir = get_user_storage_root(user_id) / "quote_compare_sessions"
    session_cache_dir.mkdir(parents=True, exist_ok=True)
    return session_cache_dir


def get_user_session_file_path(user_id: int | str | None, session_id: str) -> Path:
    normalized_session_id = str(session_id or "").strip()
    return get_user_session_cache_dir(user_id) / f"{normalized_session_id}.json"


def get_user_active_quote_compare_session_path(user_id: int | str | None) -> Path:
    return get_user_storage_root(user_id) / "active_quote_compare_session.json"


def build_templates() -> Jinja2Templates:
    ensure_app_paths()
    templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
    try:
        templates.get_template(INDEX_TEMPLATE)
    except TemplateNotFound as exc:
        raise RuntimeError(f"Required template is missing: {INDEX_TEMPLATE}") from exc
    except TemplateError as exc:
        raise RuntimeError(f"Template validation failed for {INDEX_TEMPLATE}: {exc}") from exc
    return templates


def safe_template_response(
    request: Request,
    name: str,
    context: dict,
    *,
    status_code: int = 200
):
    try:
        return templates.TemplateResponse(
            request=request,
            name=name,
            context=context,
            status_code=status_code
        )
    except TemplateNotFound:
        logger.exception("Template not found while rendering %s", name)
        return HTMLResponse(
            "<h1>Application template missing</h1><p>Please check server logs for details.</p>",
            status_code=500
        )
    except TemplateError:
        logger.exception("Template error while rendering %s", name)
        return HTMLResponse(
            "<h1>Application template error</h1><p>Please check server logs for details.</p>",
            status_code=500
        )


def redirect_if_authenticated(request: Request) -> RedirectResponse | None:
    auth_user_id = getattr(request.state, "auth_user_id", None)
    if auth_user_id:
        return RedirectResponse(url="/", status_code=303)
    return None


def build_sample_dataframe() -> pd.DataFrame:
    return pd.DataFrame({
        "Product Name": [
            "Eggs",
            "Eggs",
            "Eggs",
            "Eggs",
            "Olive Oil",
            "Olive Oil",
            "Rice",
            "Rice",
            "Coffee Beans",
            "Coffee Beans",
            "Bread",
            "Bread",
            "Sugar",
            "Sugar",
            "Olive Oil",
            "Rice"
        ],
        "Supplier": [
            "Sysco",
            "US Foods",
            "Metro",
            "Metro",
            "Sysco",
            "US Foods",
            "Metro",
            "Sysco",
            "US Foods",
            "Metro",
            "Sysco",
            "US Foods",
            "Metro",
            "Sysco",
            "US Foods",
            "Metro"
        ],
        "Unit": [
            "piece",
            "piece",
            "carton",
            "carton",
            "liter",
            "liter",
            "kg",
            "kg",
            "kg",
            "kg",
            "loaf",
            "loaf",
            "kg",
            "kg",
            "liter",
            "kg"
        ],
        "Quantity": [
            24,
            12,
            3,
            2,
            8,
            5,
            50,
            30,
            12,
            8,
            20,
            14,
            18,
            10,
            6,
            40
        ],
        "Unit Price": [
            0.42,
            0.39,
            4.80,
            5.20,
            12.60,
            13.10,
            2.10,
            2.35,
            17.40,
            18.10,
            3.25,
            3.55,
            1.28,
            1.42,
            14.00,
            2.18
        ],
        "Date": [
            "2026-03-01",
            "2026-03-02",
            "2026-03-03",
            "2026-03-04",
            "2026-03-05",
            "2026-03-06",
            "2026-03-07",
            "2026-03-08",
            "2026-03-09",
            "2026-03-10",
            "2026-03-11",
            "2026-03-12",
            "2026-03-13",
            "2026-03-14",
            "2026-03-15",
            "2026-03-16"
        ]
    })


def build_quote_compare_sample_dataframe() -> pd.DataFrame:
    return pd.DataFrame({
        "Supplier": [
            "Atlas Packaging",
            "Blue Harbor Supply",
            "Northline Goods",
            "Atlas Packaging",
            "Blue Harbor Supply",
            "Northline Goods",
            "Atlas Packaging",
            "Blue Harbor Supply",
            "Northline Goods"
        ],
        "Product Name": [
            "8 oz amber glass jar",
            "8 oz amber glass jar",
            "8 oz amber glass jar",
            "Black phenolic lid 70-400",
            "Black phenolic lid 70-400",
            "Black phenolic lid 70-400",
            "Kraft shipping box 6x6x4",
            "Kraft shipping box 6x6x4",
            "Kraft shipping box 6x6x4"
        ],
        "Unit": [
            "case",
            "case",
            "case",
            "case",
            "case",
            "case",
            "bundle",
            "bundle",
            "bundle"
        ],
        "Quantity": [
            20,
            20,
            20,
            20,
            20,
            20,
            40,
            40,
            40
        ],
        "Unit Price": [
            18.60,
            17.95,
            18.25,
            9.40,
            9.10,
            9.55,
            6.80,
            7.10,
            6.65
        ],
        "Total Price": [
            372.00,
            359.00,
            365.00,
            188.00,
            182.00,
            191.00,
            272.00,
            284.00,
            266.00
        ],
        "Currency": [
            "USD",
            "USD",
            "USD",
            "USD",
            "USD",
            "USD",
            "USD",
            "USD",
            "USD"
        ],
        "Delivery Time": [
            "12 days",
            "16 days",
            "10 days",
            "12 days",
            "16 days",
            "10 days",
            "12 days",
            "16 days",
            "10 days"
        ],
        "Payment Terms": [
            "Net 30",
            "Net 45",
            "Net 21",
            "Net 30",
            "Net 45",
            "Net 21",
            "Net 30",
            "Net 45",
            "Net 21"
        ],
        "Valid Until": [
            "2026-04-30",
            "2026-04-25",
            "2026-04-28",
            "2026-04-30",
            "2026-04-25",
            "2026-04-28",
            "2026-04-30",
            "2026-04-25",
            "2026-04-28"
        ],
        "Notes": [
            "Includes pallet wrap",
            "Freight billed separately",
            "Fastest lead time",
            "Includes liner",
            "Best payment term",
            "Smaller MOQ",
            "Standard corrugate",
            "Freight included",
            "Best landed price"
        ]
    })


def build_demo_analysis_dataframe() -> pd.DataFrame:
    suppliers = [
        {"name": "Sysco", "delivery": "2 days", "terms": "Net 14", "price_bias": 1.04},
        {"name": "US Foods", "delivery": "3 days", "terms": "Net 21", "price_bias": 1.00},
        {"name": "Restaurant Depot", "delivery": "1 day", "terms": "Due on receipt", "price_bias": 0.96},
        {"name": "Gordon Food Service", "delivery": "4 days", "terms": "Net 30", "price_bias": 1.02}
    ]
    products = [
        {"name": "Espresso Beans", "unit": "lb", "quantity": 18, "base_price": 9.85, "notes": "House espresso blend"},
        {"name": "Whole Milk", "unit": "case", "quantity": 6, "base_price": 23.40, "notes": "4 x 1 gallon"},
        {"name": "Oat Milk", "unit": "case", "quantity": 4, "base_price": 27.80, "notes": "12 barista cartons"},
        {"name": "Heavy Cream", "unit": "case", "quantity": 3, "base_price": 31.20, "notes": "Dairy program pricing"},
        {"name": "Vanilla Syrup", "unit": "case", "quantity": 2, "base_price": 34.60, "notes": "6 bottle pack"},
        {"name": "Chocolate Syrup", "unit": "case", "quantity": 2, "base_price": 29.40, "notes": "Cafe beverage station"},
        {"name": "Cane Sugar", "unit": "lb", "quantity": 25, "base_price": 0.92, "notes": "Bakery prep stock"},
        {"name": "All-Purpose Flour", "unit": "lb", "quantity": 50, "base_price": 0.68, "notes": "High-turn bakery item"},
        {"name": "Butter", "unit": "lb", "quantity": 20, "base_price": 3.42, "notes": "Unsalted"},
        {"name": "Croissant Dough", "unit": "case", "quantity": 5, "base_price": 78.00, "notes": "Frozen laminated dough"},
        {"name": "Bagels", "unit": "case", "quantity": 4, "base_price": 32.40, "notes": "Assorted breakfast mix"},
        {"name": "Sourdough Bread", "unit": "case", "quantity": 3, "base_price": 41.80, "notes": "Sandwich station"},
        {"name": "Turkey Breast", "unit": "lb", "quantity": 15, "base_price": 4.95, "notes": "Deli sliced"},
        {"name": "Smoked Bacon", "unit": "lb", "quantity": 12, "base_price": 5.85, "notes": "Breakfast line"},
        {"name": "Avocado", "unit": "case", "quantity": 2, "base_price": 48.60, "notes": "Produce market variable"},
        {"name": "Romaine Lettuce", "unit": "case", "quantity": 3, "base_price": 27.10, "notes": "Salad station"},
        {"name": "Tomatoes", "unit": "case", "quantity": 3, "base_price": 29.80, "notes": "Slicer tomatoes"},
        {"name": "Red Onion", "unit": "case", "quantity": 2, "base_price": 21.50, "notes": "Prep staple"},
        {"name": "Cheddar Cheese", "unit": "lb", "quantity": 18, "base_price": 3.78, "notes": "Shredded blend"},
        {"name": "Mozzarella Cheese", "unit": "lb", "quantity": 18, "base_price": 3.34, "notes": "Pizza and panini"},
        {"name": "Chicken Breast", "unit": "lb", "quantity": 20, "base_price": 3.96, "notes": "Grill station"},
        {"name": "Salmon Fillet", "unit": "lb", "quantity": 12, "base_price": 9.25, "notes": "Weekend feature"},
        {"name": "Ground Beef", "unit": "lb", "quantity": 20, "base_price": 4.62, "notes": "Burger program"},
        {"name": "Burger Buns", "unit": "case", "quantity": 4, "base_price": 27.20, "notes": "Brioche style"},
        {"name": "Frozen Fries", "unit": "case", "quantity": 8, "base_price": 24.90, "notes": "Straight cut"},
        {"name": "Ketchup", "unit": "case", "quantity": 2, "base_price": 18.30, "notes": "Front of house refill"},
        {"name": "Mayonnaise", "unit": "case", "quantity": 2, "base_price": 22.40, "notes": "Sauce program"},
        {"name": "Pickles", "unit": "case", "quantity": 2, "base_price": 26.70, "notes": "Burger garnish"},
        {"name": "Tortilla Chips", "unit": "case", "quantity": 3, "base_price": 29.20, "notes": "Bar snack"},
        {"name": "Flour Tortillas", "unit": "case", "quantity": 3, "base_price": 24.50, "notes": "Wrap station"},
        {"name": "Black Beans", "unit": "case", "quantity": 2, "base_price": 19.80, "notes": "Prep pantry"},
        {"name": "Jasmine Rice", "unit": "lb", "quantity": 25, "base_price": 1.08, "notes": "Rice bowl base"},
        {"name": "Penne Pasta", "unit": "lb", "quantity": 20, "base_price": 1.14, "notes": "Dinner line"},
        {"name": "Marinara Sauce", "unit": "case", "quantity": 3, "base_price": 36.20, "notes": "House red sauce"},
        {"name": "Parmesan Cheese", "unit": "lb", "quantity": 10, "base_price": 5.95, "notes": "Finishing cheese"},
        {"name": "Olive Oil", "unit": "case", "quantity": 2, "base_price": 79.50, "notes": "Kitchen premium oil"},
        {"name": "Lemons", "unit": "case", "quantity": 2, "base_price": 44.70, "notes": "Bar citrus"},
        {"name": "Limes", "unit": "case", "quantity": 2, "base_price": 42.20, "notes": "Cocktail garnish"},
        {"name": "Simple Syrup", "unit": "case", "quantity": 2, "base_price": 25.80, "notes": "Bar batch syrup"},
        {"name": "Tequila Blanco", "unit": "case", "quantity": 1, "base_price": 198.00, "notes": "12 bottle case"},
        {"name": "Vodka", "unit": "case", "quantity": 1, "base_price": 176.00, "notes": "Well spirit"},
        {"name": "IPA Draft Keg", "unit": "each", "quantity": 2, "base_price": 158.00, "notes": "Half barrel"},
        {"name": "House Red Wine", "unit": "case", "quantity": 1, "base_price": 112.00, "notes": "12 bottle case"},
        {"name": "Paper Cups 12 oz", "unit": "case", "quantity": 5, "base_price": 44.60, "notes": "To-go service"},
        {"name": "Lids 12 oz", "unit": "case", "quantity": 5, "base_price": 28.50, "notes": "Matching cup lids"},
        {"name": "Takeout Containers", "unit": "case", "quantity": 4, "base_price": 52.80, "notes": "Compostable"},
        {"name": "Napkins", "unit": "case", "quantity": 6, "base_price": 31.40, "notes": "Dining room"},
        {"name": "Dish Soap", "unit": "case", "quantity": 2, "base_price": 46.20, "notes": "Back of house"},
        {"name": "Sanitizer", "unit": "case", "quantity": 2, "base_price": 54.80, "notes": "Food-safe sanitizer"},
        {"name": "Toilet Paper", "unit": "case", "quantity": 3, "base_price": 38.60, "notes": "Restroom supplies"}
    ]
    delivery_notes = ["Stable contract pricing", "Promo lane price", "Spot market pressure", "Preferred account rate"]
    rows: list[dict[str, Any]] = []
    base_date = date(2026, 1, 6)

    for product_index, product in enumerate(products):
        for supplier_index, supplier in enumerate(suppliers):
            for week_offset in range(5):
                current_date = base_date + timedelta(days=(week_offset * 14) + ((product_index + supplier_index) % 5))
                seasonal_factor = 1 + (((product_index % 7) - 3) * 0.006) + (week_offset * 0.008)
                competitive_factor = 1 + (((supplier_index * 2) - 3) * 0.01)
                product_factor = 1 + ((product_index % 5) * 0.004)
                unit_price = round(product["base_price"] * supplier["price_bias"] * seasonal_factor * competitive_factor * product_factor, 2)
                quantity = product["quantity"] + ((week_offset + supplier_index) % 3)
                total_price = round(unit_price * quantity, 2)
                valid_until = current_date + timedelta(days=21 + (supplier_index * 3))
                rows.append({
                    "Supplier": supplier["name"],
                    "Product Name": product["name"],
                    "Unit": product["unit"],
                    "Quantity": quantity,
                    "Unit Price": unit_price,
                    "Total Price": total_price,
                    "Currency": "USD",
                    "Delivery Time": supplier["delivery"],
                    "Payment Terms": supplier["terms"],
                    "Valid Until": valid_until.strftime("%Y-%m-%d"),
                    "Notes": f"{product['notes']} | {delivery_notes[(product_index + week_offset + supplier_index) % len(delivery_notes)]}",
                    "Date": current_date.strftime("%Y-%m-%d")
                })

    return pd.DataFrame(rows)


def build_demo_recipes() -> list[dict[str, Any]]:
    return [
        {
            "recipe_id": "demo-iced-latte",
            "name": "Iced Vanilla Latte",
            "yield_portions": 1,
            "pricing_mode": "latest_price",
            "selling_price": 6.5,
            "pricing_goal_type": "food_cost_pct",
            "pricing_goal_value": 28,
            "target_food_cost_pct": 28,
            "ingredients": [
                {"product_name": "Espresso Beans", "quantity": 0.08, "unit": "lb", "purchase_unit": "lb", "purchase_size": 1, "purchase_base_unit": "lb"},
                {"product_name": "Whole Milk", "quantity": 12, "unit": "fl oz", "purchase_unit": "case", "purchase_size": 512, "purchase_base_unit": "fl oz"},
                {"product_name": "Vanilla Syrup", "quantity": 1, "unit": "fl oz", "purchase_unit": "case", "purchase_size": 202.8, "purchase_base_unit": "fl oz"}
            ],
            "created_at": "2026-03-10T09:00:00+00:00",
            "updated_at": "2026-03-10T09:00:00+00:00"
        },
        {
            "recipe_id": "demo-turkey-club",
            "name": "Turkey Club Sandwich",
            "yield_portions": 1,
            "pricing_mode": "latest_price",
            "selling_price": 14.0,
            "pricing_goal_type": "food_cost_pct",
            "pricing_goal_value": 30,
            "target_food_cost_pct": 30,
            "ingredients": [
                {"product_name": "Sourdough Bread", "quantity": 2, "unit": "each", "purchase_unit": "case", "purchase_size": 24, "purchase_base_unit": "each"},
                {"product_name": "Turkey Breast", "quantity": 0.25, "unit": "lb", "purchase_unit": "lb", "purchase_size": 1, "purchase_base_unit": "lb"},
                {"product_name": "Smoked Bacon", "quantity": 0.1, "unit": "lb", "purchase_unit": "lb", "purchase_size": 1, "purchase_base_unit": "lb"},
                {"product_name": "Tomatoes", "quantity": 2, "unit": "each", "purchase_unit": "case", "purchase_size": 80, "purchase_base_unit": "each"},
                {"product_name": "Romaine Lettuce", "quantity": 2, "unit": "each", "purchase_unit": "case", "purchase_size": 24, "purchase_base_unit": "each"},
                {"product_name": "Mayonnaise", "quantity": 1, "unit": "fl oz", "purchase_unit": "case", "purchase_size": 384, "purchase_base_unit": "fl oz"}
            ],
            "created_at": "2026-03-12T10:30:00+00:00",
            "updated_at": "2026-03-12T10:30:00+00:00"
        },
        {
            "recipe_id": "demo-house-margarita",
            "name": "House Margarita",
            "yield_portions": 1,
            "pricing_mode": "latest_price",
            "selling_price": 12.0,
            "pricing_goal_type": "gross_margin_pct",
            "pricing_goal_value": 72,
            "target_food_cost_pct": 0,
            "ingredients": [
                {"product_name": "Tequila Blanco", "quantity": 2, "unit": "fl oz", "purchase_unit": "case", "purchase_size": 304.32, "purchase_base_unit": "fl oz"},
                {"product_name": "Simple Syrup", "quantity": 0.75, "unit": "fl oz", "purchase_unit": "case", "purchase_size": 202.8, "purchase_base_unit": "fl oz"},
                {"product_name": "Limes", "quantity": 1, "unit": "each", "purchase_unit": "case", "purchase_size": 175, "purchase_base_unit": "each"}
            ],
            "created_at": "2026-03-15T16:15:00+00:00",
            "updated_at": "2026-03-15T16:15:00+00:00"
        }
    ]


def build_recipe_product_defaults(recipes: list[dict[str, Any]] | None = None) -> dict[str, dict[str, Any]]:
    defaults: dict[str, dict[str, Any]] = {}
    for recipe in recipes or []:
        for ingredient in recipe.get("ingredients", []):
            product_name = str(ingredient.get("product_name") or "").strip()
            if not product_name or product_name in defaults:
                continue
            purchase_unit = normalize_recipe_unit_name(str(ingredient.get("purchase_unit") or "").strip())
            purchase_base_unit = resolve_recipe_purchase_base_unit(
                purchase_unit,
                str(ingredient.get("unit") or "").strip(),
                ingredient.get("purchase_base_unit")
            )
            purchase_size = normalize_request_value(ingredient.get("purchase_size", 0))
            try:
                normalized_purchase_size = float(purchase_size)
            except (TypeError, ValueError):
                normalized_purchase_size = 0.0
            defaults[product_name] = {
                "purchase_unit": purchase_unit,
                "purchase_base_unit": purchase_base_unit,
                "purchase_size": normalized_purchase_size if normalized_purchase_size > 0 else None,
                "preferred_usage_unit": normalize_recipe_unit_name(str(ingredient.get("unit") or purchase_base_unit or purchase_unit or "").strip())
            }
    return defaults


def with_demo_recipe_product_fallbacks(
    frame: pd.DataFrame,
    recipe_defaults: dict[str, dict[str, Any]]
) -> dict[str, dict[str, Any]]:
    defaults = dict(recipe_defaults or {})
    if frame.empty or "Product Name" not in frame.columns:
        return defaults
    unit_column = "Normalized Unit" if "Normalized Unit" in frame.columns else "Unit"
    grouped = frame.groupby("Product Name", sort=True)
    for product_name, group in grouped:
        normalized_product_name = str(product_name or "").strip()
        if not normalized_product_name or normalized_product_name in defaults:
            continue
        units = [
            normalize_recipe_unit_name(str(unit or "").strip())
            for unit in group.get(unit_column, group["Unit"]).fillna("").tolist()
            if str(unit or "").strip()
        ]
        purchase_unit = units[0] if units else ""
        purchase_category = get_recipe_unit_category(purchase_unit)
        if purchase_category == "package":
            purchase_base_unit = "each"
            purchase_size = 1.0
            preferred_usage_unit = "each"
        else:
            purchase_base_unit = resolve_recipe_purchase_base_unit(purchase_unit, purchase_unit, purchase_unit)
            purchase_size = 1.0
            preferred_usage_unit = purchase_base_unit or purchase_unit
        defaults[normalized_product_name] = {
            "purchase_unit": purchase_unit,
            "purchase_base_unit": purchase_base_unit,
            "purchase_size": purchase_size,
            "preferred_usage_unit": preferred_usage_unit
        }
    return defaults


def ensure_demo_recipe_session_dir() -> Path:
    DEMO_RECIPES_SESSION_DIR.mkdir(parents=True, exist_ok=True)
    return DEMO_RECIPES_SESSION_DIR


def get_demo_recipe_session_path(session_id: str) -> Path:
    normalized_session_id = re.sub(r"[^A-Za-z0-9._-]+", "_", str(session_id or "default").strip()).strip("._") or "default"
    return ensure_demo_recipe_session_dir() / f"{normalized_session_id}.json"


def load_demo_recipe_store(session_id: str) -> dict[str, Any]:
    session_path = get_demo_recipe_session_path(session_id)
    if not session_path.exists():
        store = {"recipes": build_demo_recipes()}
        session_path.write_text(json.dumps(store, ensure_ascii=False, indent=2), encoding="utf-8")
        return store
    try:
        payload = json.loads(session_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise RuntimeError("Demo recipe session is invalid.") from exc
    if not isinstance(payload, dict) or not isinstance(payload.get("recipes"), list):
        payload = {"recipes": build_demo_recipes()}
    return payload


def save_demo_recipe_store(session_id: str, store: dict[str, Any]) -> None:
    session_path = get_demo_recipe_session_path(session_id)
    session_path.write_text(
        json.dumps(store, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def dataframe_to_excel_stream(dataframe: pd.DataFrame, sheet_name: str) -> io.BytesIO:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name=sheet_name)
    output.seek(0)
    return output


def get_optional_unit_column(dataframe: pd.DataFrame) -> str | None:
    for column in OPTIONAL_UNIT_COLUMNS:
        if column in dataframe.columns:
            return column
    return None


def get_optional_quantity_column(dataframe: pd.DataFrame) -> str | None:
    for column in OPTIONAL_QUANTITY_COLUMNS:
        if column in dataframe.columns:
            return column
    return None


def get_optional_unit_price_column(dataframe: pd.DataFrame) -> str | None:
    for column in OPTIONAL_UNIT_PRICE_COLUMNS:
        if column in dataframe.columns:
            return column
    return None


def normalize_header_name(value: Any) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower())
    return " ".join(cleaned.split())


def score_header_match(normalized_header: str, field_name: str, aliases: list[str]) -> int:
    if not normalized_header:
        return 0

    header_compact = normalized_header.replace(" ", "")
    header_tokens = set(normalized_header.split())
    best_score = 0

    for alias in [field_name, *aliases]:
        normalized_alias = normalize_header_name(alias)
        alias_compact = normalized_alias.replace(" ", "")
        alias_tokens = set(normalized_alias.split())

        if normalized_header == normalized_alias:
            best_score = max(best_score, 160 if alias == field_name else 150)
            continue

        if header_compact == alias_compact:
            best_score = max(best_score, 144)
            continue

        if alias_tokens and alias_tokens.issubset(header_tokens):
            best_score = max(best_score, 118 + min(len(alias_tokens), 5))
            continue

        if header_tokens and header_tokens.issubset(alias_tokens):
            best_score = max(best_score, 106 + min(len(header_tokens), 4))
            continue

        overlap = len(header_tokens & alias_tokens)
        if not overlap:
            continue

        coverage = overlap / max(len(alias_tokens), 1)
        if coverage >= 0.75:
            best_score = max(best_score, 96 + overlap * 4)
        elif coverage >= 0.5:
            best_score = max(best_score, 78 + overlap * 3)

    return best_score


def detect_column_mappings(
    columns: list[str],
    *,
    required_fields: list[str] | None = None,
    field_synonyms: dict[str, list[str]] | None = None
) -> dict[str, Any]:
    required_fields = required_fields or REQUIRED_ANALYSIS_FIELDS
    field_synonyms = field_synonyms or REQUIRED_FIELD_SYNONYMS
    normalized_columns = {column: normalize_header_name(column) for column in columns}
    candidate_matches: list[tuple[int, str, str]] = []

    for field_name in required_fields:
        aliases = field_synonyms[field_name]
        for column, normalized_column in normalized_columns.items():
            score = score_header_match(normalized_column, field_name, aliases)
            if score > 0:
                candidate_matches.append((score, field_name, column))

    field_matches: dict[str, dict[str, Any]] = {
        field_name: {
            "field": field_name,
            "detected_column": None,
            "score": 0,
            "match_quality": "missing"
        }
        for field_name in required_fields
    }
    used_columns: set[str] = set()

    for score, field_name, column in sorted(
        candidate_matches,
        key=lambda item: (-item[0], required_fields.index(item[1]), str(item[2]).lower())
    ):
        if field_matches[field_name]["detected_column"] is not None or column in used_columns:
            continue
        field_matches[field_name]["detected_column"] = column
        field_matches[field_name]["score"] = score
        field_matches[field_name]["match_quality"] = "strong" if score >= 130 else "possible"
        used_columns.add(column)

    mapping = {
        field_name: field_matches[field_name]["detected_column"]
        for field_name in required_fields
    }
    missing_fields = [field_name for field_name, column in mapping.items() if not column]
    optional_columns = [column for column in columns if column not in used_columns]

    return {
        "mapping": mapping,
        "field_reviews": [field_matches[field_name] for field_name in required_fields],
        "matched_fields": len(required_fields) - len(missing_fields),
        "missing_fields": missing_fields,
        "optional_columns": optional_columns,
        "headers": columns
    }


def apply_column_mapping(
    dataframe: pd.DataFrame,
    mapping: dict[str, str | None],
    *,
    required_fields: list[str] | None = None
) -> pd.DataFrame:
    required_fields = required_fields or REQUIRED_ANALYSIS_FIELDS
    missing_fields = [field_name for field_name in required_fields if not mapping.get(field_name)]
    if missing_fields:
        raise ValueError(f"Missing required field mappings: {', '.join(missing_fields)}")

    missing_columns = [
        mapped_column
        for mapped_column in mapping.values()
        if mapped_column and mapped_column not in dataframe.columns
    ]
    if missing_columns:
        raise ValueError(
            f"Mapped columns were not found in the uploaded file: {', '.join(sorted(set(missing_columns)))}"
        )

    return pd.DataFrame({
        field_name: dataframe[mapped_column]
        for field_name, mapped_column in mapping.items()
        if mapped_column
    })


def get_quote_compare_selected_source_columns(mapping: dict[str, Any] | None) -> list[str]:
    selected_columns: list[str] = []
    seen_columns: set[str] = set()
    for field_name in [*QUOTE_COMPARE_REQUIRED_FIELDS, *QUOTE_COMPARE_OPTIONAL_FIELDS]:
        column_name = str((mapping or {}).get(field_name) or "").strip()
        if not column_name or column_name in seen_columns:
            continue
        seen_columns.add(column_name)
        selected_columns.append(column_name)
    return selected_columns


def build_dataframe_usecols_filter(selected_columns: list[str] | None) -> Any:
    normalized_columns = [str(column).strip() for column in (selected_columns or []) if str(column).strip()]
    if not normalized_columns:
        return None
    selected_column_names = set(normalized_columns)
    return lambda column_name: str(column_name) in selected_column_names


def read_uploaded_dataframe(
    file: UploadFile,
    *,
    selected_columns: list[str] | None = None,
    perf_label_prefix: str | None = None
) -> pd.DataFrame:
    filename = file.filename or ""
    extension = Path(filename).suffix.lower()
    read_started_at = perf_counter()
    read_substeps: dict[str, Any] = {
        "filename": filename,
        "extension": extension
    }
    usecols_filter = build_dataframe_usecols_filter(selected_columns)
    normalized_selected_columns = [str(column).strip() for column in (selected_columns or []) if str(column).strip()]
    read_substeps["selected_columns_count"] = len(normalized_selected_columns)
    read_substeps["selected_columns"] = normalized_selected_columns

    if not filename:
        raise ValueError("No file was uploaded.")

    try:
        size_check_started_at = perf_counter()
        size_bytes = ensure_upload_size_within_limit(file)
        read_substeps["size_check_ms"] = round((perf_counter() - size_check_started_at) * 1000, 1)
        probe_started_at = perf_counter()
        file.file.seek(0)
        first_byte = file.file.read(1)
        file.file.seek(0)
        read_substeps["file_probe_ms"] = round((perf_counter() - probe_started_at) * 1000, 1)
    except Exception as exc:
        if isinstance(exc, ValueError):
            raise
        logger.exception("Failed to reset uploaded file pointer: %s", filename)
        raise ValueError("The uploaded file could not be prepared for reading.") from exc

    if not first_byte or size_bytes <= 0:
        raise ValueError("The uploaded file is empty.")

    try:
        logger.info(
            "[upload debug] reading uploaded file: filename=%s extension=%s",
            filename,
            extension
        )
        if extension == ".csv":
            file.file.seek(0)
            csv_started_at = perf_counter()
            dataframe = pd.read_csv(file.file, dtype=object, usecols=usecols_filter)
            read_substeps["engine"] = "csv"
            read_substeps["parse_ms"] = round((perf_counter() - csv_started_at) * 1000, 1)
            log_perf("read_file.csv", read_started_at)
        elif extension == ".xlsx":
            try:
                file.file.seek(0)
                calamine_started_at = perf_counter()
                dataframe = pd.read_excel(file.file, engine="calamine", dtype=object, usecols=usecols_filter)
                read_substeps["engine"] = "calamine"
                read_substeps["parse_ms"] = round((perf_counter() - calamine_started_at) * 1000, 1)
                logger.info(
                    "[upload debug] parsed .xlsx with engine=calamine: filename=%s",
                    filename
                )
                log_perf("read_file.xlsx.calamine", read_started_at)
            except ImportError:
                try:
                    file.file.seek(0)
                    openpyxl_started_at = perf_counter()
                    dataframe = pd.read_excel(
                        file.file,
                        engine="openpyxl",
                        dtype=object,
                        usecols=usecols_filter,
                        engine_kwargs={"read_only": True, "data_only": True}
                    )
                    read_substeps["engine"] = "openpyxl"
                    read_substeps["parse_ms"] = round((perf_counter() - openpyxl_started_at) * 1000, 1)
                    logger.info(
                        "[upload debug] parsed .xlsx with fallback engine=openpyxl: filename=%s",
                        filename
                    )
                    log_perf("read_file.xlsx.openpyxl", read_started_at)
                except TypeError:
                    file.file.seek(0)
                    openpyxl_started_at = perf_counter()
                    dataframe = pd.read_excel(file.file, engine="openpyxl", dtype=object, usecols=usecols_filter)
                    read_substeps["engine"] = "openpyxl"
                    read_substeps["parse_ms"] = round((perf_counter() - openpyxl_started_at) * 1000, 1)
                    read_substeps["engine_kwargs_supported"] = False
                    logger.info(
                        "[upload debug] parsed .xlsx with fallback engine=openpyxl: filename=%s",
                        filename
                    )
                    log_perf("read_file.xlsx.openpyxl", read_started_at)
                except ImportError:
                    raise
            except Exception as calamine_exc:
                logger.warning(
                    "[upload debug] calamine parse failed for filename=%s extension=%s error=%s",
                    filename,
                    extension,
                    calamine_exc
                )
                try:
                    file.file.seek(0)
                    openpyxl_started_at = perf_counter()
                    dataframe = pd.read_excel(
                        file.file,
                        engine="openpyxl",
                        dtype=object,
                        usecols=usecols_filter,
                        engine_kwargs={"read_only": True, "data_only": True}
                    )
                    read_substeps["engine"] = "openpyxl"
                    read_substeps["parse_ms"] = round((perf_counter() - openpyxl_started_at) * 1000, 1)
                    logger.info(
                        "[upload debug] parsed .xlsx with fallback engine=openpyxl: filename=%s",
                        filename
                    )
                    log_perf("read_file.xlsx.openpyxl", read_started_at)
                except TypeError:
                    file.file.seek(0)
                    openpyxl_started_at = perf_counter()
                    dataframe = pd.read_excel(file.file, engine="openpyxl", dtype=object, usecols=usecols_filter)
                    read_substeps["engine"] = "openpyxl"
                    read_substeps["parse_ms"] = round((perf_counter() - openpyxl_started_at) * 1000, 1)
                    read_substeps["engine_kwargs_supported"] = False
                    logger.info(
                        "[upload debug] parsed .xlsx with fallback engine=openpyxl: filename=%s",
                        filename
                    )
                    log_perf("read_file.xlsx.openpyxl", read_started_at)
                except ImportError as openpyxl_exc:
                    logger.warning(
                        "[upload debug] openpyxl unavailable for filename=%s extension=%s error=%s",
                        filename,
                        extension,
                        openpyxl_exc
                    )
                    raise ValueError(
                        "This Excel workbook could not be parsed. Install python-calamine or openpyxl to enable .xlsx uploads."
                    ) from calamine_exc
                except Exception as openpyxl_exc:
                    logger.warning(
                        "[upload debug] openpyxl parse failed for filename=%s extension=%s error=%s",
                        filename,
                        extension,
                        openpyxl_exc
                    )
                    raise ValueError(
                        "This Excel workbook could not be parsed. Please re-save the workbook as a new .xlsx file and try again."
                    ) from openpyxl_exc
        elif extension == ".xls":
            file.file.seek(0)
            xls_started_at = perf_counter()
            dataframe = pd.read_excel(file.file, dtype=object, usecols=usecols_filter)
            read_substeps["engine"] = "xls"
            read_substeps["parse_ms"] = round((perf_counter() - xls_started_at) * 1000, 1)
            log_perf("read_file.xls", read_started_at)
        else:
            raise ValueError("Unsupported file type. Please upload a CSV or Excel file (.csv, .xlsx, or .xls).")
    except ImportError as exc:
        if extension == ".xlsx":
            raise ValueError("Excel support for .xlsx files is not installed. Add openpyxl to enable Excel uploads.") from exc
        if extension == ".xls":
            raise ValueError("Legacy Excel support for .xls files is not installed. Add xlrd to enable .xls uploads.") from exc
        raise
    except ValueError:
        raise
    except Exception as exc:
        logger.exception("Failed to parse uploaded file: %s", filename)
        if extension == ".xlsx":
            raise ValueError("The .xlsx file could not be read. Please upload a valid Excel workbook.") from exc
        if extension == ".xls":
            raise ValueError("The .xls file could not be read. Please upload a valid legacy Excel workbook.") from exc
        if extension == ".csv":
            raise ValueError("The .csv file could not be read. Please upload a valid CSV file.") from exc
        raise
    finally:
        try:
            file.file.seek(0)
        except Exception:
            logger.debug("Could not reset uploaded file pointer after parsing: %s", filename)

    validation_started_at = perf_counter()
    if dataframe is None or dataframe.columns.empty:
        raise ValueError("The uploaded file does not contain any readable columns.")
    read_substeps["validation_ms"] = round((perf_counter() - validation_started_at) * 1000, 1)
    read_substeps["row_count"] = int(len(dataframe.index))
    read_substeps["column_count"] = int(len(dataframe.columns))
    if perf_label_prefix:
        log_perf_details(f"{perf_label_prefix}.read_excel.substeps", **read_substeps)

    return dataframe


def ensure_quote_compare_upload_cache_dir(user_id: int | str | None = None) -> Path:
    return get_user_upload_cache_dir(get_storage_user_id(user_id))


def ensure_quote_compare_session_cache_dir(user_id: int | str | None = None) -> Path:
    return get_user_session_cache_dir(get_storage_user_id(user_id))


def get_quote_compare_session_path(session_id: str, user_id: int | str | None = None) -> Path:
    resolved_user_id = get_storage_user_id(user_id)
    ensure_quote_compare_session_cache_dir(resolved_user_id)
    return get_user_session_file_path(resolved_user_id, session_id)


def cache_quote_compare_upload(file: UploadFile, session_id: str, user_id: int | str | None = None) -> str:
    filename = file.filename or ""
    extension = Path(filename).suffix.lower()
    if not filename:
        raise ValueError("No file was uploaded.")
    resolved_user_id = get_storage_user_id(user_id)
    upload_cache_dir = ensure_quote_compare_upload_cache_dir(resolved_user_id)
    cache_path = upload_cache_dir / f"{session_id}{extension}"
    try:
        file.file.seek(0)
        cache_path.write_bytes(file.file.read())
        file.file.seek(0)
    except Exception as exc:
        logger.exception("Failed to cache compare prices upload: %s", filename)
        raise ValueError("The uploaded supplier file could not be cached for review.") from exc
    return str(cache_path)


def read_cached_quote_compare_upload(
    cache_path: str | None,
    filename: str = "",
    user_id: int | str | None = None,
    selected_columns: list[str] | None = None,
    perf_label_prefix: str | None = None
) -> pd.DataFrame:
    normalized_path = str(cache_path or "").strip()
    if not normalized_path:
        raise ValueError("The uploaded supplier file is no longer available. Please upload it again.")
    source_path = Path(normalized_path)
    expected_cache_dir = ensure_quote_compare_upload_cache_dir(get_storage_user_id(user_id)).resolve()
    try:
        resolved_source_path = source_path.resolve()
    except FileNotFoundError:
        resolved_source_path = source_path
    if expected_cache_dir not in resolved_source_path.parents:
        raise ValueError("The uploaded supplier file is no longer available. Please upload it again.")
    if not source_path.exists() or not source_path.is_file():
        raise ValueError("The uploaded supplier file is no longer available. Please upload it again.")
    if source_path.stat().st_size > MAX_UPLOAD_SIZE_BYTES:
        raise ValueError(get_upload_size_limit_message())

    class CachedUploadFile:
        def __init__(self, path: Path, original_name: str):
            self.filename = original_name or path.name
            self.file = path.open("rb")

    cached_upload = CachedUploadFile(source_path, filename or source_path.name)
    try:
        return read_uploaded_dataframe(
            cached_upload,
            selected_columns=selected_columns,
            perf_label_prefix=perf_label_prefix
        )
    finally:
        try:
            cached_upload.file.close()
        except Exception:
            logger.debug("Could not close cached compare prices upload: %s", source_path)


def build_mapping_review_payload(
    dataframe: pd.DataFrame,
    *,
    filename: str,
    required_fields: list[str] | None = None,
    field_synonyms: dict[str, list[str]] | None = None,
    message: str | None = None,
    review_message: str | None = None
) -> dict[str, Any]:
    required_fields = required_fields or REQUIRED_ANALYSIS_FIELDS
    field_synonyms = field_synonyms or REQUIRED_FIELD_SYNONYMS
    columns = [str(column) for column in dataframe.columns]
    logger.info(
        "[upload debug] extracted dataframe columns for %s: %s",
        filename,
        columns
    )
    detection = detect_column_mappings(
        columns,
        required_fields=required_fields,
        field_synonyms=field_synonyms
    )
    logger.info(
        "[upload debug] mapping review payload for %s: matched=%s missing=%s headers=%s",
        filename,
        detection.get("matched_fields"),
        detection.get("missing_fields"),
        detection.get("headers")
    )
    return {
        "filename": filename,
        "required_fields": required_fields,
        "message": message or "We detected likely matches from your file headers.",
        "review_message": review_message or "Review and confirm the fields below before analysis.",
        **detection
    }


def ensure_codes_file() -> None:
    if CODES_PATH.exists():
        return

    default_codes = {
        "PPA-AB12-CD3": {"active": True, "session_id": None}
    }
    CODES_PATH.write_text(json.dumps(default_codes, indent=2), encoding="utf-8")


def load_codes() -> dict:
    ensure_codes_file()
    try:
        codes = json.loads(CODES_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Invalid codes file: {CODES_PATH}") from exc
    changed = False
    for code, entry in codes.items():
        if not isinstance(entry, dict):
            codes[code] = {"active": bool(entry), "session_id": None}
            changed = True
            continue
        if "active" not in entry:
            entry["active"] = False
            changed = True
        if "session_id" not in entry:
            entry["session_id"] = None
            changed = True
    if changed:
        save_codes(codes)
    return codes


def save_codes(codes: dict) -> None:
    CODES_PATH.write_text(json.dumps(codes, indent=2), encoding="utf-8")


def normalize_access_code(code: str) -> str:
    return (code or "").strip().upper()


def generate_access_code() -> str:
    prefix = "PPA"
    middle = "".join(random.choice(READABLE_CODE_ALPHABET) for _ in range(4))
    suffix = "".join(random.choice(READABLE_CODE_ALPHABET) for _ in range(3))
    return f"{prefix}-{middle}-{suffix}"


def generate_license_code() -> str:
    segments = [
        "".join(random.choice(READABLE_CODE_ALPHABET) for _ in range(4)),
        "".join(random.choice(READABLE_CODE_ALPHABET) for _ in range(4)),
        "".join(random.choice(READABLE_CODE_ALPHABET) for _ in range(4))
    ]
    return "-".join(segments)


def create_unique_access_code() -> str:
    codes = load_codes()
    while True:
        candidate = generate_access_code()
        if candidate not in codes:
            codes[candidate] = {"active": True, "session_id": None}
            save_codes(codes)
            return candidate


def create_session_id() -> str:
    return str(uuid.uuid4())


def normalize_session_id(session_id: str | None) -> str | None:
    normalized_session_id = (session_id or "").strip() or None
    if normalized_session_id and normalized_session_id.lower() in {"null", "none"}:
        return None
    return normalized_session_id


def validate_access_code_session(code: str, session_id: str | None) -> dict:
    normalized_code = normalize_access_code(code)
    normalized_session_id = normalize_session_id(session_id)
    if not normalized_code:
        response = {"success": False, "message": "Invalid code"}
        logger.info(
            "Access code validation rejected | stored session_id=%s | incoming session_id=%s | response=%s",
            None,
            normalized_session_id,
            response
        )
        return response

    codes = load_codes()
    code_entry = codes.get(normalized_code)
    if not code_entry or not code_entry.get("active"):
        response = {"success": False, "message": "Invalid code"}
        logger.info(
            "Access code validation rejected for %s | stored session_id=%s | incoming session_id=%s | response=%s",
            normalized_code,
            None if not code_entry else code_entry.get("session_id"),
            normalized_session_id,
            response
        )
        return response

    current_session_id = normalize_session_id(code_entry.get("session_id"))
    if code_entry.get("session_id") != current_session_id:
        code_entry["session_id"] = current_session_id
        save_codes(codes)

    logger.info(
        "Access code validation start for %s | stored session_id=%s | incoming session_id=%s",
        normalized_code,
        current_session_id,
        normalized_session_id
    )

    if current_session_id is None:
        new_session_id = create_session_id()
        code_entry["session_id"] = new_session_id
        save_codes(codes)
        response = {
            "success": True,
            "code": normalized_code,
            "session_id": new_session_id
        }
        print("FLOW: NEW SESSION CREATED")
        logger.info(
            "Access code validation success for %s | stored session_id=%s | incoming session_id=%s | response=%s",
            normalized_code,
            current_session_id,
            normalized_session_id,
            response
        )
        return response

    if normalized_session_id == current_session_id:
        response = {
            "success": True,
            "code": normalized_code,
            "session_id": current_session_id
        }
        print("FLOW: EXISTING SESSION MATCH")
        logger.info(
            "Access code validation success for %s | stored session_id=%s | incoming session_id=%s | response=%s",
            normalized_code,
            current_session_id,
            normalized_session_id,
            response
        )
        return response

    response = {
        "success": False,
        "message": "This access code is already in use on another session."
    }
    print("FLOW: REJECT")
    logger.info(
        "Access code validation reject for %s | stored session_id=%s | incoming session_id=%s | response=%s",
        normalized_code,
        current_session_id,
        normalized_session_id,
        response
    )
    return response


def logout_access_code_session(code: str, session_id: str | None) -> bool:
    normalized_code = normalize_access_code(code)
    normalized_session_id = (session_id or "").strip() or None
    if not normalized_code or not normalized_session_id:
        return False

    codes = load_codes()
    code_entry = codes.get(normalized_code)
    if not code_entry or code_entry.get("session_id") != normalized_session_id:
        return False

    code_entry["session_id"] = None
    save_codes(codes)
    return True


class AccessCodePayload(BaseModel):
    code: str


class AccessSessionPayload(BaseModel):
    code: str
    session_id: str | None = None


class AskDataPayload(BaseModel):
    question: str
    rows: list[dict[str, Any]] | None = None


class GuideAskPayload(BaseModel):
    question: str


class UploadMappingPayload(BaseModel):
    mappings: dict[str, str | None]


class RecipeIngredientPayload(BaseModel):
    product_name: str
    quantity: float
    unit: str
    purchase_unit: str | None = None
    purchase_size: float | None = None
    purchase_base_unit: str | None = None


class RecipePayload(BaseModel):
    recipe_id: str | None = None
    name: str
    yield_portions: float
    pricing_mode: str
    ingredients: list[RecipeIngredientPayload]
    selling_price: float | None = None
    pricing_goal_type: str | None = None
    pricing_goal_value: float | None = None
    target_food_cost_pct: float | None = None
    total_recipe_cost: float | None = None
    cost_per_portion: float | None = None
    gross_profit: float | None = None
    gross_margin_pct: float | None = None
    food_cost_pct: float | None = None
    suggested_selling_price: float | None = None


class RecipeDeletePayload(BaseModel):
    recipe_id: str


class QuoteBidPayload(BaseModel):
    supplier_name: str
    product_name: str
    unit: str
    quantity: float
    unit_price: float | None = None
    total_price: float | None = None
    quote_date: str | None = None
    currency: str
    delivery_time: str
    payment_term: str
    valid_until: str | None = None
    notes: str | None = None


class QuoteComparisonPayload(BaseModel):
    comparison_id: str | None = None
    upload_id: str | None = None
    name: str
    sourcing_need: str | None = None
    bids: list[QuoteBidPayload]
    weighting: dict[str, float] | None = None
    source_type: str | None = None
    mode: str | None = None


class QuoteComparisonDeletePayload(BaseModel):
    comparison_id: str


def format_currency(value: float) -> str:
    return f"${value:,.2f}"


def format_percent(value: float) -> str:
    return f"{value:.1f}%"


def format_period_label(period_value: pd.Period | None) -> str:
    if period_value is None:
        return "Current visible period"
    return period_value.strftime("%B %Y")


def coalesce_number(primary: Any, fallback: Any = 0) -> float:
    if pd.notna(primary):
        return float(primary)
    if pd.notna(fallback):
        return float(fallback)
    return 0.0


def parse_bool_flag(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "on"}:
        return True
    if text in {"false", "0", "no", "off", ""}:
        return False
    return False


def safe_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in {"true", "false", ""}:
        return default
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return default


def normalize_value(value: Any) -> Any:
    if value is None:
        return None
    if value is pd.NA:
        return None
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass

    if isinstance(value, bool):
        return value

    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            normalized_item = value.item()
        except (TypeError, ValueError):
            normalized_item = value
        if normalized_item is not value:
            return normalize_value(normalized_item)

    if isinstance(value, pd.Timestamp):
        return value.isoformat()

    if isinstance(value, str):
        stripped = value.strip()
        lowered = stripped.lower()
        if lowered in {"true", "false"}:
            return stripped
        if re.fullmatch(r"[+-]?\d+", stripped):
            return safe_int(stripped)
        return stripped

    return value


def normalize_request_value(value: Any) -> Any:
    return normalize_value(value)


def parse_localized_float(value: Any) -> float:
    normalized_value = normalize_value(value)
    if normalized_value is None or normalized_value == "":
        raise ValueError("Empty numeric value")
    if isinstance(normalized_value, bool):
        raise ValueError("Boolean value is not numeric")
    if isinstance(normalized_value, (int, float)):
        return float(normalized_value)

    text = str(normalized_value).strip()
    if not text:
        raise ValueError("Empty numeric text")

    text = text.replace("\u00a0", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")

    text = re.sub(r"[^0-9.\-+]", "", text)
    if not text or text in {"-", "+", ".", "-.", "+."}:
        raise ValueError(f"Invalid numeric text: {normalized_value!r}")

    return float(text)


def make_json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {
            str(key): make_json_safe(item)
            for key, item in value.items()
        }

    if isinstance(value, list):
        return [make_json_safe(item) for item in value]

    if isinstance(value, tuple):
        return [make_json_safe(item) for item in value]

    if isinstance(value, (datetime, date)):
        return value.isoformat()

    if isinstance(value, pd.Timestamp):
        return value.isoformat()

    if hasattr(value, "item") and not isinstance(value, (str, bytes)):
        try:
            normalized_item = value.item()
        except (TypeError, ValueError):
            normalized_item = value
        if normalized_item is not value:
            return make_json_safe(normalized_item)

    return normalize_value(value)


def coerce_numeric_value(value: Any, *, field_name: str, context: str) -> float:
    normalized_value = normalize_value(value)
    if normalized_value is None or normalized_value == "":
        return 0.0
    if isinstance(normalized_value, bool):
        logger.info(
            "[compare prices upload] preserving boolean-like value for %s (%s): %r",
            field_name,
            context,
            normalized_value
        )
        return 0.0
    if isinstance(normalized_value, str) and normalized_value.lower() in {"true", "false"}:
        logger.info(
            "[compare prices upload] preserving boolean-like string for %s (%s): %r",
            field_name,
            context,
            normalized_value
        )
        return 0.0

    try:
        return parse_localized_float(normalized_value)
    except (TypeError, ValueError):
        logger.warning(
            "[compare prices upload] failed numeric coercion for %s (%s): type=%s value=%r",
            field_name,
            context,
            type(normalized_value).__name__,
            normalized_value
        )
        return 0.0


def normalize_text_value(value: Any) -> str:
    normalized_value = normalize_value(value)
    if normalized_value is None:
        return ""
    if isinstance(normalized_value, bool):
        return str(normalized_value)
    return str(normalized_value).strip()


def normalize_comparison_product_name(value: Any) -> str:
    normalized_value = normalize_text_value(value)
    if not normalized_value:
        return ""
    return re.sub(r"\s+", " ", normalized_value).strip().lower()


def normalize_comparison_unit(value: Any) -> str:
    return normalize_text_value(value).lower()


def ensure_recipes_file(user_id: int | str | None = None) -> None:
    recipes_path = get_user_recipes_path(get_storage_user_id(user_id))
    if recipes_path.exists():
        return
    recipes_path.write_text(json.dumps({"recipes": []}, indent=2), encoding="utf-8")


def load_recipes_store(user_id: int | str | None = None) -> dict[str, Any]:
    resolved_user_id = get_storage_user_id(user_id)
    recipes_path = get_user_recipes_path(resolved_user_id)
    ensure_recipes_file(resolved_user_id)
    cache_signature = (
        str(recipes_path),
        recipes_path.stat().st_mtime_ns,
        recipes_path.stat().st_size
    )
    if RECIPES_STORE_CACHE["signature"] == cache_signature and isinstance(RECIPES_STORE_CACHE["store"], dict):
        return RECIPES_STORE_CACHE["store"]
    try:
        store = json.loads(recipes_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Invalid recipes file: {recipes_path}") from exc

    recipes = store.get("recipes")
    if not isinstance(recipes, list):
        store["recipes"] = []
        save_recipes_store(store, resolved_user_id)
        return RECIPES_STORE_CACHE["store"]
    RECIPES_STORE_CACHE["signature"] = cache_signature
    RECIPES_STORE_CACHE["store"] = store
    return store


def save_recipes_store(store: dict[str, Any], user_id: int | str | None = None) -> None:
    recipes_path = get_user_recipes_path(get_storage_user_id(user_id))
    recipes_path.write_text(json.dumps(store, indent=2), encoding="utf-8")
    RECIPES_STORE_CACHE["signature"] = (
        str(recipes_path),
        recipes_path.stat().st_mtime_ns,
        recipes_path.stat().st_size
    )
    RECIPES_STORE_CACHE["store"] = store
    RECIPES_BOOTSTRAP_RESPONSE_CACHE["signature"] = None
    RECIPES_BOOTSTRAP_RESPONSE_CACHE["response_json"] = None


def ensure_quote_comparisons_file(user_id: int | str | None = None) -> None:
    comparisons_path = get_user_quote_comparisons_path(get_storage_user_id(user_id))
    if comparisons_path.exists():
        return
    comparisons_path.write_text(
        json.dumps({"comparisons": [], "active_sessions": {}}, indent=2),
        encoding="utf-8"
    )


def ensure_analysis_history_file(user_id: int | str | None = None) -> None:
    analysis_history_path = get_user_analysis_history_path(get_storage_user_id(user_id))
    if analysis_history_path.exists():
        return
    analysis_history_path.write_text(
        json.dumps({"uploads": [], "rows": []}, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8"
    )


def reset_workspace_data_store(user_id: int | str | None = None) -> None:
    resolved_user_id = get_storage_user_id(user_id)
    user_storage_root = get_user_storage_root(resolved_user_id)
    active_session_path = get_user_active_quote_compare_session_path(resolved_user_id)
    latest_results_path = get_current_latest_results_path(resolved_user_id)
    analysis_history_path = get_user_analysis_history_path(resolved_user_id)
    reset_details = {
        "user_id": resolved_user_id,
        "cleared_active_session": active_session_path.exists(),
        "cleared_analysis_file": latest_results_path.exists(),
        "cleared_analysis_history": analysis_history_path.exists()
    }

    if user_storage_root.exists():
        shutil.rmtree(user_storage_root, ignore_errors=True)
    user_storage_root.mkdir(parents=True, exist_ok=True)

    QUOTE_COMPARE_STORE_CACHE["signature"] = None
    QUOTE_COMPARE_STORE_CACHE["store"] = None
    ANALYSIS_HISTORY_CACHE["signature"] = None
    ANALYSIS_HISTORY_CACHE["store"] = None
    LATEST_ANALYSIS_CACHE["signature"] = None
    LATEST_ANALYSIS_CACHE["context"] = None
    CURRENT_UPLOAD_ANALYSIS_FRAME_CACHE["signature"] = None
    CURRENT_UPLOAD_ANALYSIS_FRAME_CACHE["frame"] = None
    RECIPES_STORE_CACHE["signature"] = None
    RECIPES_STORE_CACHE["store"] = None
    RECIPE_ANALYSIS_CACHE["signature"] = None
    RECIPE_ANALYSIS_CACHE["frame"] = None
    RECIPE_ANALYSIS_CACHE["product_catalog"] = None
    RECIPE_ANALYSIS_CACHE["pricing_lookup"] = None
    LATEST_ANALYSIS_UPLOAD_META_CACHE["signature"] = None
    LATEST_ANALYSIS_UPLOAD_META_CACHE["meta"] = None
    RECIPES_BOOTSTRAP_RESPONSE_CACHE["signature"] = None
    RECIPES_BOOTSTRAP_RESPONSE_CACHE["response_json"] = None
    log_perf_details("reset.cleared_active_session", cleared=reset_details["cleared_active_session"], user_id=resolved_user_id)
    log_perf_details("reset.cleared_analysis_file", cleared=reset_details["cleared_analysis_file"], user_id=resolved_user_id)


def load_quote_comparisons_store(user_id: int | str | None = None) -> dict[str, Any]:
    resolved_user_id = get_storage_user_id(user_id)
    comparisons_path = get_user_quote_comparisons_path(resolved_user_id)
    ensure_quote_comparisons_file(resolved_user_id)
    cache_signature = (
        str(comparisons_path),
        comparisons_path.stat().st_mtime_ns,
        comparisons_path.stat().st_size
    )
    if (
        QUOTE_COMPARE_STORE_CACHE["signature"] == cache_signature
        and isinstance(QUOTE_COMPARE_STORE_CACHE["store"], dict)
    ):
        return QUOTE_COMPARE_STORE_CACHE["store"]
    try:
        store = json.loads(comparisons_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Invalid quote comparisons file: {comparisons_path}") from exc

    changed = False
    comparisons = store.get("comparisons")
    if not isinstance(comparisons, list):
        store["comparisons"] = []
        changed = True
    active_sessions = store.get("active_sessions")
    if not isinstance(active_sessions, dict):
        store["active_sessions"] = {}
        changed = True
    if changed:
        save_quote_comparisons_store(store, resolved_user_id)
        return QUOTE_COMPARE_STORE_CACHE["store"]
    QUOTE_COMPARE_STORE_CACHE["signature"] = cache_signature
    QUOTE_COMPARE_STORE_CACHE["store"] = store
    return store


def save_quote_comparisons_store(store: dict[str, Any], user_id: int | str | None = None) -> None:
    comparisons_path = get_user_quote_comparisons_path(get_storage_user_id(user_id))
    safe_store = make_json_safe(store)
    comparisons_path.write_text(
        json.dumps(safe_store, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8"
    )
    QUOTE_COMPARE_STORE_CACHE["signature"] = (
        str(comparisons_path),
        comparisons_path.stat().st_mtime_ns,
        comparisons_path.stat().st_size
    )
    QUOTE_COMPARE_STORE_CACHE["store"] = safe_store
    LATEST_ANALYSIS_UPLOAD_META_CACHE["signature"] = None
    LATEST_ANALYSIS_UPLOAD_META_CACHE["meta"] = None


def load_analysis_history_store(user_id: int | str | None = None) -> dict[str, Any]:
    resolved_user_id = get_storage_user_id(user_id)
    analysis_history_path = get_user_analysis_history_path(resolved_user_id)
    ensure_analysis_history_file(resolved_user_id)
    cache_signature = (
        str(analysis_history_path),
        analysis_history_path.stat().st_mtime_ns,
        analysis_history_path.stat().st_size
    )
    if (
        ANALYSIS_HISTORY_CACHE["signature"] == cache_signature
        and isinstance(ANALYSIS_HISTORY_CACHE["store"], dict)
    ):
        return ANALYSIS_HISTORY_CACHE["store"]

    try:
        store = json.loads(analysis_history_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Invalid analysis history file: {analysis_history_path}") from exc

    changed = False
    if not isinstance(store.get("uploads"), list):
        store["uploads"] = []
        changed = True
    if not isinstance(store.get("rows"), list):
        store["rows"] = []
        changed = True
    if changed:
        save_analysis_history_store(store, resolved_user_id)
        return ANALYSIS_HISTORY_CACHE["store"]

    ANALYSIS_HISTORY_CACHE["signature"] = cache_signature
    ANALYSIS_HISTORY_CACHE["store"] = store
    return store


def save_analysis_history_store(store: dict[str, Any], user_id: int | str | None = None) -> None:
    analysis_history_path = get_user_analysis_history_path(get_storage_user_id(user_id))
    safe_store = make_json_safe(store)
    analysis_history_path.write_text(
        json.dumps(safe_store, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8"
    )
    ANALYSIS_HISTORY_CACHE["signature"] = (
        str(analysis_history_path),
        analysis_history_path.stat().st_mtime_ns,
        analysis_history_path.stat().st_size
    )
    ANALYSIS_HISTORY_CACHE["store"] = safe_store
    LATEST_ANALYSIS_UPLOAD_META_CACHE["signature"] = None
    LATEST_ANALYSIS_UPLOAD_META_CACHE["meta"] = None


def seed_analysis_history_from_latest_results() -> None:
    latest_results_path = get_current_latest_results_path()
    if not latest_results_path.exists():
        return

    store = load_analysis_history_store()
    if store.get("rows"):
        return

    try:
        legacy_df = load_current_upload_analysis_frame()
    except Exception:
        logger.exception("Failed to seed analysis history from latest results")
        return

    if legacy_df.empty:
        return

    upload_id = f"legacy-{latest_results_path.stat().st_mtime_ns}"
    if "upload_id" not in legacy_df.columns:
        legacy_df["upload_id"] = upload_id
    else:
        legacy_df["upload_id"] = legacy_df["upload_id"].fillna("").astype(str).replace("", upload_id)
    if "row_id" not in legacy_df.columns:
        legacy_df["row_id"] = [
            str(uuid.uuid5(uuid.NAMESPACE_URL, f"{upload_id}|legacy|{index}"))
            for index in range(1, len(legacy_df.index) + 1)
        ]
    legacy_df = assign_analysis_row_ids(legacy_df, upload_id=upload_id)
    save_current_upload_analysis_frame(legacy_df)
    append_analysis_history_rows(
        legacy_df,
        upload_id=upload_id,
        source_name="Latest imported analysis",
        source_type="upload",
        comparison_name="Latest imported analysis"
    )


def normalize_analysis_scope(scope: str | None) -> str:
    normalized_scope = str(scope or "current_upload").strip().lower()
    return "demo" if normalized_scope == "demo" else "current_upload"


def build_analysis_scope_options() -> list[dict[str, str]]:
    return [dict(option) for option in ANALYSIS_SCOPE_OPTIONS]


def get_analysis_scope_label(scope: str | None) -> str:
    return "Demo Data" if normalize_analysis_scope(scope) == "demo" else "Current File"


def serialize_dataframe_records(dataframe: pd.DataFrame) -> list[dict[str, Any]]:
    dataframe = dataframe.copy()
    for column in dataframe.columns:
        if pd.api.types.is_datetime64_any_dtype(dataframe[column]):
            dataframe[column] = dataframe[column].astype(str)

    serializable = dataframe.astype(object).where(pd.notna(dataframe), None)
    records: list[dict[str, Any]] = []
    for row in serializable.to_dict(orient="records"):
        records.append({
            str(key): make_json_safe(value)
            for key, value in row.items()
        })
    return records


def build_analysis_dedupe_key_series(frame: pd.DataFrame) -> pd.Series:
    if frame.empty:
        return pd.Series(dtype="string")

    key_parts: list[pd.Series] = []
    date_series = pd.to_datetime(frame.get("Date"), errors="coerce")
    key_parts.append(date_series.dt.strftime("%Y-%m-%d").fillna("").astype("string"))

    for column in ANALYSIS_DEDUPE_TEXT_COLUMNS:
        if column in frame.columns:
            key_parts.append(frame[column].fillna("").astype("string").str.strip().str.lower())
        else:
            key_parts.append(pd.Series([""] * len(frame.index), index=frame.index, dtype="string"))

    for column, decimals in ANALYSIS_DEDUPE_NUMBER_COLUMNS.items():
        if column in frame.columns:
            numeric_series = pd.to_numeric(frame[column], errors="coerce").round(decimals)
            formatted_numeric_series = numeric_series.map(
                lambda value: "" if pd.isna(value) else f"{float(value):.{decimals}f}"
            ).astype("string")
            key_parts.append(formatted_numeric_series)
        else:
            key_parts.append(pd.Series([""] * len(frame.index), index=frame.index, dtype="string"))

    combined_keys = key_parts[0]
    for part in key_parts[1:]:
        combined_keys = combined_keys.str.cat(part, sep="|")
    return combined_keys.astype("string")


def deduplicate_analysis_frame(
    frame: pd.DataFrame,
    *,
    keep_dedupe_key: bool = False,
    copy_frame: bool = True,
    is_deduplicated: bool = False
) -> pd.DataFrame:
    if frame.empty:
        return frame.copy()

    dedupe_started_at = perf_counter()
    dedupe_substeps: dict[str, Any] = {
        "input_rows": int(len(frame.index)),
        "input_columns": int(len(frame.columns)),
        "copy_frame": copy_frame,
        "keep_dedupe_key": keep_dedupe_key,
        "is_deduplicated": is_deduplicated
    }
    skip_reason = ""
    if is_deduplicated:
        skip_reason = "caller_marked_deduplicated"
    elif "_analysis_dedupe_key" in frame.columns:
        dedupe_key_series = frame["_analysis_dedupe_key"].fillna("").astype("string")
        if dedupe_key_series.astype(bool).all() and dedupe_key_series.nunique(dropna=False) == len(frame.index):
            skip_reason = "existing_unique_dedupe_key"

    if skip_reason:
        deduped_frame = frame.copy() if copy_frame else frame
        if keep_dedupe_key and "_analysis_dedupe_key" not in deduped_frame.columns:
            build_key_started_at = perf_counter()
            deduped_frame["_analysis_dedupe_key"] = build_analysis_dedupe_key_series(deduped_frame)
            dedupe_substeps["build_key_ms"] = round((perf_counter() - build_key_started_at) * 1000, 1)
        elif not keep_dedupe_key and "_analysis_dedupe_key" in deduped_frame.columns:
            drop_key_started_at = perf_counter()
            deduped_frame = deduped_frame.drop(columns=["_analysis_dedupe_key"], errors="ignore")
            dedupe_substeps["drop_key_ms"] = round((perf_counter() - drop_key_started_at) * 1000, 1)
        dedupe_substeps["deduplicate_skipped_reason"] = skip_reason
        dedupe_substeps["output_rows"] = int(len(deduped_frame.index))
        dedupe_substeps["removed_rows"] = 0
        dedupe_substeps["total_ms"] = round((perf_counter() - dedupe_started_at) * 1000, 1)
        log_perf_details("quote_compare.deduplicate_results.substeps", **dedupe_substeps)
        return deduped_frame

    recipe_bootstrap_metrics = RECIPE_BOOTSTRAP_METRICS_CONTEXT.get()
    if isinstance(recipe_bootstrap_metrics, dict):
        recipe_bootstrap_metrics["deduplicate_call_count"] = int(recipe_bootstrap_metrics.get("deduplicate_call_count") or 0) + 1

    deduped_frame = frame.copy() if copy_frame else frame
    normalize_started_at = perf_counter()
    if "Date" in deduped_frame.columns:
        deduped_frame["Date"] = pd.to_datetime(deduped_frame["Date"], errors="coerce")
    if "Valid Until" in deduped_frame.columns:
        deduped_frame["Valid Until"] = pd.to_datetime(deduped_frame["Valid Until"], errors="coerce")
    for column in ["Quantity", "Unit Price", "Total Amount", "Average Price", "Overpay", "Savings Opportunity", "Overpay Pct"]:
        if column in deduped_frame.columns:
            deduped_frame[column] = pd.to_numeric(deduped_frame[column], errors="coerce")
    dedupe_substeps["normalize_columns_ms"] = round((perf_counter() - normalize_started_at) * 1000, 1)

    key_started_at = perf_counter()
    deduped_frame["_analysis_dedupe_key"] = build_analysis_dedupe_key_series(deduped_frame)
    dedupe_substeps["build_key_ms"] = round((perf_counter() - key_started_at) * 1000, 1)

    drop_duplicates_started_at = perf_counter()
    deduped_frame = deduped_frame.drop_duplicates(subset=["_analysis_dedupe_key"], keep="last")
    dedupe_substeps["drop_duplicates_ms"] = round((perf_counter() - drop_duplicates_started_at) * 1000, 1)
    if not keep_dedupe_key:
        drop_key_started_at = perf_counter()
        deduped_frame = deduped_frame.drop(columns=["_analysis_dedupe_key"], errors="ignore")
        dedupe_substeps["drop_key_ms"] = round((perf_counter() - drop_key_started_at) * 1000, 1)

    sort_columns = [column for column in ["Date", "Product Name", "Supplier"] if column in deduped_frame.columns]
    if sort_columns:
        sort_started_at = perf_counter()
        ascending = [False if column == "Date" else True for column in sort_columns]
        deduped_frame = deduped_frame.sort_values(sort_columns, ascending=ascending, na_position="last")
        dedupe_substeps["sort_ms"] = round((perf_counter() - sort_started_at) * 1000, 1)

    reset_index_started_at = perf_counter()
    deduped_frame = deduped_frame.reset_index(drop=True)
    dedupe_substeps["reset_index_ms"] = round((perf_counter() - reset_index_started_at) * 1000, 1)
    dedupe_substeps["output_rows"] = int(len(deduped_frame.index))
    dedupe_substeps["removed_rows"] = int(len(frame.index) - len(deduped_frame.index))
    dedupe_substeps["total_ms"] = round((perf_counter() - dedupe_started_at) * 1000, 1)
    log_perf_details("quote_compare.deduplicate_results.substeps", **dedupe_substeps)
    return deduped_frame


def assign_analysis_row_ids(frame: pd.DataFrame, *, upload_id: str, copy_frame: bool = True) -> pd.DataFrame:
    if frame.empty:
        return frame.copy()

    identified_frame = frame.copy() if copy_frame else frame
    dedupe_keys = identified_frame["_analysis_dedupe_key"] if "_analysis_dedupe_key" in identified_frame.columns else build_analysis_dedupe_key_series(identified_frame)
    identified_frame["row_id"] = [
        str(uuid.uuid5(uuid.NAMESPACE_URL, f"{str(upload_id).strip()}|{dedupe_key}"))
        for dedupe_key in dedupe_keys.tolist()
    ]
    identified_frame["upload_id"] = str(upload_id).strip()
    return identified_frame.drop(columns=["_analysis_dedupe_key"], errors="ignore")


def load_current_upload_analysis_frame() -> pd.DataFrame:
    latest_results_path = get_current_latest_results_path()
    if not latest_results_path.exists():
        return pd.DataFrame()

    cache_signature = (
        str(latest_results_path),
        latest_results_path.stat().st_mtime_ns,
        latest_results_path.stat().st_size
    )
    if (
        CURRENT_UPLOAD_ANALYSIS_FRAME_CACHE["signature"] == cache_signature
        and isinstance(CURRENT_UPLOAD_ANALYSIS_FRAME_CACHE["frame"], pd.DataFrame)
    ):
        return CURRENT_UPLOAD_ANALYSIS_FRAME_CACHE["frame"].copy()

    frame = pd.read_csv(latest_results_path)
    deduplicated_frame = deduplicate_analysis_frame(frame, is_deduplicated=True)
    CURRENT_UPLOAD_ANALYSIS_FRAME_CACHE["signature"] = cache_signature
    CURRENT_UPLOAD_ANALYSIS_FRAME_CACHE["frame"] = deduplicated_frame
    return deduplicated_frame.copy()


def save_current_upload_analysis_frame(frame: pd.DataFrame) -> None:
    latest_results_path = get_current_latest_results_path()
    frame.to_csv(latest_results_path, index=False)
    LATEST_ANALYSIS_CACHE["signature"] = None
    LATEST_ANALYSIS_CACHE["context"] = None
    CURRENT_UPLOAD_ANALYSIS_FRAME_CACHE["signature"] = None
    CURRENT_UPLOAD_ANALYSIS_FRAME_CACHE["frame"] = None
    RECIPE_ANALYSIS_CACHE["signature"] = None
    RECIPE_ANALYSIS_CACHE["frame"] = None
    RECIPE_ANALYSIS_CACHE["product_catalog"] = None
    RECIPE_ANALYSIS_CACHE["pricing_lookup"] = None
    LATEST_ANALYSIS_UPLOAD_META_CACHE["signature"] = None
    LATEST_ANALYSIS_UPLOAD_META_CACHE["meta"] = None
    RECIPES_BOOTSTRAP_RESPONSE_CACHE["signature"] = None
    RECIPES_BOOTSTRAP_RESPONSE_CACHE["response_json"] = None


def get_scope_reference_date(frame: pd.DataFrame) -> pd.Timestamp | None:
    if frame.empty or "Date" not in frame.columns:
        return None
    date_series = pd.to_datetime(frame["Date"], errors="coerce").dropna()
    if date_series.empty:
        return None
    return pd.Timestamp(date_series.max()).normalize()


def hydrate_dataframe_from_session(columns: list[str], records: list[dict[str, Any]]) -> pd.DataFrame:
    if not isinstance(columns, list):
        columns = []
    if not isinstance(records, list):
        records = []
    dataframe = pd.DataFrame(records)
    if columns:
        dataframe = dataframe.reindex(columns=columns)
    return dataframe


def get_json_payload_size_bytes(payload: Any) -> int:
    try:
        return len(json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8"))
    except Exception:
        return 0


def compact_quote_compare_evaluation_for_session(evaluation: dict[str, Any] | None) -> dict[str, Any] | None:
    if not isinstance(evaluation, dict):
        return None
    return {
        "summary": evaluation.get("summary") if isinstance(evaluation.get("summary"), dict) else {},
        "currencies": evaluation.get("currencies") if isinstance(evaluation.get("currencies"), list) else [],
        "insights": evaluation.get("insights") if isinstance(evaluation.get("insights"), list) else []
    }


def compact_quote_compare_active_session_payload(payload: dict[str, Any]) -> dict[str, Any]:
    compacted_payload = dict(payload or {})
    if str(compacted_payload.get("step") or "").strip().lower() == "analyze":
        compacted_payload["evaluation"] = compact_quote_compare_evaluation_for_session(compacted_payload.get("evaluation")) or {}
    return compacted_payload


def save_quote_compare_active_session(
    session_id: str,
    payload: dict[str, Any],
    user_id: int | str | None = None,
    perf_label_prefix: str | None = None
) -> None:
    resolved_user_id = get_storage_user_id(user_id)
    persist_started_at = perf_counter()
    persist_substeps: dict[str, Any] = {
        "step": str((payload or {}).get("step") or "").strip().lower()
    }

    compact_started_at = perf_counter()
    compacted_payload = compact_quote_compare_active_session_payload(payload)
    persist_substeps["compact_ms"] = round((perf_counter() - compact_started_at) * 1000, 1)

    normalize_started_at = perf_counter()
    payload_is_json_safe = bool(compacted_payload.pop("_json_safe_payload", False))
    if payload_is_json_safe:
        normalized_payload = dict(compacted_payload)
        persist_substeps["json_safe_skipped"] = True
    else:
        normalized_payload = make_json_safe(compacted_payload)
        persist_substeps["json_safe_skipped"] = False
    persist_substeps["json_safe_ms"] = round((perf_counter() - normalize_started_at) * 1000, 1)
    normalized_payload["session_id"] = session_id

    session_cache_dir = ensure_quote_compare_session_cache_dir(resolved_user_id)
    cleanup_started_at = perf_counter()
    removed_session_file_count = 0
    for existing_session_path in session_cache_dir.glob("*.json"):
        try:
            existing_session_path.unlink()
            removed_session_file_count += 1
        except FileNotFoundError:
            continue
    persist_substeps["cleanup_ms"] = round((perf_counter() - cleanup_started_at) * 1000, 1)
    persist_substeps["removed_session_file_count"] = removed_session_file_count

    session_path = get_quote_compare_session_path(session_id, user_id=resolved_user_id)
    serialization_started_at = perf_counter()
    session_payload_text = json.dumps(normalized_payload, ensure_ascii=False, separators=(",", ":"))
    session_payload_bytes = session_payload_text.encode("utf-8")
    persist_substeps["serialization_ms"] = round((perf_counter() - serialization_started_at) * 1000, 1)
    persist_substeps["serialized_bytes"] = len(session_payload_bytes)

    session_write_started_at = perf_counter()
    session_path.write_bytes(session_payload_bytes)
    persist_substeps["session_write_ms"] = round((perf_counter() - session_write_started_at) * 1000, 1)

    active_write_started_at = perf_counter()
    get_user_active_quote_compare_session_path(resolved_user_id).write_bytes(session_payload_bytes)
    persist_substeps["active_write_ms"] = round((perf_counter() - active_write_started_at) * 1000, 1)
    persist_substeps["file_write_ms"] = round(persist_substeps["session_write_ms"] + persist_substeps["active_write_ms"], 1)
    persist_substeps["total_ms"] = round((perf_counter() - persist_started_at) * 1000, 1)
    LATEST_ANALYSIS_UPLOAD_META_CACHE["signature"] = None
    LATEST_ANALYSIS_UPLOAD_META_CACHE["meta"] = None

    if perf_label_prefix:
        log_perf_details(f"{perf_label_prefix}.save_active_session.substeps", **persist_substeps)

    logger.info(
        "[Compare Prices active session persist] session_id=%s persisted_bytes=%s step=%s",
        bool(session_id),
        len(session_payload_bytes),
        normalized_payload.get("step")
    )


def load_quote_compare_active_session(
    session_id: str | None,
    user_id: int | str | None = None
) -> dict[str, Any] | None:
    resolved_user_id = get_storage_user_id(user_id)
    normalized_session_id = str(session_id or "").strip()
    candidate_paths: list[Path] = []
    if normalized_session_id:
        candidate_paths.append(get_quote_compare_session_path(normalized_session_id, user_id=resolved_user_id))
    candidate_paths.append(get_user_active_quote_compare_session_path(resolved_user_id))

    for session_path in candidate_paths:
        if not session_path.exists():
            continue
        try:
            payload = json.loads(session_path.read_text(encoding="utf-8"))
            return payload if isinstance(payload, dict) else None
        except json.JSONDecodeError:
            logger.warning("Compare Prices session file is invalid: %s", session_path)
            continue
    return None


def validate_quote_compare_active_session(session_payload: dict[str, Any] | None) -> dict[str, Any] | None:
    if not isinstance(session_payload, dict):
        return None

    session_id = str(session_payload.get("session_id") or "").strip()
    step = str(session_payload.get("step") or "").strip().lower()
    dataframe = session_payload.get("dataframe") or {}
    cached_upload_path = str(session_payload.get("cached_upload_path") or "").strip()
    headers = session_payload.get("headers")
    dataframe_columns = dataframe.get("columns") if isinstance(dataframe, dict) else None
    dataframe_records = dataframe.get("records") if isinstance(dataframe, dict) else None

    if not session_id or step not in {"review", "analyze"}:
        return None
    if not isinstance(headers, list) or not headers:
        return None
    has_hydrated_dataframe = isinstance(dataframe_columns, list) and dataframe_columns and isinstance(dataframe_records, list)
    has_cached_upload = bool(cached_upload_path)
    if not has_hydrated_dataframe and not has_cached_upload:
        return None

    if step == "review":
        return session_payload

    comparison = session_payload.get("comparison")
    evaluation = session_payload.get("evaluation")
    if not isinstance(comparison, dict) or not isinstance(evaluation, dict):
        return None

    session_payload = dict(session_payload)
    session_payload["evaluation"] = compact_quote_compare_evaluation_for_session(evaluation) or {}
    return session_payload


GUIDE_ASSISTANT_FALLBACK = (
    "Guide can help with upload workflow, supplier comparison, Product History, reset behavior, "
    "saved recipe behavior, and why rows may not compare cleanly. Ask what to do first, what Reset all data removes, "
    "or what Product History shows."
)
GUIDE_ASSISTANT_STOPWORDS = {
    "a", "an", "and", "are", "can", "do", "find", "for", "help", "how", "i", "in",
    "is", "me", "my", "of", "the", "this", "to", "use", "with", "you", "your"
}
GUIDE_ASSISTANT_SYNONYMS = {
    "analyse": "analyze",
    "analysing": "analyze",
    "analysis": "analyze",
    "comparing": "compare",
    "comparison": "compare",
    "differences": "compare",
    "difference": "compare",
    "inspect": "review",
    "inspecting": "review",
    "vendors": "supplier",
    "vendor": "supplier",
    "suppliers": "supplier",
    "prices": "price",
    "pricing": "price",
    "cost": "price",
    "costs": "price",
    "overpaying": "overpay",
    "overspend": "overpay",
    "overspending": "overpay",
    "expensive": "overpay",
    "opportunity": "savings",
    "opportunities": "savings",
    "margin": "savings",
    "add": "upload",
    "import": "upload",
    "importing": "upload",
    "adding": "upload",
    "recipes": "recipe",
    "ingredients": "ingredient",
    "menu": "recipe",
    "food": "recipe"
}
GUIDE_ACTION_CATALOG = {
    "go_upload": {"label": "Open Compare Prices", "href": "/quote-compare"},
    "go_top_insights": {"label": "Open Compare Prices", "href": "/quote-compare"},
    "go_workspace": {"label": "Open Compare Prices", "href": "/quote-compare"},
    "go_ask_data": {"label": "Open Compare Prices", "href": "/quote-compare"},
    "go_quote_compare": {"label": "Open Compare Prices", "href": "/quote-compare"},
    "go_recipes": {"label": "Open Recipes", "href": "/recipes"}
}


def load_guide_knowledge() -> list[dict[str, Any]]:
    try:
        payload = json.loads(GUIDE_KNOWLEDGE_PATH.read_text(encoding="utf-8"))
    except FileNotFoundError:
        logger.warning("Guide knowledge file not found: %s", GUIDE_KNOWLEDGE_PATH)
        return []
    except json.JSONDecodeError:
        logger.exception("Guide knowledge file is invalid: %s", GUIDE_KNOWLEDGE_PATH)
        return []

    entries = payload.get("entries", [])
    return entries if isinstance(entries, list) else []


def normalize_guide_text(value: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower())
    normalized_tokens = []
    for token in cleaned.split():
        normalized_tokens.append(GUIDE_ASSISTANT_SYNONYMS.get(token, token))
    return " ".join(normalized_tokens)


def tokenize_guide_text(value: str) -> set[str]:
    return {
        token for token in normalize_guide_text(value).split()
        if len(token) > 2 and token not in GUIDE_ASSISTANT_STOPWORDS
    }


def find_guide_answer(question: str) -> dict[str, Any]:
    normalized_question = normalize_guide_text(question)
    question_tokens = tokenize_guide_text(question)
    best_entry: dict[str, Any] | None = None
    best_score = 0

    for entry in load_guide_knowledge():
        title = normalize_guide_text(entry.get("title", ""))
        intents = [normalize_guide_text(intent) for intent in entry.get("intents", []) if intent]
        keywords = [normalize_guide_text(keyword) for keyword in entry.get("keywords", []) if keyword]
        synonyms = [normalize_guide_text(keyword) for keyword in entry.get("synonyms", []) if keyword]
        examples = [normalize_guide_text(example) for example in entry.get("example_questions", []) if example]
        intent_groups = [set(tokenize_guide_text(value)) for value in entry.get("intent_groups", []) if value]
        weighted_phrase_groups = [
            ("title", [value for value in [title] if value]),
            ("intent", intents),
            ("example", examples),
            ("synonym", synonyms)
        ]
        candidate_keywords = [*keywords, *synonyms]
        score = 0
        phrase_match = False
        entry_token_pool = set()

        for group_name, values in weighted_phrase_groups:
            for candidate in values:
                candidate_tokens = set(candidate.split())
                entry_token_pool.update(candidate_tokens)
                overlap = len(candidate_tokens & question_tokens)
                exact_weight = 13 if group_name == "intent" else 11
                contains_weight = 10 if group_name in {"intent", "example"} else 8
                subset_weight = 8 if group_name == "intent" else 6
                overlap_weight = 5 if group_name in {"intent", "example"} else 3

                if normalized_question == candidate:
                    score += exact_weight
                    phrase_match = True
                elif candidate in normalized_question:
                    score += contains_weight
                    phrase_match = True
                elif candidate_tokens and candidate_tokens.issubset(question_tokens):
                    score += subset_weight
                    phrase_match = True
                elif overlap >= 2:
                    score += overlap_weight
                    phrase_match = True

        for intent_group in intent_groups:
            entry_token_pool.update(intent_group)
            if intent_group and len(intent_group & question_tokens) >= min(len(intent_group), 2):
                score += 7
                phrase_match = True

        keyword_hits = 0
        for keyword in candidate_keywords:
            if not keyword:
                continue
            keyword_tokens = set(keyword.split())
            entry_token_pool.update(keyword_tokens)
            if keyword in normalized_question:
                keyword_hits += 1
            elif keyword_tokens and keyword_tokens.issubset(question_tokens):
                keyword_hits += 1
            elif len(keyword_tokens & question_tokens) >= 1:
                keyword_hits += 1

        score += keyword_hits * 2
        if phrase_match and keyword_hits:
            score += 2

        if question_tokens and entry_token_pool:
            token_overlap = len(question_tokens & entry_token_pool)
            if token_overlap:
                score += token_overlap

        if score > best_score:
            best_score = score
            best_entry = entry

    if not best_entry or best_score < 6:
        return {
            "found": False,
            "id": None,
            "title": "Guide Assistant",
            "answer": GUIDE_ASSISTANT_FALLBACK,
            "related_section": None,
            "next_step": None,
            "actions": [],
            "workflow_steps": []
        }

    return {
        "found": True,
        "id": best_entry.get("id"),
        "title": str(best_entry.get("title") or "Guide Answer"),
        "answer": str(best_entry.get("answer") or GUIDE_ASSISTANT_FALLBACK),
        "related_section": best_entry.get("related_section"),
        "next_step": best_entry.get("next_step"),
        "actions": list(best_entry.get("actions") or []),
        "workflow_steps": [str(step) for step in best_entry.get("workflow_steps", []) if step]
    }


def load_latest_results_frame() -> pd.DataFrame | None:
    latest_results_path = get_current_latest_results_path()
    if not latest_results_path.exists():
        return None
    try:
        frame = pd.read_csv(latest_results_path)
    except Exception:
        logger.exception("Failed to load latest results for Guide context")
        return None
    return frame if not frame.empty else None


def get_first_existing_column(frame: pd.DataFrame, candidates: list[str]) -> str | None:
    for candidate in candidates:
        if candidate in frame.columns:
            return candidate
    return None


def get_numeric_series(frame: pd.DataFrame, column_name: str | None) -> pd.Series:
    if not column_name or column_name not in frame.columns:
        return pd.Series(dtype="float64")
    return pd.to_numeric(frame[column_name], errors="coerce")


def count_unique_products_by_name_unit(
    frame: pd.DataFrame,
    *,
    product_column: str = "Product Name",
    unit_column: str = "Unit"
) -> int:
    if frame.empty or product_column not in frame.columns or unit_column not in frame.columns:
        return 0
    unique_products = (
        frame.loc[:, [product_column, unit_column]]
        .copy()
        .fillna("")
    )
    unique_products[product_column] = unique_products[product_column].astype(str).str.strip()
    unique_products[unit_column] = unique_products[unit_column].astype(str).str.strip()
    unique_products = unique_products[unique_products[product_column].astype(bool)]
    return int(len(unique_products.drop_duplicates(subset=[product_column, unit_column]).index))


def count_unique_product_names(
    frame: pd.DataFrame,
    *,
    product_column: str = "Product Name"
) -> int:
    if frame.empty or product_column not in frame.columns:
        return 0
    unique_products = (
        frame.loc[:, [product_column]]
        .copy()
        .fillna("")
    )
    unique_products[product_column] = unique_products[product_column].astype(str).str.strip()
    unique_products = unique_products[unique_products[product_column].astype(bool)]
    return int(len(unique_products.drop_duplicates(subset=[product_column]).index))


def build_guide_analysis_snapshot() -> dict[str, Any] | None:
    frame = load_latest_results_frame()
    if frame is None:
        return None

    product_column = get_first_existing_column(frame, ["Product Name", "Product"])
    supplier_column = get_first_existing_column(frame, ["Supplier"])
    unit_column = get_first_existing_column(frame, ["Purchase Unit", "Unit"])
    savings_column = get_first_existing_column(frame, ["Savings Opportunity"])
    status_column = get_first_existing_column(frame, ["Status"])
    unit_price_column = get_first_existing_column(frame, ["Unit Price", "Average Price"])

    total_rows = int(len(frame))
    product_count = count_unique_products_by_name_unit(frame, product_column=product_column, unit_column=unit_column) if product_column and unit_column else 0
    supplier_count = int(frame[supplier_column].nunique()) if supplier_column else 0
    savings_series = get_numeric_series(frame, savings_column).fillna(0)
    unit_price_series = get_numeric_series(frame, unit_price_column)
    overpay_rows = int((frame[status_column] == "Overpay").sum()) if status_column else 0
    good_deal_rows = int((frame[status_column] == "Good Deal").sum()) if status_column else 0
    total_savings = round(float(savings_series.sum()), 2) if not savings_series.empty else 0.0
    average_savings_per_overpay = round(total_savings / overpay_rows, 2) if overpay_rows else 0.0

    compare_ready_products = 0
    if product_column and supplier_column:
        compare_ready_products = int((frame.groupby(product_column)[supplier_column].nunique() > 1).sum())

    top_supplier = None
    top_supplier_savings = 0.0
    if supplier_column and savings_column:
        supplier_savings = (
            frame.groupby(supplier_column, as_index=False)[savings_column]
            .sum()
            .sort_values(savings_column, ascending=False)
        )
        if not supplier_savings.empty:
            top_supplier = str(supplier_savings.iloc[0][supplier_column])
            top_supplier_savings = round(float(supplier_savings.iloc[0][savings_column]), 2)

    highest_risk_product = None
    highest_risk_amount = 0.0
    top_savings_products: list[dict[str, Any]] = []
    if product_column and savings_column:
        product_savings = (
            frame.groupby(product_column, as_index=False)[savings_column]
            .sum()
            .sort_values(savings_column, ascending=False)
        )
        if not product_savings.empty:
            highest_risk_product = str(product_savings.iloc[0][product_column])
            highest_risk_amount = round(float(product_savings.iloc[0][savings_column]), 2)
            top_savings_products = [
                {
                    "product": str(row[product_column]),
                    "savings": round(float(row[savings_column]), 2)
                }
                for _, row in product_savings.head(3).iterrows()
                if str(row[product_column]).strip()
            ]

    multi_unit_products = 0
    if product_column and unit_column:
        multi_unit_products = int((frame.groupby(product_column)[unit_column].nunique() > 1).sum())

    price_movement_product = None
    price_movement_pct = 0.0
    if product_column and not unit_price_series.empty:
        price_spread_frame = frame.assign(__guide_unit_price=unit_price_series).dropna(subset=["__guide_unit_price"])
        if not price_spread_frame.empty:
            grouped_prices = (
                price_spread_frame.groupby(product_column)["__guide_unit_price"]
                .agg(["min", "max", "count"])
                .reset_index()
            )
            grouped_prices = grouped_prices[grouped_prices["count"] > 1]
            if not grouped_prices.empty:
                grouped_prices["spread_pct"] = grouped_prices.apply(
                    lambda row: ((row["max"] - row["min"]) / row["min"] * 100) if row["min"] else 0.0,
                    axis=1
                )
                grouped_prices = grouped_prices.sort_values("spread_pct", ascending=False)
                if not grouped_prices.empty:
                    price_movement_product = str(grouped_prices.iloc[0][product_column])
                    price_movement_pct = round(float(grouped_prices.iloc[0]["spread_pct"]), 1)

    snapshot_lines = [
        {"label": "Products analyzed", "value": str(product_count or total_rows or "--")},
        {"label": "Suppliers in view", "value": str(supplier_count or "--")},
        {"label": "Products with savings", "value": str(compare_ready_products or 0)},
        {"label": "Visible savings", "value": format_currency(total_savings) if total_savings > 0 else "--"}
    ]

    unique_product_name_count = int(frame[product_column].nunique()) if product_column else 0
    unique_product_name_unit_count = (
        count_unique_products_by_name_unit(frame, product_column=product_column, unit_column=unit_column)
        if product_column and unit_column
        else 0
    )
    print(
        "[DEBUG] Products analyzed count check | "
        f"unique Product Name count={unique_product_name_count} | "
        f"unique Product Name + Unit count={unique_product_name_unit_count} | "
        f"current count used={product_count}"
    )

    return {
        "total_rows": total_rows,
        "product_count": product_count,
        "supplier_count": supplier_count,
        "overpay_rows": overpay_rows,
        "good_deal_rows": good_deal_rows,
        "total_savings": total_savings,
        "average_savings_per_overpay": average_savings_per_overpay,
        "compare_ready_products": compare_ready_products,
        "top_supplier": top_supplier,
        "top_supplier_savings": top_supplier_savings,
        "highest_risk_product": highest_risk_product,
        "highest_risk_amount": highest_risk_amount,
        "multi_unit_products": multi_unit_products,
        "price_movement_product": price_movement_product,
        "price_movement_pct": price_movement_pct,
        "top_savings_products": top_savings_products,
        "snapshot_lines": snapshot_lines
    }


def resolve_guide_actions(action_ids: list[str] | None) -> list[dict[str, str]]:
    resolved: list[dict[str, str]] = []
    for action_id in action_ids or []:
        action = GUIDE_ACTION_CATALOG.get(str(action_id))
        if not action:
            continue
        if any(existing["href"] == action["href"] for existing in resolved):
            continue
        resolved.append({"label": action["label"], "href": action["href"]})
    return resolved


def append_guide_action(actions: list[dict[str, str]], action_id: str) -> list[dict[str, str]]:
    action = GUIDE_ACTION_CATALOG.get(action_id)
    if not action:
        return actions
    if any(existing["href"] == action["href"] for existing in actions):
        return actions
    return [*actions, {"label": action["label"], "href": action["href"]}]


def build_beginner_guide_steps(_: dict[str, Any] | None = None) -> list[str]:
    return [
        "Upload a CSV or Excel file in Compare Prices.",
        "Confirm the required column mappings before you continue.",
        "Use the comparison screen to review matched supplier prices on a consistent basis.",
        "Open Product History when you want to check whether a price change looks persistent or temporary.",
        "Move into Recipes only when you want to turn uploaded purchase prices into recipe costing."
    ]


def build_general_guide_response(question: str) -> dict[str, Any]:
    normalized_question = normalize_guide_text(question)

    if "compare" in normalized_question and "supplier" in normalized_question:
        return {
            "found": True,
            "id": "general-supplier-compare",
            "title": "Compare suppliers by product and unit",
            "answer": "Use supplier comparison only when the same item is being matched on the same unit and buying basis. First confirm the mapping, then compare like-for-like prices, and only after that weigh softer trade-offs such as delivery or payment terms.",
            "why_this_matters": "This keeps the comparison honest. A lower price can be misleading if the rows are really describing different units, pack sizes, or quantity assumptions.",
            "related_section": "Compare Prices",
            "next_step": "Open Compare Prices, confirm the mappings, and review the comparison table with unit consistency in mind.",
            "actions": resolve_guide_actions(["go_quote_compare"]),
            "workflow_steps": [
                "Check that product, supplier, unit, quantity, price, and date are mapped correctly.",
                "Open a matched comparison view rather than scanning raw rows first.",
                "Compare on a like-for-like basis before making a sourcing decision."
            ],
            "analysis_snapshot": [],
            "context_note": "Guide is answering as a product-usage assistant, not from uploaded dataset results.",
            "context_available": False
        }

    return {
        "found": False,
        "id": "general-product-guidance",
        "title": "Guide can explain how the product works",
        "answer": "Ask Guide about workflow, screen purpose, interpretation, reset behavior, saved recipe behavior, or how to compare rows correctly. It is designed to explain how to use the product rather than summarize your uploaded data.",
        "why_this_matters": "That keeps the answers stable and practical. You get help with what to do next, what each screen means, and how to avoid common mistakes while using the product.",
        "related_section": "Guide Workspace",
        "next_step": "Try a usage question such as how to upload a file, how to compare suppliers, what Product History shows, or what reset removes.",
        "actions": resolve_guide_actions(["go_quote_compare"]),
        "workflow_steps": [
            "Choose the screen you want to understand.",
            "Ask Guide what that screen is for or how to use it.",
            "Follow the suggested workflow in the answer."
        ],
        "analysis_snapshot": [],
        "context_note": "Guide answers from built-in product guidance only.",
        "context_available": False
    }


def build_guide_response(question: str) -> dict[str, Any]:
    response = find_guide_answer(question)
    actions = resolve_guide_actions(response.get("actions"))
    workflow_steps = [str(step) for step in response.get("workflow_steps", []) if step]
    context_note = "Guide answers from built-in product guidance only."
    why_this_matters = None

    if response.get("id") == "beginner-flow":
        workflow_steps = build_beginner_guide_steps(None)
        for action_id in ["go_upload", "go_top_insights", "go_workspace", "go_ask_data"]:
            actions = append_guide_action(actions, action_id)

    if response.get("found"):
        why_this_matters = "This guidance is meant to help you understand the product flow, avoid common usage mistakes, and choose the right workspace for the task in front of you."
    else:
        fallback_response = build_general_guide_response(question)
        return fallback_response

    return {
        **response,
        "actions": actions,
        "workflow_steps": workflow_steps,
        "context_note": context_note,
        "why_this_matters": why_this_matters,
        "analysis_snapshot": [],
        "context_available": False
    }


def parse_days_from_text(value: str | None) -> int | None:
    value = normalize_request_value(value)
    text = str(value or "").strip().lower()
    if not text:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)", text)
    if not match:
        return None
    return max(safe_int(match.group(1)), 0)


def parse_payment_term_days(value: str | None) -> int | None:
    text = str(value or "").strip().lower()
    if not text:
        return None
    if "advance" in text or "prepay" in text or "upfront" in text or "cash" in text:
        return 0
    return parse_days_from_text(text)


def normalize_quote_weighting(weighting: dict[str, Any] | None = None) -> dict[str, float]:
    normalized = {}
    source = weighting or {}
    for key, default_value in QUOTE_COMPARE_DEFAULT_WEIGHTS.items():
        raw_value = normalize_value(source.get(key, default_value))
        if isinstance(raw_value, bool) or (
            isinstance(raw_value, str) and raw_value.lower() in {"true", "false", ""}
        ) or raw_value is None:
            logger.info(
                "[Compare Prices upload] preserving non-numeric weighting for %s: %r",
                key,
                raw_value
            )
            normalized[key] = default_value
            continue
        try:
            normalized[key] = max(float(raw_value), 0.0)
        except (TypeError, ValueError):
            logger.warning(
                "[Compare Prices upload] failed numeric coercion for %s (quote weighting): type=%s value=%r",
                key,
                type(raw_value).__name__,
                raw_value
            )
            normalized[key] = default_value

    total_weight = sum(normalized.values())
    if total_weight <= 0:
        return dict(QUOTE_COMPARE_DEFAULT_WEIGHTS)

    return {
        key: round(value / total_weight, 4)
        for key, value in normalized.items()
    }


def normalize_quote_comparison_payload(payload: dict[str, Any]) -> dict[str, Any]:
    normalized_bids: list[dict[str, Any]] = []
    for index, bid in enumerate(payload.get("bids", []), start=1):
        row_context = f"normalized payload row {index}"
        supplier_name = normalize_text_value(bid.get("supplier_name", ""))
        product_name = normalize_text_value(bid.get("product_name", ""))
        unit = normalize_text_value(bid.get("unit", ""))
        quote_date = normalize_text_value(bid.get("quote_date", bid.get("date", "")))
        currency = normalize_text_value(bid.get("currency", "")).upper()
        delivery_time = normalize_text_value(bid.get("delivery_time", ""))
        payment_term = normalize_text_value(bid.get("payment_term", ""))
        valid_until = normalize_text_value(bid.get("valid_until", ""))
        notes = normalize_text_value(bid.get("notes", ""))

        quantity = coerce_numeric_value(
            bid.get("quantity", 0),
            field_name="quantity",
            context=row_context
        )
        unit_price = coerce_numeric_value(
            bid.get("unit_price", 0),
            field_name="unit_price",
            context=row_context
        )
        total_price = coerce_numeric_value(
            bid.get("total_price", 0),
            field_name="total_price",
            context=row_context
        )

        if not supplier_name and not product_name and quantity <= 0 and unit_price <= 0 and total_price <= 0:
            continue

        resolved_total = total_price if total_price > 0 else unit_price * quantity
        normalized_bids.append({
            "supplier_name": supplier_name,
            "product_name": product_name,
            "unit": unit,
            "quantity": quantity,
            "unit_price": round(unit_price, 4),
            "total_price": round(resolved_total, 4),
            "quote_date": quote_date,
            "currency": currency or "USD",
            "delivery_time": delivery_time,
            "payment_term": payment_term,
            "valid_until": valid_until,
            "notes": notes
        })

    return {
        "comparison_id": normalize_text_value(payload.get("comparison_id", "")) or None,
        "upload_id": normalize_text_value(payload.get("upload_id", "")) or None,
        "name": normalize_text_value(payload.get("name", "")),
        "sourcing_need": normalize_text_value(payload.get("sourcing_need", "")),
        "weighting": normalize_quote_weighting(payload.get("weighting")),
        "source_type": normalize_text_value(payload.get("source_type", "")) or "manual",
        "mode": normalize_text_value(payload.get("mode", "")) or "compare",
        "bids": normalized_bids
    }


def validate_quote_comparison_payload(comparison: dict[str, Any], *, require_name: bool = False) -> None:
    if require_name and not comparison["name"]:
        raise ValueError("Enter a comparison name before evaluating or saving.")
    if not comparison["bids"]:
        raise ValueError("Add at least one supplier offer before evaluating the comparison.")

    for bid in comparison["bids"]:
        if not bid["supplier_name"]:
            raise ValueError("Each offer must include a supplier name.")
        if not bid["product_name"]:
            raise ValueError("Each offer must include a product name.")
        if bid["quantity"] <= 0:
            raise ValueError("Each offer must include a quantity greater than zero.")
        if bid["total_price"] <= 0:
            raise ValueError("Each offer must include a total price greater than zero.")


def build_pre_normalized_quote_comparison(
    *,
    bids: list[dict[str, Any]],
    upload_id: str | None = None,
    name: str = "",
    sourcing_need: str = "",
    weighting: dict[str, Any] | None = None,
    source_type: str = "manual",
    mode: str = "compare"
) -> dict[str, Any]:
    return {
        "comparison_id": None,
        "upload_id": normalize_text_value(upload_id) or None,
        "name": normalize_text_value(name),
        "sourcing_need": normalize_text_value(sourcing_need),
        "weighting": normalize_quote_weighting(weighting),
        "source_type": normalize_text_value(source_type) or "manual",
        "mode": normalize_text_value(mode) or "compare",
        "bids": list(bids or [])
    }


def build_quote_bid_import_result(dataframe: pd.DataFrame, *, text_already_normalized: bool = False) -> dict[str, Any]:
    import_started_at = perf_counter()
    import_substeps: dict[str, Any] = {
        "text_already_normalized": text_already_normalized,
        "source_rows": int(len(dataframe.index)),
        "source_columns": int(len(dataframe.columns))
    }
    bids: list[dict[str, Any]] = []
    skipped_row_count = 0
    numeric_preview: list[dict[str, float]] = []
    positive_total_count = 0
    has_total_price = "Total Price" in dataframe.columns
    column_prepare_started_at = perf_counter()
    row_count = len(dataframe.index)

    def get_column_values(column_name: str) -> list[Any]:
        if column_name in dataframe.columns:
            return dataframe[column_name].tolist()
        return [""] * row_count

    supplier_values = get_column_values("Supplier")
    supplier_name_values = get_column_values("Supplier Name")
    product_name_values = get_column_values("Product Name")
    unit_values = get_column_values("Unit")
    quantity_values = get_column_values("Quantity")
    unit_price_values = get_column_values("Unit Price")
    total_price_values = get_column_values("Total Price") if has_total_price else None
    quote_date_values = get_column_values("Date")
    currency_values = get_column_values("Currency")
    delivery_time_values = get_column_values("Delivery Time")
    payment_terms_values = get_column_values("Payment Terms")
    valid_until_values = get_column_values("Valid Until")
    notes_values = get_column_values("Notes")
    import_substeps["prepare_columns_ms"] = round((perf_counter() - column_prepare_started_at) * 1000, 1)

    def normalize_import_text(value: Any) -> str:
        if text_already_normalized and isinstance(value, str):
            return value
        return normalize_text_value(value)

    row_loop_started_at = perf_counter()
    append_bid = bids.append
    coerce_numeric = coerce_numeric_value
    for index in range(row_count):
        row_number = index + 1
        supplier_value = supplier_values[index]
        supplier_name_value = supplier_name_values[index]
        product_name_value = product_name_values[index]
        unit_value = unit_values[index]
        quantity_value = quantity_values[index]
        unit_price_value = unit_price_values[index]
        total_price_value = total_price_values[index] if total_price_values is not None else 0.0
        quote_date_value = quote_date_values[index]
        currency_value = currency_values[index]
        delivery_time_value = delivery_time_values[index]
        payment_terms_value = payment_terms_values[index]
        valid_until_value = valid_until_values[index]
        notes_value = notes_values[index]
        row_context = f"uploaded dataframe row {row_number}"
        supplier_name = normalize_import_text(supplier_value or supplier_name_value)
        product_name = normalize_import_text(product_name_value)
        unit = normalize_import_text(unit_value)

        quantity = coerce_numeric(
            quantity_value,
            field_name="Quantity",
            context=row_context
        )
        unit_price = coerce_numeric(
            unit_price_value,
            field_name="Unit Price",
            context=row_context
        )
        total_price = (
            coerce_numeric(
                total_price_value,
                field_name="Total Price",
                context=row_context
            )
            if has_total_price
            else 0.0
        )

        if not supplier_name and not product_name and quantity <= 0 and unit_price <= 0 and total_price <= 0:
            skipped_row_count += 1
            continue
        if not supplier_name:
            logger.warning(
                "[Compare Prices upload] skipping row with empty resolved supplier | row_index=%s | supplier_value=%r | product_name=%r | quantity=%r | unit_price=%r",
                row_number,
                supplier_value or supplier_name_value,
                product_name,
                quantity,
                unit_price
            )
            skipped_row_count += 1
            continue
        if quantity <= 0:
            logger.warning(
                "[Compare Prices upload] skipping row with invalid resolved quantity | row_index=%s | supplier_name=%r | quantity=%r | unit_price=%r | total_price=%r",
                row_number,
                supplier_name,
                quantity,
                unit_price,
                total_price
            )
            skipped_row_count += 1
            continue

        resolved_total = total_price if total_price > 0 else quantity * unit_price
        if unit_price <= 0 and resolved_total <= 0:
            logger.warning(
                "[Compare Prices upload] skipping row with invalid resolved pricing | row_index=%s | supplier_name=%r | quantity=%r | unit_price=%r | total_price=%r | resolved_total=%r",
                row_number,
                supplier_name,
                quantity,
                unit_price,
                total_price,
                resolved_total
            )
            skipped_row_count += 1
            continue
        if len(numeric_preview) < 10:
            numeric_preview.append({
                "quantity": round(quantity, 4),
                "unit_price": round(unit_price, 4),
                "total_price": round(total_price, 4),
                "resolved_total": round(resolved_total, 4)
            })
        if resolved_total > 0:
            positive_total_count += 1
        append_bid({
            "supplier_name": supplier_name,
            "product_name": product_name,
            "unit": unit,
            "quantity": round(quantity, 4),
            "unit_price": round(unit_price, 4),
            "total_price": round(resolved_total, 4),
            "quote_date": normalize_import_text(quote_date_value),
            "currency": normalize_import_text(currency_value).upper() or "USD",
            "delivery_time": normalize_import_text(delivery_time_value),
            "payment_term": normalize_import_text(payment_terms_value),
            "valid_until": normalize_import_text(valid_until_value),
            "notes": normalize_import_text(notes_value)
        })
    import_substeps["row_loop_ms"] = round((perf_counter() - row_loop_started_at) * 1000, 1)

    logging_started_at = perf_counter()
    logger.info(
        "[Compare Prices upload] numeric debug | first_10_numeric_values=%s | rows_with_resolved_total_gt_zero=%s | skipped_row_count=%s | valid_row_count=%s",
        numeric_preview,
        positive_total_count,
        skipped_row_count,
        len(bids)
    )
    import_substeps["logging_ms"] = round((perf_counter() - logging_started_at) * 1000, 1)
    import_substeps["valid_row_count"] = len(bids)
    import_substeps["skipped_row_count"] = skipped_row_count
    import_substeps["positive_total_count"] = positive_total_count
    import_substeps["total_ms"] = round((perf_counter() - import_started_at) * 1000, 1)
    log_perf_details("confirm.import_rows.substeps", **import_substeps)

    return {
        "bids": bids,
        "skipped_row_count": skipped_row_count,
        "valid_row_count": len(bids)
    }


def build_quote_bids_from_dataframe(dataframe: pd.DataFrame) -> list[dict[str, Any]]:
    return build_quote_bid_import_result(dataframe)["bids"]


def normalize_quote_compare_mapped_dataframe(
    dataframe: pd.DataFrame,
    *,
    selected_mapping: dict[str, Any] | None = None,
    source_columns: list[str] | None = None
) -> pd.DataFrame:
    normalized = dataframe.copy()
    text_columns = [
        "Supplier",
        "Product Name",
        "Unit",
        "Date",
        "Currency",
        "Delivery Time",
        "Payment Terms",
        "Valid Until",
        "Notes"
    ]

    for column in text_columns:
        if column in normalized.columns:
            normalized[column] = normalized[column].fillna("").map(normalize_text_value)

    supplier_source_column = str((selected_mapping or {}).get("Supplier") or "").strip()
    supplier_column_exists = supplier_source_column in (source_columns or [])
    supplier_preview: list[str] = []
    supplier_non_empty_count = 0
    if "Supplier" in normalized.columns:
        supplier_series = normalized["Supplier"]
        supplier_preview = supplier_series.head(10).tolist()
        supplier_non_empty_count = int(supplier_series.astype(bool).sum())

    logger.info(
        "[Compare Prices upload] supplier mapping debug | selected_supplier_column=%s | exists_in_source=%s | mapped_columns=%s | first_10_supplier_values=%s | non_empty_supplier_count=%s",
        supplier_source_column or "<empty>",
        supplier_column_exists,
        list(normalized.columns),
        supplier_preview,
        supplier_non_empty_count
    )

    return normalized


def normalize_metric_scores(values: list[float], *, reverse: bool = False) -> list[float]:
    if not values:
        return []
    minimum = min(values)
    maximum = max(values)
    if maximum == minimum:
        return [100.0 for _ in values]
    scores: list[float] = []
    for value in values:
        base = (value - minimum) / (maximum - minimum)
        score = (1 - base) * 100 if reverse else base * 100
        scores.append(round(score, 2))
    return scores


def calculate_quote_comparison(comparison: dict[str, Any]) -> dict[str, Any]:
    calculation_started_at = perf_counter()
    calculation_substeps: dict[str, Any] = {
        "input_bids": int(len((comparison or {}).get("bids", [])))
    }
    validate_quote_comparison_payload(comparison)
    build_breakdown_started_at = perf_counter()
    currency_set: set[str] = set()
    bid_breakdown: list[dict[str, Any]] = []
    product_groups: dict[tuple[str, str], list[dict[str, Any]]] = {}
    lowest_price_bid: dict[str, Any] | None = None
    lowest_price_key: tuple[float, str] | None = None
    fastest_delivery_bid: dict[str, Any] | None = None
    fastest_delivery_key: tuple[int, str] | None = None
    best_payment_bid: dict[str, Any] | None = None
    best_payment_key: tuple[int, str] | None = None
    for bid in comparison["bids"]:
        enriched_bid = {
            **bid,
            "delivery_days": parse_days_from_text(bid.get("delivery_time")),
            "payment_days": parse_payment_term_days(bid.get("payment_term")) or 0
        }
        bid_breakdown.append(enriched_bid)
        currency_set.add(enriched_bid.get("currency") or "USD")
        comparison_key = (
            normalize_comparison_product_name(enriched_bid.get("product_name")),
            normalize_comparison_unit(enriched_bid.get("unit"))
        )
        product_groups.setdefault(comparison_key, []).append(enriched_bid)
        lowest_key = (float(enriched_bid["total_price"]), str(enriched_bid["supplier_name"]).lower())
        if lowest_price_key is None or lowest_key < lowest_price_key:
            lowest_price_key = lowest_key
            lowest_price_bid = enriched_bid
        delivery_key = (
            enriched_bid["delivery_days"] if enriched_bid["delivery_days"] is not None else 9999,
            str(enriched_bid["supplier_name"]).lower()
        )
        if fastest_delivery_key is None or delivery_key < fastest_delivery_key:
            fastest_delivery_key = delivery_key
            fastest_delivery_bid = enriched_bid
        payment_key = (int(enriched_bid["payment_days"]), str(enriched_bid["supplier_name"]).lower())
        if best_payment_key is None or payment_key > best_payment_key:
            best_payment_key = payment_key
            best_payment_bid = enriched_bid
    calculation_substeps["build_breakdown_ms"] = round((perf_counter() - build_breakdown_started_at) * 1000, 1)
    currencies = sorted(currency_set)

    group_products_started_at = perf_counter()
    grouped_products: list[dict[str, Any]] = []
    supplier_wins: dict[str, dict[str, Any]] = {}
    for comparison_key, offers in sorted(
        product_groups.items(),
        key=lambda item: item[0]
    ):
        product_name, unit = comparison_key
        sorted_offers = sorted(
            offers,
            key=lambda item: (
                item["total_price"],
                item["delivery_days"] if item["delivery_days"] is not None else 9999,
                -item["payment_days"],
                item["supplier_name"].lower()
            )
        )
        best_offer = sorted_offers[0]
        fastest_days = min(
            offer["delivery_days"] if offer["delivery_days"] is not None else 9999
            for offer in sorted_offers
        )
        best_payment_days = max(offer["payment_days"] for offer in sorted_offers)

        for offer in sorted_offers:
            badges: list[str] = []
            if offer is best_offer:
                badges.append("Best Price")
            if offer["delivery_days"] is not None and offer["delivery_days"] == fastest_days:
                badges.append("Fastest Delivery")
            if offer["payment_days"] == best_payment_days:
                badges.append("Best Payment")
            offer["badges"] = badges

        supplier_entry = supplier_wins.setdefault(best_offer["supplier_name"], {"wins": 0, "total_best_value": 0.0})
        supplier_entry["wins"] += 1
        supplier_entry["total_best_value"] += best_offer["total_price"]
        grouped_products.append({
            "product_name": best_offer["product_name"],
            "unit": best_offer["unit"],
            "offer_count": len(sorted_offers),
            "best_offer_supplier": best_offer["supplier_name"],
            "best_offer_value": best_offer["total_price"],
            "offers": sorted_offers
        })
    calculation_substeps["group_products_ms"] = round((perf_counter() - group_products_started_at) * 1000, 1)

    insights_started_at = perf_counter()
    recommended_supplier_name = min(
        supplier_wins.items(),
        key=lambda item: (-item[1]["wins"], item[1]["total_best_value"], item[0].lower())
    )[0]
    recommended_meta = supplier_wins[recommended_supplier_name]
    delivery_copy = (
        f"{fastest_delivery_bid['delivery_days']} days"
        if fastest_delivery_bid["delivery_days"] is not None
        else "Not provided"
    )
    payment_copy = best_payment_bid["payment_term"] or "Not provided"

    insights = [
        f"{lowest_price_bid['supplier_name']} is the lowest-price offer at {format_currency(lowest_price_bid['total_price'])}.",
        f"{fastest_delivery_bid['supplier_name']} is the fastest delivery option"
        f"{f' at {fastest_delivery_bid['delivery_days']} days' if fastest_delivery_bid['delivery_days'] is not None else ''}.",
        f"{best_payment_bid['supplier_name']} offers the strongest payment position with {best_payment_bid['payment_term'] or 'the most favorable term entered'}.",
        f"{recommended_supplier_name} is recommended because it wins {recommended_meta['wins']} product group"
        f"{'' if recommended_meta['wins'] == 1 else 's'} on best price."
    ]
    if len(currencies) > 1:
        insights.append("Mixed currencies were detected, so price ranking assumes the entered totals are directly comparable.")
    calculation_substeps["insights_ms"] = round((perf_counter() - insights_started_at) * 1000, 1)
    calculation_substeps["product_groups"] = len(grouped_products)
    calculation_substeps["currencies"] = len(currencies)
    calculation_substeps["total_ms"] = round((perf_counter() - calculation_started_at) * 1000, 1)
    log_perf_details("quote_compare.compare_calculation.substeps", **calculation_substeps)

    return {
        "summary": {
            "lowest_price_supplier": lowest_price_bid["supplier_name"],
            "lowest_price_value": lowest_price_bid["total_price"],
            "fastest_delivery_supplier": fastest_delivery_bid["supplier_name"],
            "fastest_delivery_value": delivery_copy,
            "best_payment_supplier": best_payment_bid["supplier_name"],
            "best_payment_value": payment_copy,
            "best_overall_supplier": recommended_supplier_name,
            "recommended_supplier": recommended_supplier_name,
            "recommended_reason": (
                f"Wins {recommended_meta['wins']} product group"
                f"{'' if recommended_meta['wins'] == 1 else 's'} on best price, "
                "with delivery and payment terms used as supporting context."
            )
        },
        "currencies": currencies,
        "insights": insights,
        "bids": bid_breakdown,
        "products": grouped_products
    }


def build_analysis_dataframe_from_quote_comparison(
    comparison: dict[str, Any],
    *,
    comparison_is_normalized: bool = False
) -> pd.DataFrame:
    normalized_comparison = comparison if comparison_is_normalized else normalize_quote_comparison_payload(comparison or {})
    validate_quote_comparison_payload(normalized_comparison)
    upload_id = normalized_comparison.get("upload_id") or str(uuid.uuid4())

    rows: list[dict[str, Any]] = []
    for bid in normalized_comparison["bids"]:
        rows.append({
            "upload_id": upload_id,
            "Product Name": bid.get("product_name", ""),
            "Supplier": bid.get("supplier_name", ""),
            "Unit": bid.get("unit", ""),
            "Quantity": bid.get("quantity", 0),
            "Unit Price": bid.get("unit_price", 0),
            "Date": bid.get("quote_date", ""),
            "Currency": bid.get("currency", ""),
            "Delivery Time": bid.get("delivery_time", ""),
            "Payment Terms": bid.get("payment_term", ""),
            "Valid Until": bid.get("valid_until", ""),
            "Notes": bid.get("notes", "")
        })

    dataframe_columns = [
        "upload_id",
        *REQUIRED_ANALYSIS_FIELDS,
        "Currency",
        "Delivery Time",
        "Payment Terms",
        "Valid Until",
        "Notes"
    ]
    return pd.DataFrame(rows, columns=dataframe_columns)


def append_analysis_history_rows(
    result_df: pd.DataFrame,
    *,
    upload_id: str,
    source_name: str,
    source_type: str,
    comparison_name: str,
    is_deduplicated: bool = False
) -> None:
    now = datetime.now(timezone.utc).isoformat()
    current_upload_df = result_df.copy() if is_deduplicated else deduplicate_analysis_frame(result_df)
    latest_date_series = current_upload_df["Date"].dropna() if "Date" in current_upload_df.columns else pd.Series(dtype="datetime64[ns]")
    store = {
        "uploads": [{
            "upload_id": upload_id,
            "source_name": str(source_name or "").strip() or "Compare Prices analysis",
            "comparison_name": str(comparison_name or "").strip() or str(source_name or "").strip() or "Compare Prices analysis",
            "source_type": str(source_type or "manual").strip() or "manual",
            "created_at": now,
            "updated_at": now,
            "row_count": int(len(current_upload_df.index)),
            "product_count": count_unique_products_by_name_unit(current_upload_df),
            "latest_date": latest_date_series.max().isoformat() if not latest_date_series.empty else ""
        }],
        "rows": serialize_dataframe_records(current_upload_df)
    }
    save_analysis_history_store(store)


def persist_quote_compare_analysis_results(
    comparison: dict[str, Any],
    *,
    source_name: str,
    upload_id: str | None = None,
    comparison_is_normalized: bool = False,
    perf_label_prefix: str | None = None
) -> pd.DataFrame:
    persist_started_at = perf_counter()
    persist_substeps: dict[str, Any] = {
        "comparison_already_normalized": comparison_is_normalized,
        "deduplicate_copy_frame": False,
        "assign_row_ids_copy_frame": False
    }

    normalize_started_at = perf_counter()
    normalized_comparison = comparison if comparison_is_normalized else normalize_quote_comparison_payload(comparison or {})
    persist_substeps["normalize_comparison_ms"] = round((perf_counter() - normalize_started_at) * 1000, 1)

    build_df_started_at = perf_counter()
    source_df = build_analysis_dataframe_from_quote_comparison(
        normalized_comparison,
        comparison_is_normalized=True
    )
    persist_substeps["build_analysis_dataframe_ms"] = round((perf_counter() - build_df_started_at) * 1000, 1)
    log_perf("quote_compare.build_analysis_dataframe", build_df_started_at)
    normalized_upload_id = str(upload_id or source_df["upload_id"].iloc[0]).strip()
    source_df["upload_id"] = normalized_upload_id
    persist_substeps["source_rows"] = int(len(source_df.index))
    persist_substeps["source_columns"] = int(len(source_df.columns))

    analysis_started_at = perf_counter()
    _, analyzed_df = analyze_dataframe(
        source_df,
        source_name=source_name,
        persist_latest_results=False,
        return_result_df=True
    )
    persist_substeps["analyze_dataframe_ms"] = round((perf_counter() - analysis_started_at) * 1000, 1)
    log_perf("quote_compare.analyze_dataframe", analysis_started_at)

    dedupe_started_at = perf_counter()
    result_df = deduplicate_analysis_frame(analyzed_df, keep_dedupe_key=True, copy_frame=False)
    persist_substeps["deduplicate_ms"] = round((perf_counter() - dedupe_started_at) * 1000, 1)
    log_perf("quote_compare.deduplicate_results", dedupe_started_at)
    if "upload_id" in result_df.columns:
        upload_id_normalize_started_at = perf_counter()
        result_df["upload_id"] = result_df["upload_id"].fillna("").astype(str)
        result_df.loc[:, "upload_id"] = normalized_upload_id
        persist_substeps["upload_id_normalize_ms"] = round((perf_counter() - upload_id_normalize_started_at) * 1000, 1)

    persist_storage_started_at = perf_counter()
    assign_ids_started_at = perf_counter()
    result_df = assign_analysis_row_ids(result_df, upload_id=normalized_upload_id, copy_frame=False)
    persist_substeps["assign_row_ids_ms"] = round((perf_counter() - assign_ids_started_at) * 1000, 1)
    log_perf("quote_compare.assign_row_ids", assign_ids_started_at)

    save_frame_started_at = perf_counter()
    save_current_upload_analysis_frame(result_df)
    persist_substeps["save_latest_results_ms"] = round((perf_counter() - save_frame_started_at) * 1000, 1)
    latest_results_path = get_current_latest_results_path()
    persist_substeps["saved_file_bytes"] = latest_results_path.stat().st_size if latest_results_path.exists() else 0
    persist_substeps["result_rows"] = int(len(result_df.index))
    persist_substeps["result_columns"] = int(len(result_df.columns))
    persist_substeps["total_ms"] = round((perf_counter() - persist_started_at) * 1000, 1)
    log_perf("quote_compare.save_latest_results", save_frame_started_at)
    log_perf("quote_compare.persist_active_analysis", persist_storage_started_at)
    log_perf("quote_compare.total_analysis_pipeline", persist_started_at)
    if perf_label_prefix:
        log_perf_details(f"{perf_label_prefix}.persist_analysis.substeps", **persist_substeps)
    return result_df


def get_latest_quote_compare_analysis_session(user_id: int | str | None = None) -> dict[str, Any] | None:
    validated_payload = validate_quote_compare_active_session(load_quote_compare_active_session(None, user_id=user_id))
    if validated_payload and isinstance(validated_payload.get("comparison"), dict):
        return validated_payload
    return None


def get_latest_saved_quote_compare_analysis() -> dict[str, Any] | None:
    store = load_quote_comparisons_store()
    comparisons = [item for item in store.get("comparisons", []) if isinstance(item, dict) and item.get("bids")]
    if not comparisons:
        return None

    def get_sort_key(item: dict[str, Any]) -> tuple[datetime, str]:
        raw_timestamp = item.get("updated_at") or item.get("created_at") or ""
        try:
            timestamp = datetime.fromisoformat(str(raw_timestamp).replace("Z", "+00:00"))
        except ValueError:
            timestamp = datetime.min.replace(tzinfo=timezone.utc)
        return (timestamp, str(item.get("name") or ""))

    return sorted(comparisons, key=get_sort_key, reverse=True)[0]


def restore_latest_quote_compare_analysis_results() -> pd.DataFrame | None:
    latest_session = get_latest_quote_compare_analysis_session()
    if latest_session and isinstance(latest_session.get("comparison"), dict):
        source_name = (
            str(latest_session.get("filename") or "").strip()
            or str(latest_session["comparison"].get("name") or "").strip()
            or "Compare Prices analysis"
        )
        return persist_quote_compare_analysis_results(latest_session["comparison"], source_name=source_name)

    latest_comparison = get_latest_saved_quote_compare_analysis()
    if latest_comparison:
        source_name = str(latest_comparison.get("name") or "").strip() or "Compare Prices analysis"
        return persist_quote_compare_analysis_results(latest_comparison, source_name=source_name)

    return None


def has_recipe_analysis_source() -> bool:
    latest_results_path = get_current_latest_results_path()
    if latest_results_path.exists():
        cache_signature = get_recipe_analysis_cache_signature()
        if (
            cache_signature is not None
            and RECIPE_ANALYSIS_CACHE["signature"] == cache_signature
            and isinstance(RECIPE_ANALYSIS_CACHE["frame"], pd.DataFrame)
        ):
            return not RECIPE_ANALYSIS_CACHE["frame"].empty
        try:
            quick_probe = pd.read_csv(latest_results_path, nrows=1)
            if not quick_probe.empty:
                return True
        except Exception:
            logger.exception("Failed to inspect latest results while checking recipe analysis availability")
    return False


def get_path_cache_signature(path: Path) -> tuple[str, int, int] | None:
    if not path.exists():
        return None
    stat_result = path.stat()
    return (str(path), stat_result.st_mtime_ns, stat_result.st_size)


def get_latest_analysis_upload_meta_from_history() -> dict[str, Any] | None:
    seed_analysis_history_from_latest_results()
    store = load_analysis_history_store()
    uploads = [item for item in store.get("uploads", []) if isinstance(item, dict) and item.get("upload_id")]
    if not uploads:
        return None

    def sort_key(item: dict[str, Any]) -> tuple[str, str]:
        return (str(item.get("updated_at") or ""), str(item.get("upload_id") or ""))

    return sorted(uploads, key=sort_key, reverse=True)[0]


def get_latest_analysis_upload_meta(
    *,
    prepared_frame: pd.DataFrame | None = None,
    user_id: int | str | None = None,
    perf_label: str | None = None
) -> dict[str, Any] | None:
    started_at = perf_counter()
    substeps: dict[str, Any] = {
        "cache_hit": False,
        "cache_reason": "miss"
    }
    resolved_user_id = get_storage_user_id(user_id)
    current_upload_id = ""
    prepared_frame_to_use = prepared_frame if isinstance(prepared_frame, pd.DataFrame) else None
    if prepared_frame_to_use is not None and not prepared_frame_to_use.empty and "upload_id" in prepared_frame_to_use.columns:
        upload_id_started_at = perf_counter()
        upload_id_series = prepared_frame_to_use["upload_id"].dropna()
        if not upload_id_series.empty:
            current_upload_id = str(upload_id_series.iloc[0]).strip()
        substeps["prepared_frame_upload_id_ms"] = round((perf_counter() - upload_id_started_at) * 1000, 1)

    active_session_path = get_user_active_quote_compare_session_path(resolved_user_id)
    ensure_quote_comparisons_file(resolved_user_id)
    comparisons_path = get_user_quote_comparisons_path(resolved_user_id)
    ensure_analysis_history_file(resolved_user_id)
    analysis_history_path = get_user_analysis_history_path(resolved_user_id)
    cache_signature = (
        get_recipe_analysis_cache_signature(),
        current_upload_id,
        get_path_cache_signature(active_session_path),
        get_path_cache_signature(comparisons_path),
        get_path_cache_signature(analysis_history_path)
    )
    if LATEST_ANALYSIS_UPLOAD_META_CACHE["signature"] == cache_signature:
        substeps["cache_hit"] = True
        substeps["cache_reason"] = "signature_match"
        substeps["current_upload_id"] = current_upload_id
        substeps["total_ms"] = round((perf_counter() - started_at) * 1000, 1)
        if perf_label:
            log_perf_details(perf_label, **substeps)
        return LATEST_ANALYSIS_UPLOAD_META_CACHE["meta"]

    substeps["current_upload_id"] = current_upload_id
    active_session_started_at = perf_counter()
    latest_session = get_latest_quote_compare_analysis_session(user_id=resolved_user_id)
    substeps["active_session_lookup_ms"] = round((perf_counter() - active_session_started_at) * 1000, 1)
    if latest_session and isinstance(latest_session.get("comparison"), dict):
        comparison = latest_session["comparison"]
        session_upload_id = str(comparison.get("upload_id") or "").strip()
        if current_upload_id and session_upload_id == current_upload_id:
            source_name = (
                str(latest_session.get("filename") or "").strip()
                or str(comparison.get("name") or "").strip()
                or "Compare Prices analysis"
            )
            meta = {
                "upload_id": current_upload_id,
                "source_name": source_name
            }
            substeps["cache_reason"] = "active_session_match"
            substeps["resolved_via"] = "active_session"
            substeps["total_ms"] = round((perf_counter() - started_at) * 1000, 1)
            LATEST_ANALYSIS_UPLOAD_META_CACHE["signature"] = cache_signature
            LATEST_ANALYSIS_UPLOAD_META_CACHE["meta"] = meta
            if perf_label:
                log_perf_details(perf_label, **substeps)
            return meta

    saved_comparison_started_at = perf_counter()
    latest_comparison = get_latest_saved_quote_compare_analysis()
    substeps["saved_comparison_lookup_ms"] = round((perf_counter() - saved_comparison_started_at) * 1000, 1)
    if latest_comparison:
        comparison_upload_id = str(latest_comparison.get("upload_id") or "").strip()
        if current_upload_id and comparison_upload_id == current_upload_id:
            meta = {
                "upload_id": current_upload_id,
                "source_name": str(latest_comparison.get("name") or "").strip() or "Compare Prices analysis"
            }
            substeps["cache_reason"] = "saved_comparison_match"
            substeps["resolved_via"] = "saved_comparison"
            substeps["total_ms"] = round((perf_counter() - started_at) * 1000, 1)
            LATEST_ANALYSIS_UPLOAD_META_CACHE["signature"] = cache_signature
            LATEST_ANALYSIS_UPLOAD_META_CACHE["meta"] = meta
            if perf_label:
                log_perf_details(perf_label, **substeps)
            return meta

    history_started_at = perf_counter()
    meta = get_latest_analysis_upload_meta_from_history()
    substeps["history_lookup_ms"] = round((perf_counter() - history_started_at) * 1000, 1)
    substeps["cache_reason"] = "history_fallback"
    substeps["resolved_via"] = "analysis_history"
    substeps["total_ms"] = round((perf_counter() - started_at) * 1000, 1)
    LATEST_ANALYSIS_UPLOAD_META_CACHE["signature"] = cache_signature
    LATEST_ANALYSIS_UPLOAD_META_CACHE["meta"] = meta
    if perf_label:
        log_perf_details(perf_label, **substeps)
    return meta


def load_analysis_history_frame() -> pd.DataFrame:
    if not get_current_latest_results_path().exists():
        return pd.DataFrame()
    return load_current_upload_analysis_frame()


def filter_analysis_frame_by_scope(frame: pd.DataFrame, scope: str) -> pd.DataFrame:
    if normalize_analysis_scope(scope) == "demo":
        return build_demo_analysis_dataframe()
    return frame.copy()


def build_analysis_scope_summary(frame: pd.DataFrame, *, scope: str) -> dict[str, Any]:
    normalized_scope = normalize_analysis_scope(scope)
    latest_upload = None if normalized_scope == "demo" else get_latest_analysis_upload_meta()
    dated_rows = frame["Date"].dropna() if "Date" in frame.columns else pd.Series(dtype="datetime64[ns]")
    return {
        "scope": normalized_scope,
        "scope_label": get_analysis_scope_label(normalized_scope),
        "row_count": int(len(frame.index)),
        "product_count": count_unique_products_by_name_unit(frame),
        "product_name_count": count_unique_product_names(frame),
        "supplier_count": int(frame["Supplier"].nunique()) if "Supplier" in frame.columns and not frame.empty else 0,
        "current_upload_id": str((latest_upload or {}).get("upload_id") or (normalized_scope if normalized_scope == "demo" else "")).strip(),
        "current_upload_name": str((latest_upload or {}).get("source_name") or ("Demo Data" if normalized_scope == "demo" else "")).strip(),
        "date_range": {
            "start": dated_rows.min().strftime("%Y-%m-%d") if not dated_rows.empty else "",
            "end": dated_rows.max().strftime("%Y-%m-%d") if not dated_rows.empty else ""
        }
    }


def build_analysis_scope_summary_with_upload(
    frame: pd.DataFrame,
    *,
    latest_upload: dict[str, Any] | None,
    scope: str = "current_upload"
) -> dict[str, Any]:
    normalized_scope = normalize_analysis_scope(scope)
    dated_rows = frame["Date"].dropna() if "Date" in frame.columns else pd.Series(dtype="datetime64[ns]")
    return {
        "scope": normalized_scope,
        "scope_label": get_analysis_scope_label(normalized_scope),
        "row_count": int(len(frame.index)),
        "product_count": count_unique_products_by_name_unit(frame),
        "product_name_count": count_unique_product_names(frame),
        "supplier_count": int(frame["Supplier"].nunique()) if "Supplier" in frame.columns and not frame.empty else 0,
        "current_upload_id": str((latest_upload or {}).get("upload_id") or (normalized_scope if normalized_scope == "demo" else "")).strip(),
        "current_upload_name": str((latest_upload or {}).get("source_name") or ("Demo Data" if normalized_scope == "demo" else "")).strip(),
        "date_range": {
            "start": dated_rows.min().strftime("%Y-%m-%d") if not dated_rows.empty else "",
            "end": dated_rows.max().strftime("%Y-%m-%d") if not dated_rows.empty else ""
        }
    }


def build_analysis_scope_summary_from_comparison(
    comparison: dict[str, Any] | None,
    *,
    latest_upload: dict[str, Any] | None = None,
    scope: str = "current_upload"
) -> dict[str, Any]:
    normalized_scope = normalize_analysis_scope(scope)
    normalized_comparison = comparison if isinstance(comparison, dict) else {}
    bids = normalized_comparison.get("bids")
    normalized_bids = bids if isinstance(bids, list) else []
    product_keys: set[tuple[str, str]] = set()
    suppliers: set[str] = set()
    date_values: list[datetime] = []

    for bid in normalized_bids:
        if not isinstance(bid, dict):
            continue
        product_name = str(bid.get("product_name") or "").strip()
        unit = str(bid.get("unit") or "").strip()
        supplier_name = str(bid.get("supplier_name") or "").strip()
        if product_name:
            product_keys.add((
                normalize_comparison_product_name(product_name),
                normalize_comparison_unit(unit)
            ))
        if supplier_name:
            suppliers.add(supplier_name)
        raw_bid_date = str(bid.get("date") or "").strip()
        if raw_bid_date:
            try:
                date_values.append(datetime.fromisoformat(raw_bid_date.replace("Z", "+00:00")))
            except ValueError:
                pass

    sorted_dates = sorted(date_values)
    product_names = {
        str(bid.get("product_name") or "").strip()
        for bid in normalized_bids
        if isinstance(bid, dict) and str(bid.get("product_name") or "").strip()
    }
    return {
        "scope": normalized_scope,
        "scope_label": get_analysis_scope_label(normalized_scope),
        "row_count": len(normalized_bids),
        "product_count": len(product_keys),
        "product_name_count": len(product_names),
        "supplier_count": len(suppliers),
        "current_upload_id": str((latest_upload or {}).get("upload_id") or normalized_comparison.get("upload_id") or (normalized_scope if normalized_scope == "demo" else "")).strip(),
        "current_upload_name": str((latest_upload or {}).get("source_name") or normalized_comparison.get("name") or ("Demo Data" if normalized_scope == "demo" else "")).strip(),
        "date_range": {
            "start": sorted_dates[0].strftime("%Y-%m-%d") if sorted_dates else "",
            "end": sorted_dates[-1].strftime("%Y-%m-%d") if sorted_dates else ""
        }
    }


def normalize_recipe_analysis_frame(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        raise ValueError("No analyzed dataset is available yet. Please upload and analyze a file first.")

    frame = frame.copy()
    frame["Product Name"] = frame["Product Name"].fillna("").astype(str).str.strip()
    frame["Unit"] = frame["Unit"].fillna("").astype(str).str.strip()
    frame["Normalized Unit"] = frame["Unit"].map(normalize_recipe_unit_name)
    frame["Supplier"] = frame["Supplier"].fillna("").astype(str).str.strip()
    frame["Unit Price"] = pd.to_numeric(frame["Unit Price"], errors="coerce")
    frame["Average Price"] = pd.to_numeric(frame.get("Average Price"), errors="coerce")
    frame["Date"] = pd.to_datetime(frame.get("Date"), errors="coerce")
    frame = frame[frame["Product Name"].astype(bool)].copy()

    if frame.empty:
        raise ValueError("No analyzed dataset is available yet. Please upload and analyze a file first.")

    return frame


def get_recipe_analysis_cache_signature() -> tuple[str, int, int] | None:
    latest_results_path = get_current_latest_results_path()
    if not latest_results_path.exists():
        return None
    return (
        str(latest_results_path),
        latest_results_path.stat().st_mtime_ns,
        latest_results_path.stat().st_size
    )


def get_recipes_store_cache_signature(user_id: int | str | None = None) -> tuple[str, int, int]:
    resolved_user_id = get_storage_user_id(user_id)
    recipes_path = get_user_recipes_path(resolved_user_id)
    ensure_recipes_file(resolved_user_id)
    return (
        str(recipes_path),
        recipes_path.stat().st_mtime_ns,
        recipes_path.stat().st_size
    )


def build_recipe_pricing_lookup(frame: pd.DataFrame) -> dict[tuple[str, str], dict[str, Any]]:
    pricing_lookup_started_at = perf_counter()
    pricing_lookup_substeps: dict[str, Any] = {
        "cache_hit": False,
        "cache_reason": "build_required",
        "frame_rows": int(len(frame.index))
    }
    pricing_lookup: dict[tuple[str, str], dict[str, Any]] = {}
    if frame.empty:
        pricing_lookup_substeps["total_ms"] = round((perf_counter() - pricing_lookup_started_at) * 1000, 1)
        log_perf_details("recipes.bootstrap.pricing_lookup.substeps", **pricing_lookup_substeps)
        return pricing_lookup

    group_unit_column = "Normalized Unit" if "Normalized Unit" in frame.columns else "Unit"
    group_columns = ["Product Name", group_unit_column]
    subset_started_at = perf_counter()
    latest_row_frame = frame.loc[:, group_columns + ["Date", "Unit Price", "Supplier"]]
    aggregate_frame = frame.loc[:, group_columns + ["Average Price", "Unit Price"]]
    pricing_lookup_substeps["select_columns_ms"] = round((perf_counter() - subset_started_at) * 1000, 1)

    sort_started_at = perf_counter()
    sorted_frame = latest_row_frame.sort_values(group_columns + ["Date"], ascending=[True, True, False], na_position="last")
    pricing_lookup_substeps["sort_once_ms"] = round((perf_counter() - sort_started_at) * 1000, 1)

    latest_rows_started_at = perf_counter()
    latest_rows = sorted_frame.drop_duplicates(subset=group_columns, keep="first")
    pricing_lookup_substeps["latest_rows_ms"] = round((perf_counter() - latest_rows_started_at) * 1000, 1)

    aggregate_started_at = perf_counter()
    grouped_means = aggregate_frame.groupby(group_columns, sort=False, dropna=False, as_index=False).agg(
        average_price_from_average=("Average Price", "mean"),
        average_price_from_unit=("Unit Price", "mean")
    )
    grouped_means["average_price"] = grouped_means["average_price_from_average"].fillna(grouped_means["average_price_from_unit"])
    pricing_lookup_substeps["aggregate_ms"] = round((perf_counter() - aggregate_started_at) * 1000, 1)

    join_started_at = perf_counter()
    pricing_source = latest_rows.loc[:, group_columns + ["Unit Price", "Supplier"]].merge(
        grouped_means.loc[:, group_columns + ["average_price"]],
        on=group_columns,
        how="left",
        sort=False
    )
    pricing_lookup_substeps["join_ms"] = round((perf_counter() - join_started_at) * 1000, 1)

    build_dict_started_at = perf_counter()
    product_values = pricing_source["Product Name"].tolist()
    unit_values = pricing_source[group_unit_column].tolist()
    latest_price_values = pricing_source["Unit Price"].tolist()
    supplier_values = pricing_source["Supplier"].tolist()
    average_price_values = pricing_source["average_price"].tolist()
    for product_name, unit, latest_price, supplier, average_price in zip(
        product_values,
        unit_values,
        latest_price_values,
        supplier_values,
        average_price_values
    ):
        pricing_lookup[(str(product_name), str(unit))] = {
            "latest_price": float(latest_price),
            "latest_supplier": str(supplier or "").strip() or "N/A",
            "average_price": float(average_price)
        }
    pricing_lookup_substeps["build_dict_ms"] = round((perf_counter() - build_dict_started_at) * 1000, 1)
    pricing_lookup_substeps["group_count"] = len(pricing_lookup)
    pricing_lookup_substeps["total_ms"] = round((perf_counter() - pricing_lookup_started_at) * 1000, 1)
    log_perf_details("recipes.bootstrap.pricing_lookup.substeps", **pricing_lookup_substeps)
    return pricing_lookup


def load_recipe_analysis_bundle(*, scope: str = "current_upload", session_id: str | None = None) -> dict[str, Any]:
    normalized_scope = normalize_analysis_scope(scope)
    if normalized_scope == "demo":
        frame = normalize_recipe_analysis_frame(build_demo_analysis_dataframe())
        demo_recipe_defaults = with_demo_recipe_product_fallbacks(
            frame,
            build_recipe_product_defaults(build_demo_recipes())
        )
        return {
            "frame": frame,
            "product_catalog": build_recipe_product_catalog(frame, recipe_defaults=demo_recipe_defaults),
            "pricing_lookup": build_recipe_pricing_lookup(frame)
        }

    cache_signature = get_recipe_analysis_cache_signature()
    if cache_signature is None:
        raise ValueError("No analyzed dataset is available yet. Please upload and analyze a file first.")

    analysis_load_metrics = {
        "scope": scope,
        "cache_hit": False,
        "cache_reason": "miss"
    }
    if RECIPE_ANALYSIS_CACHE["signature"] == cache_signature and isinstance(RECIPE_ANALYSIS_CACHE["frame"], pd.DataFrame):
        analysis_load_metrics["cache_hit"] = True
        analysis_load_metrics["cache_reason"] = "signature_match"
        log_perf_details(
            "recipes.bootstrap.pricing_lookup.substeps",
            cache_hit=True,
            cache_reason="analysis_bundle_cache_hit",
            frame_rows=int(len(RECIPE_ANALYSIS_CACHE["frame"].index)),
            group_count=len(RECIPE_ANALYSIS_CACHE["pricing_lookup"] or {}),
            total_ms=0.0
        )
        log_perf_details("recipes.bootstrap.analysis_load.substeps", **analysis_load_metrics)
        return {
            "frame": RECIPE_ANALYSIS_CACHE["frame"],
            "product_catalog": RECIPE_ANALYSIS_CACHE["product_catalog"] or [],
            "pricing_lookup": RECIPE_ANALYSIS_CACHE["pricing_lookup"] or {}
        }

    frame_load_started_at = perf_counter()
    current_upload_frame = load_current_upload_analysis_frame()
    analysis_load_metrics["load_current_upload_frame_ms"] = round((perf_counter() - frame_load_started_at) * 1000, 1)
    normalize_started_at = perf_counter()
    frame = normalize_recipe_analysis_frame(current_upload_frame)
    analysis_load_metrics["normalize_recipe_frame_ms"] = round((perf_counter() - normalize_started_at) * 1000, 1)
    product_catalog_started_at = perf_counter()
    product_catalog = build_recipe_product_catalog(frame)
    analysis_load_metrics["product_catalog_ms"] = round((perf_counter() - product_catalog_started_at) * 1000, 1)
    pricing_lookup_started_at = perf_counter()
    pricing_lookup = build_recipe_pricing_lookup(frame)
    analysis_load_metrics["pricing_lookup_ms"] = round((perf_counter() - pricing_lookup_started_at) * 1000, 1)
    RECIPE_ANALYSIS_CACHE["signature"] = cache_signature
    RECIPE_ANALYSIS_CACHE["frame"] = frame
    RECIPE_ANALYSIS_CACHE["product_catalog"] = product_catalog
    RECIPE_ANALYSIS_CACHE["pricing_lookup"] = pricing_lookup
    analysis_load_metrics["frame_rows"] = int(len(frame.index))
    analysis_load_metrics["product_count"] = len(product_catalog)
    analysis_load_metrics["pricing_lookup_entries"] = len(pricing_lookup)
    log_perf_details("recipes.bootstrap.analysis_load.substeps", **analysis_load_metrics)
    return {
        "frame": frame,
        "product_catalog": product_catalog,
        "pricing_lookup": pricing_lookup
    }


def load_recipe_analysis_dataframe(*, scope: str = "current_upload", session_id: str | None = None) -> pd.DataFrame:
    if normalize_analysis_scope(scope) == "demo":
        return load_recipe_analysis_bundle(scope=scope, session_id=session_id)["frame"]
    if get_current_latest_results_path().exists():
        try:
            return load_recipe_analysis_bundle(scope=scope, session_id=session_id)["frame"]
        except ValueError:
            pass
        except Exception:
            logger.exception("Failed to load latest results for Recipes")

    raise ValueError("No analyzed dataset is available yet. Please upload and analyze a file first.")


def build_recipe_product_catalog(
    frame: pd.DataFrame,
    recipe_defaults: dict[str, dict[str, Any]] | None = None
) -> list[dict[str, Any]]:
    catalog: list[dict[str, Any]] = []
    for product_name, group in frame.groupby("Product Name", sort=True):
        units = sorted({
            str(unit)
            for unit in group.get("Normalized Unit", group["Unit"]).fillna("").tolist()
            if str(unit)
        })
        product_defaults = (recipe_defaults or {}).get(str(product_name), {})
        purchase_unit = normalize_recipe_unit_name(str(product_defaults.get("purchase_unit") or (units[0] if units else "")).strip())
        purchase_base_unit = normalize_recipe_unit_name(str(product_defaults.get("purchase_base_unit") or "").strip())
        preferred_usage_unit = normalize_recipe_unit_name(str(product_defaults.get("preferred_usage_unit") or purchase_base_unit or purchase_unit).strip())
        purchase_size_value = product_defaults.get("purchase_size")
        catalog.append({
            "product_name": str(product_name),
            "units": units,
            "purchase_unit": purchase_unit,
            "purchase_base_unit": purchase_base_unit,
            "purchase_size": purchase_size_value if purchase_size_value is not None else "",
            "preferred_usage_unit": preferred_usage_unit
        })
    return catalog


RECIPE_UNIT_ALIASES = {
    "g": "g",
    "gram": "g",
    "grams": "g",
    "kg": "kg",
    "kilogram": "kg",
    "kilograms": "kg",
    "oz": "oz",
    "ounce": "oz",
    "ounces": "oz",
    "lb": "lb",
    "lbs": "lb",
    "pound": "lb",
    "pounds": "lb",
    "ml": "ml",
    "milliliter": "ml",
    "milliliters": "ml",
    "millilitre": "ml",
    "millilitres": "ml",
    "fl oz": "fl oz",
    "fl. oz": "fl oz",
    "floz": "fl oz",
    "fluid ounce": "fl oz",
    "fluid ounces": "fl oz",
    "l": "l",
    "lt": "l",
    "liter": "l",
    "liters": "l",
    "litre": "l",
    "litres": "l",
    "each": "each",
    "ea": "each",
    "piece": "each",
    "pieces": "each",
    "pc": "each",
    "pcs": "each",
    "portion": "portion",
    "portions": "portion",
    "package": "pack",
    "packages": "pack",
    "pack": "pack",
    "packs": "pack",
    "box": "box",
    "boxes": "box",
    "case": "case",
    "cases": "case",
    "carton": "carton",
    "cartons": "carton",
    "bottle": "bottle",
    "bottles": "bottle",
    "can": "can",
    "cans": "can",
    "bag": "bag",
    "bags": "bag",
    "jar": "jar",
    "jars": "jar"
}

RECIPE_UNIT_FACTORS = {
    "g": ("weight", 1.0),
    "oz": ("weight", 28.3495),
    "kg": ("weight", 1000.0),
    "lb": ("weight", 453.592),
    "ml": ("volume", 1.0),
    "fl oz": ("volume", 29.5735),
    "l": ("volume", 1000.0),
    "each": ("count", 1.0),
    "portion": ("count", 1.0)
}

RECIPE_UNIT_CATEGORIES = {
    "g": "weight",
    "kg": "weight",
    "oz": "weight",
    "lb": "weight",
    "ml": "volume",
    "l": "volume",
    "fl oz": "volume",
    "each": "count",
    "portion": "count",
    "pack": "package",
    "box": "package",
    "case": "package",
    "carton": "package",
    "bottle": "package",
    "can": "package",
    "bag": "package",
    "jar": "package"
}

RECIPE_BASE_UNITS = {
    "weight": "g",
    "volume": "ml",
    "count": "each"
}


def normalize_recipe_unit_name(unit: str) -> str:
    normalized_unit = str(unit or "").strip()
    return RECIPE_UNIT_ALIASES.get(normalized_unit.lower(), normalized_unit)


def get_recipe_unit_category(unit: str) -> str:
    return RECIPE_UNIT_CATEGORIES.get(normalize_recipe_unit_name(unit), "")


def resolve_recipe_purchase_base_unit(
    purchase_unit: str,
    usage_unit: str,
    purchase_base_unit: str | None = None
) -> str:
    normalized_purchase_unit = normalize_recipe_unit_name(purchase_unit)
    normalized_usage_unit = normalize_recipe_unit_name(usage_unit)
    normalized_base_unit = normalize_recipe_unit_name(purchase_base_unit or "")
    purchase_category = get_recipe_unit_category(normalized_purchase_unit)
    base_category = get_recipe_unit_category(normalized_base_unit)

    if purchase_category == "package":
        if base_category and base_category != "package":
            return RECIPE_BASE_UNITS.get(base_category, normalized_base_unit)
        usage_category = get_recipe_unit_category(normalized_usage_unit)
        if usage_category and usage_category != "package":
            return RECIPE_BASE_UNITS.get(usage_category, normalized_usage_unit)
        return "each"

    if normalized_base_unit and base_category and base_category != "package":
        return normalized_base_unit
    return RECIPE_BASE_UNITS.get(purchase_category, normalized_purchase_unit)


def infer_recipe_purchase_size(
    purchase_unit: str,
    usage_unit: str,
    purchase_base_unit: str | None = None
) -> float:
    normalized_purchase_unit = normalize_recipe_unit_name(purchase_unit)
    normalized_usage_unit = normalize_recipe_unit_name(usage_unit)
    resolved_base_unit = resolve_recipe_purchase_base_unit(purchase_unit, usage_unit, purchase_base_unit)
    if not normalized_purchase_unit or not normalized_usage_unit:
        return 1.0
    if normalized_purchase_unit == resolved_base_unit:
        return 1.0

    if get_recipe_unit_category(normalized_purchase_unit) == "package":
        return 0.0

    purchase_meta = RECIPE_UNIT_FACTORS.get(normalized_purchase_unit)
    base_meta = RECIPE_UNIT_FACTORS.get(resolved_base_unit)
    if purchase_meta and base_meta and purchase_meta[0] == base_meta[0] and base_meta[1] > 0:
        return purchase_meta[1] / base_meta[1]

    return 0.0


def convert_recipe_quantity_to_base(quantity: float, source_unit: str, base_unit: str) -> float:
    normalized_source_unit = normalize_recipe_unit_name(source_unit)
    normalized_base_unit = normalize_recipe_unit_name(base_unit)
    if normalized_source_unit == normalized_base_unit:
        return quantity

    source_meta = RECIPE_UNIT_FACTORS.get(normalized_source_unit)
    base_meta = RECIPE_UNIT_FACTORS.get(normalized_base_unit)
    if not source_meta or not base_meta or source_meta[0] != base_meta[0] or base_meta[1] <= 0:
        raise ValueError("Selected unit type does not match product type.")

    return quantity * (source_meta[1] / base_meta[1])


def resolve_recipe_usage_ratio(
    quantity: float,
    usage_unit: str,
    purchase_unit: str,
    purchase_size: float,
    purchase_base_unit: str | None = None
) -> tuple[float, float, str]:
    normalized_usage_unit = normalize_recipe_unit_name(usage_unit)
    normalized_purchase_unit = normalize_recipe_unit_name(purchase_unit)
    resolved_base_unit = resolve_recipe_purchase_base_unit(
        normalized_purchase_unit,
        normalized_usage_unit,
        purchase_base_unit
    )
    purchase_category = get_recipe_unit_category(normalized_purchase_unit)
    usage_category = get_recipe_unit_category(normalized_usage_unit)
    base_category = get_recipe_unit_category(resolved_base_unit)

    if (
        not normalized_usage_unit
        or not normalized_purchase_unit
        or not resolved_base_unit
        or usage_category == "package"
        or base_category == "package"
        or (purchase_category == "package" and usage_category != base_category)
        or (purchase_category and purchase_category != "package" and purchase_category != usage_category)
        or (not purchase_category and usage_category != base_category)
    ):
        raise ValueError("Selected unit type does not match product type.")

    effective_purchase_size = purchase_size
    if effective_purchase_size <= 0:
        effective_purchase_size = infer_recipe_purchase_size(
            normalized_purchase_unit,
            normalized_usage_unit,
            resolved_base_unit
        )
    if effective_purchase_size <= 0:
        raise ValueError("Each ingredient must include a valid conversion basis.")

    usage_quantity_in_base_unit = convert_recipe_quantity_to_base(
        quantity,
        normalized_usage_unit,
        resolved_base_unit
    )
    purchase_ratio = usage_quantity_in_base_unit / effective_purchase_size
    return purchase_ratio, usage_quantity_in_base_unit, resolved_base_unit


def normalize_recipe_payload(payload: dict[str, Any]) -> dict[str, Any]:
    normalized_ingredients: list[dict[str, Any]] = []
    for ingredient in payload.get("ingredients", []):
        product_name = str(ingredient.get("product_name", "")).strip()
        unit = normalize_recipe_unit_name(ingredient.get("unit", ""))
        purchase_unit = normalize_recipe_unit_name(ingredient.get("purchase_unit") or ingredient.get("unit", ""))
        purchase_base_unit = resolve_recipe_purchase_base_unit(
            purchase_unit,
            unit,
            ingredient.get("purchase_base_unit") or ingredient.get("conversion_unit")
        )
        if get_recipe_unit_category(unit) == "package":
            unit = purchase_base_unit
        quantity_raw = normalize_request_value(ingredient.get("quantity", 0))
        purchase_size_raw = normalize_request_value(ingredient.get("purchase_size", 0))
        try:
            quantity = float(quantity_raw)
        except (TypeError, ValueError):
            quantity = 0.0
        try:
            purchase_size = float(purchase_size_raw)
        except (TypeError, ValueError):
            purchase_size = 0.0
        if purchase_size <= 0:
            purchase_size = infer_recipe_purchase_size(purchase_unit, unit, purchase_base_unit)
        if not product_name and not unit and not purchase_unit and quantity <= 0:
            continue
        normalized_ingredients.append({
            "product_name": product_name,
            "quantity": quantity,
            "unit": unit,
            "purchase_unit": purchase_unit,
            "purchase_size": purchase_size,
            "purchase_base_unit": purchase_base_unit
        })

    pricing_goal_value_raw = normalize_request_value(payload.get("pricing_goal_value"))
    target_food_cost_raw = normalize_request_value(payload.get("target_food_cost_pct", 0))

    return {
        "recipe_id": str(payload.get("recipe_id") or "").strip() or None,
        "name": str(payload.get("name", "")).strip(),
        "yield_portions": float(normalize_request_value(payload.get("yield_portions", 0)) or 0),
        "pricing_mode": str(payload.get("pricing_mode", "")).strip(),
        "ingredients": normalized_ingredients,
        "selling_price": float(normalize_request_value(payload.get("selling_price", 0)) or 0),
        "pricing_goal_type": str(payload.get("pricing_goal_type") or "food_cost_pct").strip() or "food_cost_pct",
        "pricing_goal_value": (
            float(pricing_goal_value_raw)
            if pricing_goal_value_raw is not None
            else (
                float(target_food_cost_raw)
                if "pricing_goal_value" not in payload and target_food_cost_raw is not None
                else None
            )
        ),
        "target_food_cost_pct": float(target_food_cost_raw or 0),
        "total_recipe_cost": float(normalize_request_value(payload.get("total_recipe_cost", 0)) or 0),
        "cost_per_portion": float(normalize_request_value(payload.get("cost_per_portion", 0)) or 0),
        "gross_profit": float(normalize_request_value(payload.get("gross_profit", 0)) or 0),
        "gross_margin_pct": float(normalize_request_value(payload.get("gross_margin_pct", 0)) or 0),
        "food_cost_pct": float(normalize_request_value(payload.get("food_cost_pct", 0)) or 0),
        "suggested_selling_price": float(normalize_request_value(payload.get("suggested_selling_price", 0)) or 0)
    }


def validate_recipe_payload(recipe: dict[str, Any]) -> None:
    if not recipe["name"]:
        raise ValueError("Enter a recipe name before calculating or saving.")
    if recipe["yield_portions"] <= 0:
        raise ValueError("Yield must be greater than zero.")
    if recipe["pricing_mode"] not in RECIPE_PRICING_MODES:
        raise ValueError("Choose a pricing mode before calculating the recipe.")
    if not recipe["ingredients"]:
        raise ValueError("Add at least one ingredient before calculating the recipe.")

    for ingredient in recipe["ingredients"]:
        if (
            not ingredient["product_name"]
            or not ingredient["unit"]
            or not ingredient["purchase_unit"]
            or not ingredient.get("purchase_base_unit")
            or ingredient["quantity"] <= 0
            or ingredient["purchase_size"] <= 0
        ):
            raise ValueError("Each ingredient must include a product, usage quantity, recipe unit, purchase unit, and valid conversion basis.")

        purchase_category = get_recipe_unit_category(ingredient["purchase_unit"])
        usage_category = get_recipe_unit_category(ingredient["unit"])
        base_category = get_recipe_unit_category(ingredient.get("purchase_base_unit", ""))
        if (
            usage_category == "package"
            or base_category == "package"
            or (purchase_category == "package" and usage_category != base_category)
            or (purchase_category and purchase_category != "package" and purchase_category != usage_category)
            or (not purchase_category and usage_category != base_category)
        ):
            raise ValueError("Selected unit type does not match product type.")


def resolve_recipe_price(filtered_rows: pd.DataFrame, pricing_mode: str) -> tuple[float, str]:
    if pricing_mode == "latest_price":
        sorted_rows = filtered_rows.sort_values("Date", ascending=False, na_position="last")
        latest_row = sorted_rows.iloc[0]
        return float(latest_row["Unit Price"]), "Latest Price"
    if pricing_mode == "average_price":
        average_series = filtered_rows["Average Price"].dropna()
        if not average_series.empty:
            return float(average_series.mean()), "Average Price"
        return float(filtered_rows["Unit Price"].mean()), "Average Price"
    raise ValueError("Unsupported pricing mode selected.")


def calculate_recipe_cost(
    recipe: dict[str, Any],
    frame: pd.DataFrame,
    *,
    pricing_lookup: dict[tuple[str, str], dict[str, Any]] | None = None
) -> dict[str, Any]:
    validate_recipe_payload(recipe)
    breakdown: list[dict[str, Any]] = []

    for ingredient in recipe["ingredients"]:
        product_name = ingredient["product_name"]
        purchase_unit = ingredient["purchase_unit"]
        product_rows = frame[frame["Product Name"] == product_name]
        if product_rows.empty:
            raise ValueError(f"'{product_name}' was not found in the analyzed purchase dataset.")

        pricing_meta = (pricing_lookup or {}).get((product_name, purchase_unit))
        if pricing_meta is None:
            if "Normalized Unit" in product_rows.columns:
                unit_rows = product_rows[product_rows["Normalized Unit"] == purchase_unit]
            else:
                normalized_units = product_rows["Unit"].map(normalize_recipe_unit_name)
                unit_rows = product_rows[normalized_units == purchase_unit]
            if unit_rows.empty:
                raise ValueError(
                    f"No analyzed pricing was found for {product_name} with purchase unit '{purchase_unit}'."
                )
            average_series = unit_rows["Average Price"].dropna()
            sorted_rows = unit_rows.sort_values("Date", ascending=False, na_position="last")
            latest_row = sorted_rows.iloc[0]
            pricing_meta = {
                "latest_price": float(latest_row["Unit Price"]),
                "latest_supplier": str(latest_row["Supplier"] or "").strip() or "N/A",
                "average_price": float(average_series.mean()) if not average_series.empty else float(unit_rows["Unit Price"].mean())
            }

        if pricing_meta is None:
            raise ValueError(
                f"No analyzed pricing was found for {product_name} with purchase unit '{purchase_unit}'."
            )

        if recipe["pricing_mode"] == "latest_price":
            price_used = float(pricing_meta["latest_price"])
            pricing_label = "Latest Price"
        elif recipe["pricing_mode"] == "average_price":
            price_used = float(pricing_meta["average_price"])
            pricing_label = "Average Price"
        else:
            raise ValueError("Unsupported pricing mode selected.")
        purchase_ratio, usage_quantity_in_base_unit, purchase_base_unit = resolve_recipe_usage_ratio(
            float(ingredient["quantity"]),
            ingredient["unit"],
            purchase_unit,
            float(ingredient["purchase_size"] or 1),
            ingredient.get("purchase_base_unit")
        )
        ingredient_cost = round(price_used * purchase_ratio, 4)
        latest_supplier = pricing_meta.get("latest_supplier", "N/A")

        breakdown.append({
            "product_name": product_name,
            "quantity": round(float(ingredient["quantity"]), 4),
            "unit": ingredient["unit"],
            "purchase_unit": purchase_unit,
            "purchase_base_unit": purchase_base_unit,
            "purchase_size": round(float(ingredient["purchase_size"] or 1), 4),
            "purchase_ratio": round(float(purchase_ratio), 6),
            "usage_quantity_in_base_unit": round(float(usage_quantity_in_base_unit), 6),
            "price_used": round(price_used, 4),
            "pricing_label": pricing_label,
            "ingredient_cost": round(ingredient_cost, 4),
            "supplier": str(latest_supplier or "").strip() or "N/A"
        })

    total_recipe_cost = round(sum(item["ingredient_cost"] for item in breakdown), 2)
    cost_per_portion = round(total_recipe_cost / recipe["yield_portions"], 2) if recipe["yield_portions"] else 0.0
    main_cost_driver = max(breakdown, key=lambda item: item["ingredient_cost"], default=None)

    return {
        "pricing_mode": recipe["pricing_mode"],
        "pricing_mode_label": RECIPE_PRICING_MODES[recipe["pricing_mode"]],
        "total_recipe_cost": total_recipe_cost,
        "cost_per_portion": cost_per_portion,
        "main_cost_driver": main_cost_driver,
        "ingredient_breakdown": breakdown
    }


def calculate_recipe_pricing_metrics(
    recipe: dict[str, Any],
    calculation: dict[str, Any]
) -> dict[str, Any]:
    cost_per_portion = float(calculation.get("cost_per_portion") or 0)
    total_recipe_cost = float(calculation.get("total_recipe_cost") or 0)
    selling_price = float(recipe.get("selling_price") or 0)
    pricing_goal_type = str(recipe.get("pricing_goal_type") or "food_cost_pct").strip()
    if pricing_goal_type not in RECIPE_PRICING_GOAL_TYPES:
        pricing_goal_type = "food_cost_pct"
    pricing_goal_value = (
        float(recipe.get("pricing_goal_value"))
        if recipe.get("pricing_goal_value") is not None
        else None
    )
    target_food_cost_pct = (
        float(pricing_goal_value or 0)
        if pricing_goal_type == "food_cost_pct"
        else float(recipe.get("target_food_cost_pct") or 0)
    )
    gross_profit = selling_price - cost_per_portion if selling_price > 0 else 0.0
    gross_margin_pct = ((gross_profit / selling_price) * 100) if selling_price > 0 else 0.0
    food_cost_pct = ((cost_per_portion / selling_price) * 100) if selling_price > 0 else 0.0
    suggested_selling_price = 0.0
    if pricing_goal_value is not None and pricing_goal_value >= 0 and cost_per_portion > 0:
        goal_rate = pricing_goal_value / 100
        if pricing_goal_type == "food_cost_pct" and goal_rate > 0:
            suggested_selling_price = cost_per_portion / goal_rate
        elif pricing_goal_type == "gross_margin_pct" and 0 <= goal_rate < 1:
            suggested_selling_price = cost_per_portion / (1 - goal_rate)
        elif pricing_goal_type == "markup_pct":
            suggested_selling_price = cost_per_portion * (1 + goal_rate)
    return {
        "selling_price": round(selling_price, 2),
        "pricing_goal_type": pricing_goal_type,
        "pricing_goal_value": round(pricing_goal_value, 2) if pricing_goal_value is not None else None,
        "target_food_cost_pct": round(target_food_cost_pct, 2),
        "total_recipe_cost": round(total_recipe_cost, 2),
        "cost_per_portion": round(cost_per_portion, 2),
        "gross_profit": round(gross_profit, 2),
        "gross_margin_pct": round(gross_margin_pct, 2),
        "food_cost_pct": round(food_cost_pct, 2),
        "suggested_selling_price": round(suggested_selling_price, 2)
    }


def enrich_saved_recipes_with_costs(
    recipes: list[dict[str, Any]],
    frame: pd.DataFrame,
    *,
    pricing_lookup: dict[tuple[str, str], dict[str, Any]] | None = None
) -> list[dict[str, Any]]:
    enriched_recipes: list[dict[str, Any]] = []
    for recipe in recipes or []:
        enriched_recipe = {**recipe}
        try:
            normalized_recipe = normalize_recipe_payload(recipe)
            calculation = calculate_recipe_cost(normalized_recipe, frame, pricing_lookup=pricing_lookup)
            if "total_recipe_cost" not in enriched_recipe:
                enriched_recipe["total_recipe_cost"] = calculation["total_recipe_cost"]
            if "cost_per_portion" not in enriched_recipe:
                enriched_recipe["cost_per_portion"] = calculation["cost_per_portion"]
        except ValueError:
            enriched_recipe["total_recipe_cost"] = float(recipe.get("total_recipe_cost") or 0)
            enriched_recipe["cost_per_portion"] = float(recipe.get("cost_per_portion") or 0)
        enriched_recipes.append(enriched_recipe)
    return enriched_recipes


def load_recipe_store_for_scope(
    *,
    scope: str = "current_upload",
    session_id: str | None = None,
    user_id: int | str | None = None
) -> dict[str, Any]:
    if normalize_analysis_scope(scope) == "demo":
        return load_demo_recipe_store(str(session_id or "").strip())
    return load_recipes_store(user_id)


def save_recipe_store_for_scope(
    store: dict[str, Any],
    *,
    scope: str = "current_upload",
    session_id: str | None = None,
    user_id: int | str | None = None
) -> None:
    if normalize_analysis_scope(scope) == "demo":
        save_demo_recipe_store(str(session_id or "").strip(), store)
        return
    save_recipes_store(store, user_id)


def get_saved_recipe_by_id(
    recipe_id: str,
    user_id: int | str | None = None,
    *,
    scope: str = "current_upload",
    session_id: str | None = None
) -> dict[str, Any] | None:
    normalized_recipe_id = str(recipe_id or "").strip()
    if not normalized_recipe_id:
        return None
    recipes = load_recipe_store_for_scope(scope=scope, session_id=session_id, user_id=user_id).get("recipes", [])
    return next((recipe for recipe in recipes if str(recipe.get("recipe_id") or "").strip() == normalized_recipe_id), None)


def get_saved_recipe_export_metrics(recipe: dict[str, Any]) -> dict[str, float | None]:
    yield_portions = float(recipe.get("yield_portions") or 0)
    selling_price_total = float(recipe.get("selling_price") or 0)
    selling_price_per_portion = (selling_price_total / yield_portions) if selling_price_total > 0 and yield_portions > 0 else None
    gross_profit_per_portion = float(recipe.get("gross_profit") or 0) if selling_price_total > 0 else None
    gross_profit_total = (gross_profit_per_portion * yield_portions) if gross_profit_per_portion is not None and yield_portions > 0 else None
    gross_margin_pct = float(recipe.get("gross_margin_pct") or 0) if selling_price_total > 0 else None
    food_cost_pct = float(recipe.get("food_cost_pct") or 0) if selling_price_total > 0 else None
    suggested_selling_price_per_portion = float(recipe.get("suggested_selling_price") or 0) or None
    suggested_selling_price_total = (
        suggested_selling_price_per_portion * yield_portions
        if suggested_selling_price_per_portion is not None and yield_portions > 0
        else None
    )
    pricing_goal_value = recipe.get("pricing_goal_value")
    if pricing_goal_value is None:
        pricing_goal_value = recipe.get("target_food_cost_pct")
    return {
        "selling_price_total": round(selling_price_total, 2),
        "selling_price_per_portion": round(selling_price_per_portion, 2) if selling_price_per_portion is not None else None,
        "gross_profit_per_portion": round(gross_profit_per_portion, 2) if gross_profit_per_portion is not None else None,
        "gross_profit_total": round(gross_profit_total, 2) if gross_profit_total is not None else None,
        "gross_margin_pct": round(gross_margin_pct, 2) if gross_margin_pct is not None else None,
        "food_cost_pct": round(food_cost_pct, 2) if food_cost_pct is not None else None,
        "suggested_selling_price_per_portion": round(suggested_selling_price_per_portion, 2) if suggested_selling_price_per_portion is not None else None,
        "suggested_selling_price_total": round(suggested_selling_price_total, 2) if suggested_selling_price_total is not None else None,
        "pricing_goal_value": round(float(pricing_goal_value), 2) if pricing_goal_value is not None else None
    }


def build_saved_recipe_export_context(
    recipe: dict[str, Any],
    *,
    scope: str = "current_upload",
    session_id: str | None = None
) -> dict[str, Any]:
    normalized_recipe = normalize_recipe_payload(recipe)
    ingredient_breakdown: list[dict[str, Any]] = []
    try:
        analysis_bundle = load_recipe_analysis_bundle(scope=scope, session_id=session_id)
        calculation = calculate_recipe_cost(
            normalized_recipe,
            analysis_bundle["frame"],
            pricing_lookup=analysis_bundle.get("pricing_lookup")
        )
        ingredient_breakdown = calculation.get("ingredient_breakdown", [])
    except ValueError:
        ingredient_breakdown = []

    created_date = recipe.get("created_at") or recipe.get("updated_at") or ""
    summary = {
        "Recipe Name": normalized_recipe.get("name") or "",
        "Yield / Portions": normalized_recipe.get("yield_portions") or "",
        "Total Recipe Cost": recipe.get("total_recipe_cost") if recipe.get("total_recipe_cost") is not None else "",
        "Cost Per Portion": recipe.get("cost_per_portion") if recipe.get("cost_per_portion") is not None else "",
        "Created Date": created_date
    }

    ingredient_rows: list[dict[str, Any]] = []
    for index, ingredient in enumerate(normalized_recipe.get("ingredients", [])):
        breakdown_item = ingredient_breakdown[index] if index < len(ingredient_breakdown) else {}
        ingredient_rows.append({
            "Ingredient Name": ingredient.get("product_name") or "",
            "Quantity": ingredient.get("quantity") if ingredient.get("quantity") is not None else "",
            "Unit": ingredient.get("unit") or "",
            "Purchase Unit": ingredient.get("purchase_unit") or "",
            "Purchase Size": ingredient.get("purchase_size") if ingredient.get("purchase_size") is not None else "",
            "Unit Cost": breakdown_item.get("price_used") if breakdown_item.get("price_used") is not None else "",
            "Ingredient Cost": breakdown_item.get("ingredient_cost") if breakdown_item.get("ingredient_cost") is not None else "",
            "Supplier / Source": str(breakdown_item.get("supplier") or "").strip() or "N/A"
        })

    return {
        "summary": summary,
        "ingredient_rows": ingredient_rows
    }


def build_saved_recipe_export_rows(
    recipe: dict[str, Any],
    *,
    scope: str = "current_upload",
    session_id: str | None = None
) -> list[dict[str, Any]]:
    export_context = build_saved_recipe_export_context(recipe, scope=scope, session_id=session_id)
    summary_columns = export_context["summary"]
    ingredient_rows = export_context["ingredient_rows"]

    if not ingredient_rows:
        return [{
            **summary_columns,
            "Ingredient Name": "",
            "Quantity": "",
            "Unit": "",
            "Purchase Unit": "",
            "Purchase Size": "",
            "Unit Cost": "",
            "Ingredient Cost": "",
            "Supplier / Source": ""
        }]

    rows: list[dict[str, Any]] = []
    for ingredient_row in ingredient_rows:
        rows.append({
            **summary_columns,
            **ingredient_row
        })
    return rows


def build_saved_recipe_export_dataframe(
    recipe: dict[str, Any],
    *,
    scope: str = "current_upload",
    session_id: str | None = None
) -> pd.DataFrame:
    return pd.DataFrame(build_saved_recipe_export_rows(recipe, scope=scope, session_id=session_id))


def build_saved_recipe_export_summary_rows(
    recipe: dict[str, Any],
    *,
    scope: str = "current_upload",
    session_id: str | None = None
) -> list[tuple[str, Any]]:
    summary = build_saved_recipe_export_context(recipe, scope=scope, session_id=session_id)["summary"]
    return [
        ("Recipe Name", summary.get("Recipe Name")),
        ("Yield / Portions", summary.get("Yield / Portions")),
        ("Total Recipe Cost", summary.get("Total Recipe Cost")),
        ("Cost Per Portion", summary.get("Cost Per Portion")),
        ("Created Date", summary.get("Created Date"))
    ]


def build_saved_recipe_excel_stream(
    recipe: dict[str, Any],
    *,
    scope: str = "current_upload",
    session_id: str | None = None
) -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Recipe Export"

    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    section_fill = PatternFill(fill_type="solid", fgColor="EAF2F8")
    bold_font = Font(bold=True)

    sheet["A1"] = "Recipe Summary"
    sheet["A1"].font = Font(bold=True, size=14)

    summary_rows = build_saved_recipe_export_summary_rows(recipe, scope=scope, session_id=session_id)
    summary_start_row = 3
    for offset, (label, value) in enumerate(summary_rows):
        row_number = summary_start_row + offset
        sheet.cell(row=row_number, column=1, value=label).font = bold_font
        sheet.cell(row=row_number, column=2, value=value if value not in (None, "") else "--")

    ingredient_header_row = summary_start_row + len(summary_rows) + 2
    sheet.cell(row=ingredient_header_row, column=1, value="Ingredients").font = Font(bold=True, size=13)
    ingredient_table_row = ingredient_header_row + 1

    ingredient_headers = [
        "Ingredient",
        "Quantity Used",
        "Usage Unit",
        "Purchase Unit",
        "Purchase Size",
        "Unit Cost",
        "Ingredient Cost",
        "Supplier / Source"
    ]
    for column_index, header in enumerate(ingredient_headers, start=1):
        cell = sheet.cell(row=ingredient_table_row, column=column_index, value=header)
        cell.font = bold_font
        cell.fill = header_fill

    ingredient_rows = build_saved_recipe_export_context(recipe, scope=scope, session_id=session_id)["ingredient_rows"]
    for row_offset, ingredient_row in enumerate(ingredient_rows, start=1):
        row_number = ingredient_table_row + row_offset
        sheet.cell(row=row_number, column=1, value=ingredient_row.get("Ingredient Name"))
        sheet.cell(row=row_number, column=2, value=ingredient_row.get("Quantity"))
        sheet.cell(row=row_number, column=3, value=ingredient_row.get("Unit"))
        sheet.cell(row=row_number, column=4, value=ingredient_row.get("Purchase Unit"))
        sheet.cell(row=row_number, column=5, value=ingredient_row.get("Purchase Size"))
        sheet.cell(row=row_number, column=6, value=ingredient_row.get("Unit Cost"))
        sheet.cell(row=row_number, column=7, value=ingredient_row.get("Ingredient Cost"))
        sheet.cell(row=row_number, column=8, value=ingredient_row.get("Supplier / Source"))

    for cell_ref in ("A1", f"A{ingredient_header_row}"):
        sheet[cell_ref].fill = section_fill

    for column_letter, width in {
        "A": 28,
        "B": 16,
        "C": 14,
        "D": 16,
        "E": 14,
        "F": 14,
        "G": 16,
        "H": 20
    }.items():
        sheet.column_dimensions[column_letter].width = width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_saved_recipe_export_filename(recipe: dict[str, Any], extension: str) -> str:
    recipe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", str(recipe.get("name") or "recipe").strip()).strip("._") or "recipe"
    recipe_id = re.sub(r"[^A-Za-z0-9._-]+", "_", str(recipe.get("recipe_id") or "snapshot").strip()).strip("._") or "snapshot"
    return f"{recipe_name}_{recipe_id}.{extension}"


def normalize_ai_rows(rows: list[dict[str, Any]] | None = None) -> pd.DataFrame:
    if rows is None:
        latest_results_path = get_current_latest_results_path()
        if not latest_results_path.exists():
            raise ValueError("No analyzed dataset is available yet.")
        source_rows = pd.read_csv(latest_results_path).to_dict(orient="records")
    else:
        source_rows = rows

    normalized_rows = []
    for row in source_rows or []:
        product_name = (row.get("Product Name") or row.get("productName") or "").strip()
        supplier = (row.get("Supplier") or row.get("supplier") or "").strip()
        if not product_name or not supplier:
            continue

        normalized_rows.append({
            "Product Name": product_name,
            "Supplier": supplier,
            "Unit": (row.get("Unit") or row.get("purchaseUnit") or "unit" or "").strip() or "unit",
            "Quantity": row.get("Quantity", row.get("quantity", 1)),
            "Unit Price": row.get("Unit Price", row.get("unitPrice", row.get("price", 0))),
            "Total Amount": row.get("Total Amount", row.get("totalAmount", None)),
            "Average Price": row.get("Average Price", row.get("averagePrice", None)),
            "Overpay": row.get("Overpay", row.get("overpay", None)),
            "Savings Opportunity": row.get("Savings Opportunity", row.get("savingsOpportunity", None)),
            "Date": row.get("Date", row.get("date", None)),
            "Status": row.get("Status", row.get("status", "Normal")),
            "Overpay Pct": row.get("Overpay Pct", row.get("overpayPct", None))
        })

    frame = pd.DataFrame(normalized_rows)
    if frame.empty:
        raise ValueError("No analyzed rows are available for AI insights yet.")

    numeric_columns = [
        "Quantity",
        "Unit Price",
        "Total Amount",
        "Average Price",
        "Overpay",
        "Savings Opportunity",
        "Overpay Pct"
    ]
    for column in numeric_columns:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")

    frame["Quantity"] = frame["Quantity"].fillna(1).clip(lower=0)
    frame["Unit Price"] = frame["Unit Price"].fillna(0)
    frame["Date"] = pd.to_datetime(frame["Date"], errors="coerce")
    frame["Product Display"] = frame["Product Name"] + " (" + frame["Unit"].fillna("unit").astype(str) + ")"
    frame["Average Price"] = frame["Average Price"].fillna(
        frame.groupby("Product Display")["Unit Price"].transform("mean")
    )
    frame["Total Amount"] = frame["Total Amount"].fillna(frame["Unit Price"] * frame["Quantity"])
    computed_overpay = (frame["Unit Price"] - frame["Average Price"]).clip(lower=0)
    frame["Overpay"] = frame["Overpay"].fillna(computed_overpay)
    frame["Savings Opportunity"] = frame["Savings Opportunity"].fillna(frame["Overpay"] * frame["Quantity"])
    computed_overpay_pct = (
        (frame["Unit Price"] - frame["Average Price"])
        / frame["Average Price"].replace(0, pd.NA)
    ) * 100
    frame["Overpay Pct"] = frame["Overpay Pct"].fillna(computed_overpay_pct).fillna(0)
    frame["Period"] = frame["Date"].dt.to_period("M")
    return frame


def build_ai_snapshot(frame: pd.DataFrame) -> dict[str, Any]:
    dated_periods = frame["Period"].dropna()
    current_period = dated_periods.max() if not dated_periods.empty else None
    current_frame = frame[frame["Period"] == current_period].copy() if current_period is not None else frame.copy()
    baseline_frame = frame[frame["Period"] != current_period].copy() if current_period is not None else frame.iloc[0:0].copy()
    basis_label = "yearly average" if frame["Period"].nunique(dropna=True) >= 10 else "historical average"

    product_totals = frame.groupby("Product Display", as_index=False).agg(
        total_savings=("Savings Opportunity", "sum"),
        total_overpay=("Overpay", "sum"),
        total_spend=("Total Amount", "sum"),
        avg_overpay_pct=("Overpay Pct", "mean"),
        rows=("Product Display", "size"),
        supplier_count=("Supplier", "nunique")
    )
    current_product = current_frame.groupby("Product Display", as_index=False).agg(
        current_avg=("Unit Price", "mean"),
        current_savings=("Savings Opportunity", "sum"),
        current_overpay=("Overpay", "sum"),
        current_rows=("Product Display", "size")
    )
    baseline_product = baseline_frame.groupby("Product Display", as_index=False).agg(
        baseline_avg=("Unit Price", "mean"),
        baseline_rows=("Product Display", "size")
    )
    product_compare = product_totals.merge(current_product, on="Product Display", how="left").merge(
        baseline_product,
        on="Product Display",
        how="left"
    )
    product_compare["current_avg"] = product_compare["current_avg"].fillna(product_compare["total_spend"] / product_compare["rows"].replace(0, pd.NA))
    product_compare["price_delta"] = product_compare["current_avg"] - product_compare["baseline_avg"]
    product_compare["delta_pct"] = (
        product_compare["price_delta"]
        / product_compare["baseline_avg"].replace(0, pd.NA)
    ) * 100
    product_compare["delta_pct"] = product_compare["delta_pct"].fillna(0)

    supplier_totals = frame.groupby("Supplier", as_index=False).agg(
        total_savings=("Savings Opportunity", "sum"),
        total_spend=("Total Amount", "sum"),
        avg_overpay_pct=("Overpay Pct", "mean"),
        overpay_rows=("Status", lambda values: int((values == "Overpay").sum())),
        row_count=("Supplier", "size")
    )
    current_supplier = current_frame.groupby("Supplier", as_index=False).agg(
        current_avg=("Unit Price", "mean"),
        current_savings=("Savings Opportunity", "sum")
    )
    baseline_supplier = baseline_frame.groupby("Supplier", as_index=False).agg(
        baseline_avg=("Unit Price", "mean")
    )
    supplier_compare = supplier_totals.merge(current_supplier, on="Supplier", how="left").merge(
        baseline_supplier,
        on="Supplier",
        how="left"
    )
    supplier_compare["price_delta"] = supplier_compare["current_avg"] - supplier_compare["baseline_avg"]
    supplier_compare["delta_pct"] = (
        supplier_compare["price_delta"]
        / supplier_compare["baseline_avg"].replace(0, pd.NA)
    ) * 100
    supplier_compare["delta_pct"] = supplier_compare["delta_pct"].fillna(0)
    supplier_compare["risk_score"] = (
        supplier_compare["total_savings"] * 1.35
        + supplier_compare["overpay_rows"] * 8
        + supplier_compare["avg_overpay_pct"].clip(lower=0) * 4
        + supplier_compare["delta_pct"].clip(lower=0) * 5
    )

    supplier_product_risk = current_frame.groupby(["Product Display", "Supplier"], as_index=False).agg(
        supplier_savings=("Savings Opportunity", "sum"),
        supplier_avg_price=("Unit Price", "mean")
    )

    visible_total_spend = float(frame["Total Amount"].sum())
    visible_extra_spend = float(frame["Savings Opportunity"].sum())
    overpay_rows = int((frame["Status"] == "Overpay").sum())
    latest_rows = len(current_frame)
    latest_label = format_period_label(current_period)

    return {
        "frame": frame,
        "current_period": current_period,
        "current_period_label": latest_label,
        "basis_label": basis_label,
        "visible_total_spend": visible_total_spend,
        "visible_extra_spend": visible_extra_spend,
        "overpay_rows": overpay_rows,
        "latest_rows": latest_rows,
        "product_compare": product_compare.sort_values(["current_savings", "total_savings"], ascending=False),
        "supplier_compare": supplier_compare.sort_values(["risk_score", "total_savings"], ascending=False),
        "supplier_product_risk": supplier_product_risk.sort_values("supplier_savings", ascending=False)
    }


def build_insight_response(
    headline: str,
    insights: list[str],
    action: str,
    *,
    period_label: str,
    question_type: str,
    suggestions: list[str]
) -> dict[str, Any]:
    return {
        "headline": headline,
        "insights": insights[:4],
        "recommended_action": action,
        "period_label": period_label,
        "question_type": question_type,
        "suggestions": suggestions[:3]
    }


def resolve_question_type(question: str) -> str:
    prompt = (question or "").strip().lower()
    if any(keyword in prompt for keyword in ["renegotiate", "renegotiation", "what should i renegotiate"]):
        return "renegotiate_first"
    if any(keyword in prompt for keyword in ["supplier increased", "increased prices", "fastest", "price increase"]):
        return "supplier_price_increase"
    if any(keyword in prompt for keyword in ["yearly average", "historical average", "above average"]):
        return "above_baseline"
    if any(keyword in prompt for keyword in ["what changed", "changed this month", "versus"]):
        return "period_change"
    if any(keyword in prompt for keyword in ["spike", "unusual", "price spike"]):
        return "price_spike"
    if any(keyword in prompt for keyword in ["supplier", "pricing risk", "risk"]):
        return "supplier_risk"
    if any(keyword in prompt for keyword in ["margin", "losing money", "hurt", "overpay", "money right now"]):
        return "margin_pressure"
    return "overview"


def answer_margin_pressure(snapshot: dict[str, Any]) -> dict[str, Any]:
    product = snapshot["product_compare"].sort_values(
        ["current_savings", "delta_pct", "total_savings"],
        ascending=False
    ).head(1)
    if product.empty:
        return build_insight_response(
            "No margin pressure is visible in the current slice.",
            ["The current visible rows do not contain recoverable overpay right now."],
            "Broaden the visible selection or upload a dataset with recent purchasing activity.",
            period_label=snapshot["current_period_label"],
            question_type="margin_pressure",
            suggestions=[
                "Which supplier has the highest pricing risk?",
                "What should I renegotiate first?"
            ]
        )

    item = product.iloc[0]
    supplier_risk = snapshot["supplier_product_risk"]
    top_supplier = supplier_risk[supplier_risk["Product Display"] == item["Product Display"]].head(1)
    supplier_line = (
        f"Highest supplier pressure appears under {top_supplier.iloc[0]['Supplier']} at {format_currency(top_supplier.iloc[0]['supplier_savings'])} of excess spend."
        if not top_supplier.empty else
        "Supplier concentration is limited in the current visible slice."
    )
    return build_insight_response(
        f"{item['Product Display']} shows the strongest margin pressure in {snapshot['current_period_label']}.",
        [
            f"Visible recoverable spend is {format_currency(coalesce_number(item['current_savings'], item['total_savings']))}.",
            f"Average price is {format_percent(item['delta_pct'])} above the {snapshot['basis_label']}." if pd.notna(item["baseline_avg"]) else "There is not enough historical baseline to compare against prior periods.",
            supplier_line,
            f"{int(item['rows'])} visible purchase rows contribute to this signal."
        ],
        f"Review {item['Product Display']} first and compare current supplier pricing against the baseline before the next purchasing cycle.",
        period_label=snapshot["current_period_label"],
        question_type="margin_pressure",
        suggestions=[
            "What should I renegotiate first?",
            "Which supplier has the highest pricing risk?",
            "Which items show unusual price spikes?"
        ]
    )


def answer_supplier_price_increase(snapshot: dict[str, Any]) -> dict[str, Any]:
    supplier = snapshot["supplier_compare"].sort_values(
        ["delta_pct", "total_savings"],
        ascending=False
    ).head(1)
    if supplier.empty or supplier.iloc[0]["delta_pct"] <= 0:
        return build_insight_response(
            "No supplier is showing a clear price acceleration in the current slice.",
            [
                "Either the visible period is stable or there is not enough dated history for a month-over-month comparison."
            ],
            "Broaden the date range to compare current prices against a stronger historical baseline.",
            period_label=snapshot["current_period_label"],
            question_type="supplier_price_increase",
            suggestions=[
                "What changed this month versus my yearly average?",
                "Which supplier has the highest pricing risk?"
            ]
        )

    item = supplier.iloc[0]
    return build_insight_response(
        f"{item['Supplier']} shows the fastest visible price increase in {snapshot['current_period_label']}.",
        [
            f"Average unit price is {format_percent(item['delta_pct'])} above the {snapshot['basis_label']}.",
            f"Visible excess spend tied to this supplier is {format_currency(item['total_savings'])}.",
            f"{int(item['overpay_rows'])} visible rows are currently flagged as overpay.",
            f"Visible spend under this supplier is {format_currency(item['total_spend'])}."
        ],
        f"Review {item['Supplier']} first, then compare its current pricing against the strongest alternative suppliers in the same product set.",
        period_label=snapshot["current_period_label"],
        question_type="supplier_price_increase",
        suggestions=[
            "Which supplier has the highest pricing risk?",
            "What should I renegotiate first?"
        ]
    )


def answer_renegotiate_first(snapshot: dict[str, Any]) -> dict[str, Any]:
    ranked = snapshot["product_compare"].copy()
    ranked["priority_score"] = (
        ranked["current_savings"].fillna(ranked["total_savings"]) * 1.4
        + ranked["delta_pct"].clip(lower=0) * 6
        + ranked["supplier_count"] * 3
    )
    item = ranked.sort_values("priority_score", ascending=False).head(1)
    if item.empty:
        return answer_margin_pressure(snapshot)

    winner = item.iloc[0]
    return build_insight_response(
        f"{winner['Product Display']} is the first renegotiation target right now.",
        [
            f"Priority score is {winner['priority_score']:.1f} based on excess spend, price drift, and supplier spread.",
            f"Recoverable spend is {format_currency(coalesce_number(winner['current_savings'], winner['total_savings']))}.",
            f"Current pricing is {format_percent(winner['delta_pct'])} above the {snapshot['basis_label']}." if pd.notna(winner["baseline_avg"]) else "Historical comparison is limited, so the score is being driven mostly by current excess spend.",
            f"{int(winner['supplier_count'])} supplier comparison points are available."
        ],
        f"Use {winner['Product Display']} as the first negotiation brief, then work down the next-highest savings opportunities.",
        period_label=snapshot["current_period_label"],
        question_type="renegotiate_first",
        suggestions=[
            "Which product hurt my margin the most this month?",
            "Which supplier increased prices the fastest?"
        ]
    )


def answer_above_baseline(snapshot: dict[str, Any]) -> dict[str, Any]:
    item = snapshot["product_compare"].sort_values("delta_pct", ascending=False).head(1)
    if item.empty or item.iloc[0]["delta_pct"] <= 0:
        return build_insight_response(
            "No product is currently sitting above its visible baseline.",
            ["The current slice does not show a positive variance against the available historical average."],
            "Review a broader date range if you want a deeper baseline comparison.",
            period_label=snapshot["current_period_label"],
            question_type="above_baseline",
            suggestions=[
                "What changed this month versus my yearly average?",
                "Which items show unusual price spikes?"
            ]
        )

    winner = item.iloc[0]
    return build_insight_response(
        f"{winner['Product Display']} is running furthest above its {snapshot['basis_label']}.",
        [
            f"Current average unit price is {format_currency(winner['current_avg'])}.",
            f"That is {format_percent(winner['delta_pct'])} above the baseline average of {format_currency(winner['baseline_avg'])}.",
            f"Visible savings opportunity already totals {format_currency(coalesce_number(winner['current_savings'], winner['total_savings']))}.",
            f"{int(winner['rows'])} visible rows support the signal."
        ],
        f"Treat {winner['Product Display']} as a watchlist product and compare the latest purchases against your baseline target price before reordering.",
        period_label=snapshot["current_period_label"],
        question_type="above_baseline",
        suggestions=[
            "Which items show unusual price spikes?",
            "What should I renegotiate first?"
        ]
    )


def answer_period_change(snapshot: dict[str, Any]) -> dict[str, Any]:
    changed = snapshot["product_compare"].sort_values(["delta_pct", "current_savings"], ascending=False).head(1)
    if changed.empty:
        return answer_overview(snapshot)

    item = changed.iloc[0]
    return build_insight_response(
        f"{item['Product Display']} shows the sharpest shift in {snapshot['current_period_label']}.",
        [
            f"Current average price is {format_currency(item['current_avg'])}.",
            f"Variance versus the {snapshot['basis_label']} is {format_percent(item['delta_pct'])}.",
            f"Visible excess spend in the current slice is {format_currency(coalesce_number(item['current_savings'], item['total_savings']))}.",
            f"Current period coverage includes {int(coalesce_number(item['current_rows'], 0))} visible rows."
        ],
        f"Use {item['Product Display']} as the lead story for this period review, then validate whether the shift is supplier-specific or market-wide.",
        period_label=snapshot["current_period_label"],
        question_type="period_change",
        suggestions=[
            "Which supplier increased prices the fastest?",
            "Which products are above yearly average price?"
        ]
    )


def answer_price_spike(snapshot: dict[str, Any]) -> dict[str, Any]:
    spiking = snapshot["product_compare"].sort_values(["delta_pct", "current_avg"], ascending=False).head(1)
    if spiking.empty or spiking.iloc[0]["delta_pct"] <= 0:
        return build_insight_response(
            "No unusual price spike is standing out in the current slice.",
            ["The visible price pattern looks relatively stable against the available baseline."],
            "If you expect volatility, widen the visible date range and rerun the comparison.",
            period_label=snapshot["current_period_label"],
            question_type="price_spike",
            suggestions=[
                "What changed this month versus my yearly average?",
                "Which supplier has the highest pricing risk?"
            ]
        )

    winner = spiking.iloc[0]
    return build_insight_response(
        f"{winner['Product Display']} shows the clearest unusual price spike right now.",
        [
            f"Current average price is {format_percent(winner['delta_pct'])} above the {snapshot['basis_label']}.",
            f"The latest average is {format_currency(winner['current_avg'])} versus a baseline of {format_currency(winner['baseline_avg'])}.",
            f"Visible recoverable spend tied to the spike is {format_currency(coalesce_number(winner['current_savings'], winner['total_savings']))}.",
            f"The product appears across {int(winner['supplier_count'])} supplier comparison points."
        ],
        f"Flag {winner['Product Display']} for immediate price verification and validate whether the spike is coming from one supplier or across the market.",
        period_label=snapshot["current_period_label"],
        question_type="price_spike",
        suggestions=[
            "Which supplier increased prices the fastest?",
            "What should I renegotiate first?"
        ]
    )


def answer_supplier_risk(snapshot: dict[str, Any]) -> dict[str, Any]:
    supplier = snapshot["supplier_compare"].sort_values(["risk_score", "total_savings"], ascending=False).head(1)
    if supplier.empty:
        return answer_overview(snapshot)

    item = supplier.iloc[0]
    return build_insight_response(
        f"{item['Supplier']} carries the highest pricing risk in the current view.",
        [
            f"Risk score is {item['risk_score']:.1f}, driven by excess spend, overpay frequency, and price drift.",
            f"Visible recoverable spend tied to this supplier is {format_currency(item['total_savings'])}.",
            f"Average pricing sits {format_percent(item['delta_pct'])} above the {snapshot['basis_label']}." if pd.notna(item["baseline_avg"]) else "Historical supplier comparison is limited, so the risk score is being driven mostly by current overpay concentration.",
            f"{int(item['overpay_rows'])} visible rows are currently flagged as overpay."
        ],
        f"Prioritize a supplier review with {item['Supplier']} and compare its latest pricing against both baseline and alternate supplier quotes.",
        period_label=snapshot["current_period_label"],
        question_type="supplier_risk",
        suggestions=[
            "Which supplier increased prices the fastest?",
            "What should I renegotiate first?"
        ]
    )


def answer_overview(snapshot: dict[str, Any]) -> dict[str, Any]:
    product = snapshot["product_compare"].sort_values(["current_savings", "total_savings"], ascending=False).head(1)
    supplier = snapshot["supplier_compare"].sort_values(["risk_score", "total_savings"], ascending=False).head(1)
    product_name = product.iloc[0]["Product Display"] if not product.empty else "the current product mix"
    supplier_name = supplier.iloc[0]["Supplier"] if not supplier.empty else "the visible supplier set"
    return build_insight_response(
        f"{product_name} and {supplier_name} are driving the clearest purchasing pressure right now.",
        [
            f"Visible recoverable spend totals {format_currency(snapshot['visible_extra_spend'])}.",
            f"{snapshot['overpay_rows']} visible row{' is' if snapshot['overpay_rows'] == 1 else 's are'} currently flagged as overpay.",
            f"The current view covers {snapshot['latest_rows']} row{'s' if snapshot['latest_rows'] != 1 else ''} in {snapshot['current_period_label']}.",
            f"Visible spend in scope is {format_currency(snapshot['visible_total_spend'])}."
        ],
        "Start with the top margin-pressure product and the highest-risk supplier, then move into renegotiation targets from there.",
        period_label=snapshot["current_period_label"],
        question_type="overview",
        suggestions=[
            "Which product hurt my margin the most this month?",
            "Which supplier has the highest pricing risk?",
            "What should I renegotiate first?"
        ]
    )


def build_ai_answer(question: str, rows: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    snapshot = build_ai_snapshot(normalize_ai_rows(rows))
    question_type = resolve_question_type(question)
    handlers = {
        "margin_pressure": answer_margin_pressure,
        "supplier_price_increase": answer_supplier_price_increase,
        "renegotiate_first": answer_renegotiate_first,
        "above_baseline": answer_above_baseline,
        "period_change": answer_period_change,
        "price_spike": answer_price_spike,
        "supplier_risk": answer_supplier_risk,
        "overview": answer_overview
    }
    answer = handlers[question_type](snapshot)
    answer["question"] = question.strip()
    answer["source_row_count"] = len(snapshot["frame"])
    return answer


def analyze_dataframe(
    df: pd.DataFrame,
    *,
    source_name: str,
    persist_latest_results: bool = True,
    return_result_df: bool = False
) -> dict | tuple[dict, pd.DataFrame]:
    analysis_started_at = perf_counter()
    required_columns = REQUIRED_ANALYSIS_FIELDS
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    normalize_started_at = perf_counter()
    working_df = df.copy()
    working_df["Product Name"] = working_df["Product Name"].astype(str).str.strip()
    working_df["Supplier"] = working_df["Supplier"].astype(str).str.strip()
    working_df["Unit"] = working_df["Unit"].fillna("unit").astype(str).str.strip().replace("", "unit")
    log_perf("analyze.normalize", normalize_started_at)

    cleaning_started_at = perf_counter()
    working_df["Quantity"] = pd.to_numeric(working_df["Quantity"], errors="coerce")
    working_df["Unit Price"] = pd.to_numeric(working_df["Unit Price"], errors="coerce")
    working_df["Date"] = pd.to_datetime(working_df["Date"], errors="coerce")
    working_df = working_df.dropna(subset=["Product Name", "Supplier", "Unit Price", "Quantity", "Date"])

    if working_df.empty:
        raise ValueError("No valid price data found in the uploaded file.")

    working_df["Quantity"] = working_df["Quantity"].clip(lower=0)
    working_df["Date"] = working_df["Date"].dt.strftime("%Y-%m-%d")
    if "Currency" in working_df.columns:
        working_df["Currency"] = working_df["Currency"].fillna("").astype(str).str.strip().str.upper()
    if "Delivery Time" in working_df.columns:
        working_df["Delivery Time"] = working_df["Delivery Time"].fillna("").astype(str).str.strip()
    if "Payment Terms" in working_df.columns:
        working_df["Payment Terms"] = working_df["Payment Terms"].fillna("").astype(str).str.strip()
    if "Valid Until" in working_df.columns:
        working_df["Valid Until"] = pd.to_datetime(working_df["Valid Until"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
    if "Notes" in working_df.columns:
        working_df["Notes"] = working_df["Notes"].fillna("").astype(str).str.strip()
    if "upload_id" in working_df.columns:
        working_df["upload_id"] = working_df["upload_id"].fillna("").astype(str).str.strip()
    if "row_id" in working_df.columns:
        working_df["row_id"] = working_df["row_id"].fillna("").astype(str).str.strip()
    working_df["__comparison_product_name"] = working_df["Product Name"].map(normalize_comparison_product_name)
    working_df["__comparison_unit"] = working_df["Unit"].map(normalize_comparison_unit)
    log_perf("analyze.clean_types", cleaning_started_at)

    grouping_started_at = perf_counter()
    grouping_columns = ["__comparison_product_name", "__comparison_unit"]
    working_df["Average Price"] = working_df.groupby(grouping_columns)["Unit Price"].transform("mean")
    log_perf("analyze.groupby_average", grouping_started_at)

    calculations_started_at = perf_counter()
    working_df["Overpay Pct"] = ((working_df["Unit Price"] - working_df["Average Price"]) / working_df["Average Price"].replace(0, pd.NA)) * 100
    working_df["Overpay Pct"] = working_df["Overpay Pct"].fillna(0)
    working_df["Overpay"] = (working_df["Unit Price"] - working_df["Average Price"]).clip(lower=0)
    working_df["Savings Opportunity"] = working_df["Overpay"] * working_df["Quantity"]
    working_df["Total Amount"] = working_df["Unit Price"] * working_df["Quantity"]
    working_df["Product Key"] = working_df["__comparison_product_name"] + " | " + working_df["__comparison_unit"]
    working_df["Product Display"] = working_df["Product Name"] + " (" + working_df["Unit"] + ")"
    working_df["Status"] = "Normal"
    overpay_pct = working_df["Overpay Pct"]
    working_df.loc[overpay_pct >= 5, "Status"] = "Overpay"
    working_df.loc[overpay_pct <= -5, "Status"] = "Good Deal"
    log_perf("analyze.price_calculations", calculations_started_at)

    result_started_at = perf_counter()
    total_rows = len(working_df)
    overpay_items = int((working_df["Status"] == "Overpay").sum())
    estimated_extra_spend = round(working_df["Savings Opportunity"].sum(), 2)
    total_spend = round(working_df["Total Amount"].sum(), 2)
    overpay_rate = round((overpay_items / total_rows) * 100, 2) if total_rows else 0

    result_columns = [
        "Product Name",
        "Supplier",
        "Unit",
        "Quantity",
        "Unit Price",
        "Total Amount",
        "Average Price",
        "Overpay",
        "Savings Opportunity",
        "Date",
        "Status",
        "Overpay Pct",
        "Product Key",
        "Product Display"
    ]
    leading_columns = [column for column in ["upload_id", "row_id"] if column in working_df.columns]
    trailing_columns = [
        column for column in ["Currency", "Delivery Time", "Payment Terms", "Valid Until", "Notes"]
        if column in working_df.columns
    ]
    result_columns = leading_columns + result_columns + trailing_columns

    result_df = working_df[result_columns].copy()
    rounded_columns = [
        "Quantity",
        "Unit Price",
        "Total Amount",
        "Average Price",
        "Overpay",
        "Savings Opportunity",
        "Overpay Pct"
    ]
    existing_rounded_columns = [column for column in rounded_columns if column in result_df.columns]
    if existing_rounded_columns:
        result_df.loc[:, existing_rounded_columns] = result_df[existing_rounded_columns].round(2)
    log_perf("analyze.result_materialization", result_started_at)

    if persist_latest_results:
        write_started_at = perf_counter()
        latest_results_path = get_current_latest_results_path()
        result_df.to_csv(latest_results_path, index=False)
        log_perf("analyze.write_csv", write_started_at)
        LATEST_ANALYSIS_CACHE["signature"] = None
        LATEST_ANALYSIS_CACHE["context"] = None
    context_started_at = perf_counter()
    context = build_analysis_context_from_results(
        result_df,
        source_name=source_name,
        precomputed_summary={
            "total_rows": total_rows,
            "overpay_items": overpay_items,
            "estimated_extra_spend": estimated_extra_spend,
            "total_spend": total_spend,
            "overpay_rate": overpay_rate
        }
    )
    log_perf("analyze.response_build", context_started_at)
    log_perf("analyze.total", analysis_started_at)
    if return_result_df:
        return context, result_df
    return context


def build_analysis_context_from_results(
    result_df: pd.DataFrame,
    *,
    source_name: str,
    precomputed_summary: dict[str, Any] | None = None
) -> dict[str, Any]:
    context_started_at = perf_counter()
    summary_started_at = perf_counter()
    rows = result_df.to_dict(orient="records")
    if precomputed_summary is None:
        total_rows = len(result_df)
        overpay_items = int((result_df["Status"] == "Overpay").sum())
        estimated_extra_spend = round(result_df["Savings Opportunity"].sum(), 2)
        total_spend = round(result_df["Total Amount"].sum(), 2)
        overpay_rate = round((overpay_items / total_rows) * 100, 2) if total_rows else 0
    else:
        total_rows = int(precomputed_summary.get("total_rows", len(result_df)))
        overpay_items = int(precomputed_summary.get("overpay_items", 0))
        estimated_extra_spend = round(float(precomputed_summary.get("estimated_extra_spend", 0)), 2)
        total_spend = round(float(precomputed_summary.get("total_spend", 0)), 2)
        overpay_rate = round(float(precomputed_summary.get("overpay_rate", 0)), 2)
    log_perf("analysis_context.summary", summary_started_at)

    aggregation_started_at = perf_counter()
    overpay_df = result_df[result_df["Status"] == "Overpay"]
    top_overpay = overpay_df.nlargest(5, "Savings Opportunity")
    top_savings_by_product = (
        result_df.groupby("Product Display", sort=False)["Savings Opportunity"]
        .sum()
        .nlargest(5)
    )
    status_counts = result_df["Status"].value_counts().to_dict()
    savings_by_supplier = result_df.groupby("Supplier", sort=False)["Savings Opportunity"].sum()
    best_saving_supplier = savings_by_supplier.idxmax() if not savings_by_supplier.empty else "N/A"
    highest_risk_product_series = overpay_df.groupby("Product Display", sort=False)["Savings Opportunity"].sum()
    highest_risk_product = highest_risk_product_series.idxmax() if not highest_risk_product_series.empty else "N/A"
    log_perf("analysis_context.aggregate_rank", aggregation_started_at)

    insights = []
    if highest_risk_product != "N/A":
        insights.append(f"Highest savings risk is concentrated in {highest_risk_product}.")
    if best_saving_supplier != "N/A":
        insights.append(f"Best savings opportunity is linked to supplier comparison against {best_saving_supplier}.")
    if overpay_rate > 0:
        insights.append(f"{overpay_rate}% of analyzed rows are currently flagged as overpay.")
    if estimated_extra_spend > 0:
        insights.append(f"Estimated extra spend identified: ${estimated_extra_spend:.2f}.")

    charts = {
        "top_overpay_labels": top_overpay["Product Display"].tolist(),
        "top_overpay_values": top_overpay["Savings Opportunity"].round(2).tolist(),
        "savings_labels": top_savings_by_product.index.tolist(),
        "savings_values": top_savings_by_product.round(2).tolist(),
        "status_labels": list(status_counts.keys()),
        "status_values": list(status_counts.values())
    }
    log_perf("analysis_context.serialize", context_started_at)

    return {
        "filename": source_name,
        "summary": {
            "total_rows": total_rows,
            "overpay_items": overpay_items,
            "estimated_extra_spend": estimated_extra_spend,
            "total_spend": total_spend,
            "overpay_rate": overpay_rate,
            "best_saving_supplier": best_saving_supplier,
            "highest_risk_product": highest_risk_product
        },
        "has_unit": True,
        "rows": rows,
        "insights": insights,
        "charts_json": json.dumps(charts)
    }


def build_home_redirect(**params) -> RedirectResponse:
    filtered_params = {
        key: value
        for key, value in params.items()
        if value not in (None, "", False)
    }
    query_string = urlencode(filtered_params)
    target = "/" if not query_string else f"/?{query_string}"
    return RedirectResponse(url=target, status_code=303)


def load_latest_analysis_context(*, source_name: str | None = None, demo_mode: bool = False) -> dict:
    latest_results_path = get_current_latest_results_path()
    if not latest_results_path.exists():
        raise ValueError("No results available yet. Please upload a file first.")

    cache_signature = (
        str(latest_results_path),
        latest_results_path.stat().st_mtime_ns,
        latest_results_path.stat().st_size,
        source_name or "Previous analysis",
        bool(demo_mode)
    )
    if (
        LATEST_ANALYSIS_CACHE["signature"] == cache_signature
        and isinstance(LATEST_ANALYSIS_CACHE["context"], dict)
    ):
        return dict(LATEST_ANALYSIS_CACHE["context"])

    latest_results_df = pd.read_csv(latest_results_path)
    analysis_context = build_analysis_context_from_results(
        latest_results_df,
        source_name=source_name or "Previous analysis"
    )
    if demo_mode:
        analysis_context["demo_mode"] = True
    analysis_context["persisted_analysis"] = True
    analysis_context["has_analysis"] = True
    LATEST_ANALYSIS_CACHE["signature"] = cache_signature
    LATEST_ANALYSIS_CACHE["context"] = dict(analysis_context)
    return dict(analysis_context)


def build_page_context(
    request: Request,
    *,
    active_view: str = "quote_compare"
) -> dict[str, Any]:
    context_started_at = perf_counter()
    is_authenticated = bool(getattr(request.state, "auth_user_id", None))
    demo_mode = bool(getattr(request.state, "demo_mode", False))
    has_workspace_access = is_authenticated or demo_mode
    current_user_id = get_current_user_id(request)
    latest_results_path = get_current_latest_results_path(current_user_id)
    has_saved_recipes = False
    if active_view == "recipes":
        try:
            if demo_mode:
                has_saved_recipes = True
            else:
                has_saved_recipes = bool(load_recipes_store().get("recipes", []))
        except Exception:
            logger.exception("Failed to inspect saved recipes while building page context")
            has_saved_recipes = False
    has_analysis = (
        True
        if demo_mode
        else (has_recipe_analysis_source() if active_view == "recipes" else latest_results_path.exists())
    )
    recipes_has_visible_content = has_analysis or has_saved_recipes
    context: dict[str, Any] = {
        "request": request,
        "active_view": active_view,
        "is_authenticated": is_authenticated,
        "has_workspace_access": has_workspace_access,
        "demo_mode": demo_mode,
        "current_user_id": current_user_id,
        "has_analysis": has_analysis,
        "has_saved_recipes": has_saved_recipes,
        "recipes_has_visible_content": recipes_has_visible_content,
        "persisted_analysis": has_analysis,
        "rows": [],
        "summary": None,
        "insights": [],
        "charts_json": "{}",
        "has_unit": True
    }
    error = request.query_params.get("error")
    source_name = request.query_params.get("filename")
    context["filename"] = source_name or ("Previous analysis" if has_analysis else "")

    if error:
        context["error"] = error

    logger.info(
        "[page context timing] active_view=%s is_authenticated=%s has_analysis=%s has_saved_recipes=%s total_ms=%.1f",
        active_view,
        is_authenticated,
        has_analysis,
        has_saved_recipes,
        (perf_counter() - context_started_at) * 1000
    )
    return context


def wants_json_response(request: Request) -> bool:
    accept_header = (request.headers.get("accept") or "").lower()
    requested_with = (request.headers.get("x-requested-with") or "").lower()
    return "application/json" in accept_header or requested_with == "xmlhttprequest"


def create_app() -> FastAPI:
    app = FastAPI()
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
    app.state.templates = templates

    @app.on_event("startup")
    async def log_startup_state():
        logger.info("Starting price analyzer app")
        logger.info(
            "Runtime config | app_env=%s debug=%s secure_cookies=%s max_upload_size=%s",
            APP_ENV,
            DEBUG_MODE,
            AUTH_COOKIE_SECURE,
            format_upload_size_limit(MAX_UPLOAD_SIZE_BYTES)
        )
        if IS_PRODUCTION and not RESEND_API_KEY:
            logger.warning("Production mode is enabled but RESEND_API_KEY is not set. Password reset emails will fail to send.")
        logger.info("Base directory: %s", BASE_DIR)
        logger.info("Templates directory verified: %s", TEMPLATES_DIR)
        logger.info("Static directory mounted: %s", STATIC_DIR)
        ensure_auth_db()
        ensure_recipes_file()
        logger.info("Recipes file ready: %s", RECIPES_PATH)
        ensure_quote_comparisons_file()
        logger.info("Quote comparisons file ready: %s", QUOTE_COMPARISONS_PATH)
        ensure_analysis_history_file()
        logger.info("Analysis history file ready: %s", ANALYSIS_HISTORY_PATH)
        try:
            templates.get_template(INDEX_TEMPLATE)
            logger.info("Template preflight passed: %s", INDEX_TEMPLATE)
        except Exception:
            logger.exception("Template preflight failed during startup")
            raise

    @app.get("/")
    def home(request: Request):
        if not getattr(request.state, "auth_user_id", None) and not getattr(request.state, "demo_mode", False):
            return RedirectResponse(url="/login", status_code=303)
        return safe_template_response(
            request,
            INDEX_TEMPLATE,
            build_page_context(request, active_view="quote_compare")
        )

    @app.get("/recipes")
    def recipes_view(request: Request):
        return safe_template_response(
            request,
            INDEX_TEMPLATE,
            build_page_context(request, active_view="recipes")
        )

    @app.get("/guide")
    def guide_view(request: Request):
        return safe_template_response(
            request,
            INDEX_TEMPLATE,
            build_page_context(request, active_view="guide")
        )

    @app.get("/notes")
    def notes_view(request: Request):
        return safe_template_response(
            request,
            INDEX_TEMPLATE,
            build_page_context(request, active_view="notes")
        )

    @app.get("/quote-compare")
    def quote_compare_view(request: Request):
        return safe_template_response(
            request,
            INDEX_TEMPLATE,
            build_page_context(request, active_view="quote_compare")
        )

    @app.post("/validate-code")
    def validate_code(payload: AccessSessionPayload):
        logger.warning(
            "Legacy access-code validation attempt blocked | code=%s | session_id=%s",
            normalize_access_code(payload.code),
            normalize_session_id(payload.session_id)
        )
        return JSONResponse(
            content={"success": False, "detail": "Legacy access system disabled"},
            status_code=410
        )

    @app.post("/logout")
    def logout(payload: AccessSessionPayload):
        logger.warning(
            "Legacy access-code logout attempt blocked | code=%s | session_id=%s",
            normalize_access_code(payload.code),
            normalize_session_id(payload.session_id)
        )
        return JSONResponse(
            content={"success": False, "detail": "Legacy access system disabled"},
            status_code=410
        )

    @app.get("/generate-code")
    def generate_code():
        logger.warning("Legacy access-code generation attempt blocked")
        return JSONResponse(
            content={"success": False, "detail": "Legacy access system disabled"},
            status_code=410
        )

    @app.get("/quote-compare/download-sample-csv")
    def download_quote_compare_sample_csv():
        sample_df = build_quote_compare_sample_dataframe()
        output = io.StringIO()
        sample_df.to_csv(output, index=False)

        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=quote_compare_sample.csv"}
        )

    @app.get("/quote-compare/download-sample-excel")
    def download_quote_compare_sample_excel():
        sample_df = build_quote_compare_sample_dataframe()
        output = dataframe_to_excel_stream(sample_df, "Compare Prices Sample")

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=quote_compare_sample.xlsx"}
        )

    @app.post("/guide/ask")
    def guide_ask(payload: GuideAskPayload):
        question = (payload.question or "").strip()
        if not question:
            return JSONResponse(
                {"success": False, "message": "Enter a product-help question first."},
                status_code=400
            )

        answer = build_guide_response(question)
        return JSONResponse({
            "success": True,
            **answer
        })

    @app.get("/recipes/bootstrap")
    def recipes_bootstrap(scope: str = "current_upload", session_id: str | None = None):
        bootstrap_started_at = perf_counter()
        normalized_scope = normalize_analysis_scope(scope)
        bootstrap_substeps: dict[str, Any] = {
            "scope": normalized_scope
        }
        recipe_bootstrap_metrics = {
            "deduplicate_call_count": 0
        }
        bootstrap_metrics_token = RECIPE_BOOTSTRAP_METRICS_CONTEXT.set(recipe_bootstrap_metrics)
        analysis_signature = ("demo_analysis", len(build_demo_analysis_dataframe().index)) if normalized_scope == "demo" else get_recipe_analysis_cache_signature()
        if normalized_scope == "demo":
            demo_store = load_demo_recipe_store(str(session_id or "").strip())
            demo_session_path = get_demo_recipe_session_path(str(session_id or "").strip())
            recipes_signature = (
                str(demo_session_path),
                demo_session_path.stat().st_mtime_ns,
                demo_session_path.stat().st_size,
                len(demo_store.get("recipes", []))
            )
        else:
            recipes_signature = get_recipes_store_cache_signature()
        response_cache_signature = (
            normalized_scope,
            str(session_id or "").strip() if normalized_scope == "demo" else "",
            analysis_signature,
            recipes_signature
        )
        try:
            cache_hit = (
                RECIPES_BOOTSTRAP_RESPONSE_CACHE["signature"] == response_cache_signature
                and isinstance(RECIPES_BOOTSTRAP_RESPONSE_CACHE["response_json"], str)
            )
            log_perf_details(
                "recipes.bootstrap.cache_status",
                cache_hit=cache_hit,
                cache_reason="signature_match" if cache_hit else "signature_miss",
                has_analysis_signature=analysis_signature is not None,
                deduplicate_call_count=int(recipe_bootstrap_metrics.get("deduplicate_call_count") or 0)
            )
            if cache_hit:
                cached_response_json = RECIPES_BOOTSTRAP_RESPONSE_CACHE["response_json"]
                bootstrap_substeps["deduplicate_call_count"] = int(recipe_bootstrap_metrics.get("deduplicate_call_count") or 0)
                bootstrap_substeps["payload_bytes"] = len(cached_response_json.encode("utf-8"))
                bootstrap_substeps["total_ms"] = round((perf_counter() - bootstrap_started_at) * 1000, 1)
                log_perf_details("recipes.bootstrap.backend", **bootstrap_substeps)
                return Response(content=cached_response_json, media_type="application/json")

            try:
                analysis_load_started_at = perf_counter()
                analysis_bundle = load_recipe_analysis_bundle(scope=normalized_scope, session_id=session_id)
                frame = analysis_bundle["frame"]
                bootstrap_substeps["analysis_load_ms"] = round((perf_counter() - analysis_load_started_at) * 1000, 1)
            except ValueError:
                frame = pd.DataFrame()
                analysis_bundle = {"product_catalog": [], "pricing_lookup": {}}
                bootstrap_substeps["analysis_load_ms"] = round((perf_counter() - bootstrap_started_at) * 1000, 1)

            recipes_load_started_at = perf_counter()
            recipes = load_recipe_store_for_scope(scope=normalized_scope, session_id=session_id).get("recipes", [])
            bootstrap_substeps["load_recipes_store_ms"] = round((perf_counter() - recipes_load_started_at) * 1000, 1)
            if frame.empty:
                products = []
                enriched_recipes = recipes
                bootstrap_substeps["product_catalog_ms"] = 0.0
                bootstrap_substeps["pricing_lookup_ms"] = 0.0
                bootstrap_substeps["recipe_enrichment_ms"] = 0.0
            else:
                product_catalog_started_at = perf_counter()
                products = analysis_bundle["product_catalog"]
                bootstrap_substeps["product_catalog_ms"] = round((perf_counter() - product_catalog_started_at) * 1000, 1)
                pricing_lookup_started_at = perf_counter()
                pricing_lookup = analysis_bundle.get("pricing_lookup") or {}
                bootstrap_substeps["pricing_lookup_ms"] = round((perf_counter() - pricing_lookup_started_at) * 1000, 1)
                enrich_started_at = perf_counter()
                enriched_recipes = enrich_saved_recipes_with_costs(
                    recipes,
                    frame,
                    pricing_lookup=pricing_lookup
                )
                bootstrap_substeps["recipe_enrichment_ms"] = round((perf_counter() - enrich_started_at) * 1000, 1)

            payload_started_at = perf_counter()
            build_payload_substeps: dict[str, Any] = {}
            scope_options_started_at = perf_counter()
            scope_options = build_analysis_scope_options()
            build_payload_substeps["scope_options_ms"] = round((perf_counter() - scope_options_started_at) * 1000, 1)
            latest_upload_started_at = perf_counter()
            latest_upload = (
                {"upload_id": "demo", "source_name": "Demo Data"}
                if normalized_scope == "demo"
                else get_latest_analysis_upload_meta(
                    prepared_frame=frame,
                    perf_label="recipes.bootstrap.latest_upload_meta.substeps"
                )
            )
            build_payload_substeps["latest_upload_meta_ms"] = round((perf_counter() - latest_upload_started_at) * 1000, 1)
            scope_summary_started_at = perf_counter()
            scope_summary = build_analysis_scope_summary_with_upload(frame, latest_upload=latest_upload, scope=normalized_scope)
            build_payload_substeps["scope_summary_ms"] = round((perf_counter() - scope_summary_started_at) * 1000, 1)
            pricing_modes_started_at = perf_counter()
            pricing_modes = [
                {"value": value, "label": label}
                for value, label in RECIPE_PRICING_MODES.items()
            ]
            build_payload_substeps["pricing_modes_ms"] = round((perf_counter() - pricing_modes_started_at) * 1000, 1)
            response_payload = {
                "success": True,
                "scope_options": scope_options,
                "scope_summary": scope_summary,
                "pricing_modes": pricing_modes,
                "products": products,
                "recipes": enriched_recipes
            }
            bootstrap_substeps["build_payload_ms"] = round((perf_counter() - payload_started_at) * 1000, 1)
            build_payload_substeps["products_count"] = len(products)
            build_payload_substeps["recipes_count"] = len(enriched_recipes)
            build_payload_substeps["frame_rows"] = int(len(frame.index))
            log_perf_details("recipes.bootstrap.build_payload.substeps", **build_payload_substeps)

            serialize_started_at = perf_counter()
            response_json = json.dumps(response_payload, ensure_ascii=False, separators=(",", ":"))
            bootstrap_substeps["serialize_ms"] = round((perf_counter() - serialize_started_at) * 1000, 1)
            bootstrap_substeps["payload_bytes"] = len(response_json.encode("utf-8"))
            bootstrap_substeps["products_count"] = len(products)
            bootstrap_substeps["recipes_count"] = len(enriched_recipes)
            bootstrap_substeps["frame_rows"] = int(len(frame.index))
            bootstrap_substeps["deduplicate_call_count"] = int(recipe_bootstrap_metrics.get("deduplicate_call_count") or 0)
            bootstrap_substeps["total_ms"] = round((perf_counter() - bootstrap_started_at) * 1000, 1)
            log_perf_details("recipes.bootstrap.deduplicate_call_count", count=bootstrap_substeps["deduplicate_call_count"])
            log_perf_details("recipes.bootstrap.backend", **bootstrap_substeps)
            RECIPES_BOOTSTRAP_RESPONSE_CACHE["signature"] = response_cache_signature
            RECIPES_BOOTSTRAP_RESPONSE_CACHE["response_json"] = response_json
            return Response(content=response_json, media_type="application/json")
        finally:
            RECIPE_BOOTSTRAP_METRICS_CONTEXT.reset(bootstrap_metrics_token)

    @app.get("/analysis/scope-bootstrap")
    def analysis_scope_bootstrap(scope: str = "current_upload", session_id: str | None = None):
        request_started_at = perf_counter()
        normalized_scope = normalize_analysis_scope(scope)
        fast_path_details: dict[str, Any] = {
            "scope": normalized_scope,
            "has_session_id": bool(str(session_id or "").strip()),
            "fast_path": False,
            "fast_path_reason": "frame_required"
        }
        overlap_reason = "analysis_frame_required"
        endpoint_selection = "analysis.scope_bootstrap.frame"

        if normalized_scope == "demo":
            frame = build_demo_analysis_dataframe()
            summary = build_analysis_scope_summary(frame, scope="demo")
            response_payload = {
                "success": True,
                "has_analysis": bool(summary["row_count"]),
                "scope_options": build_analysis_scope_options(),
                "scope_summary": summary
            }
            response_json = json.dumps(response_payload, ensure_ascii=False, separators=(",", ":"))
            return Response(content=response_json, media_type="application/json")

        if normalized_scope == "current_upload" and str(session_id or "").strip():
            active_session_started_at = perf_counter()
            active_session = validate_quote_compare_active_session(load_quote_compare_active_session(session_id))
            fast_path_details["active_session_lookup_ms"] = round((perf_counter() - active_session_started_at) * 1000, 1)
            if (
                isinstance(active_session, dict)
                and str(active_session.get("step") or "").strip().lower() == "analyze"
                and isinstance(active_session.get("comparison"), dict)
            ):
                comparison = active_session["comparison"]
                latest_upload = {
                    "upload_id": str(comparison.get("upload_id") or "").strip(),
                    "source_name": (
                        str(active_session.get("filename") or "").strip()
                        or str(comparison.get("name") or "").strip()
                    )
                }
                summary = build_analysis_scope_summary_from_comparison(comparison, latest_upload=latest_upload, scope="current_upload")
                response_payload = {
                    "success": True,
                    "has_analysis": bool(summary["row_count"]),
                    "scope_options": build_analysis_scope_options(),
                    "scope_summary": summary
                }
                response_json = json.dumps(response_payload, ensure_ascii=False, separators=(",", ":"))
                fast_path_details["fast_path"] = True
                fast_path_details["fast_path_reason"] = "active_session_comparison"
                fast_path_details["response_bytes"] = len(response_json.encode("utf-8"))
                fast_path_details["total_ms"] = round((perf_counter() - request_started_at) * 1000, 1)
                overlap_reason = "reuse_active_session_comparison"
                endpoint_selection = "analysis.scope_bootstrap.active_session"
                log_perf_details("analysis.scope_bootstrap.fast_path", **fast_path_details)
                log_perf_details("bootstrap_overlap_reason", endpoint="analysis.scope_bootstrap", reason=overlap_reason)
                log_perf_details("step3_restore_endpoint_selection", endpoint="analysis.scope_bootstrap", selection=endpoint_selection)
                return Response(content=response_json, media_type="application/json")

        try:
            frame = load_recipe_analysis_dataframe()
        except ValueError:
            frame = pd.DataFrame()
        summary = build_analysis_scope_summary(frame, scope="current_upload")
        response_payload = {
            "success": True,
            "has_analysis": not frame.empty,
            "scope_options": build_analysis_scope_options(),
            "scope_summary": summary
        }
        response_json = json.dumps(response_payload, ensure_ascii=False, separators=(",", ":"))
        fast_path_details["response_bytes"] = len(response_json.encode("utf-8"))
        fast_path_details["total_ms"] = round((perf_counter() - request_started_at) * 1000, 1)
        if frame.empty:
            log_perf_details("bootstrap.empty_state_detected", endpoint="analysis.scope_bootstrap", reason="no_analysis_frame")
        log_perf_details("analysis.scope_bootstrap.fast_path", **fast_path_details)
        log_perf_details("bootstrap_overlap_reason", endpoint="analysis.scope_bootstrap", reason=overlap_reason)
        log_perf_details("step3_restore_endpoint_selection", endpoint="analysis.scope_bootstrap", selection=endpoint_selection)
        return Response(content=response_json, media_type="application/json")

    @app.post("/workspace/reset")
    def workspace_reset():
        try:
            reset_workspace_data_store()
            return JSONResponse({
                "success": True,
                "has_analysis": False,
                "scope_summary": build_analysis_scope_summary(pd.DataFrame(), scope="current_upload"),
                "message": "Workspace data reset successfully."
            })
        except Exception:
            logger.exception("Workspace reset failed")
            return JSONResponse(
                {"success": False, "message": "Workspace reset failed."},
                status_code=500
            )

    @app.post("/recipes/calculate")
    def calculate_recipe(payload: RecipePayload, scope: str = "current_upload", session_id: str | None = None):
        calculate_started_at = perf_counter()
        try:
            analysis_load_started_at = perf_counter()
            analysis_bundle = load_recipe_analysis_bundle(scope=scope, session_id=session_id)
            frame = analysis_bundle["frame"]
            normalized_started_at = perf_counter()
            normalized_recipe = normalize_recipe_payload(payload.model_dump())
            normalize_ms = round((perf_counter() - normalized_started_at) * 1000, 1)
            calculation_started_at = perf_counter()
            calculation = calculate_recipe_cost(
                normalized_recipe,
                frame,
                pricing_lookup=analysis_bundle.get("pricing_lookup")
            )
            log_perf_details(
                "recipes.calculate.backend",
                analysis_load_ms=round((normalized_started_at - analysis_load_started_at) * 1000, 1),
                normalize_recipe_ms=normalize_ms,
                calculation_ms=round((perf_counter() - calculation_started_at) * 1000, 1),
                ingredient_count=len(normalized_recipe.get("ingredients", [])),
                total_ms=round((perf_counter() - calculate_started_at) * 1000, 1)
            )
        except ValueError as exc:
            status_code = 413 if str(exc) == get_upload_size_limit_message() else 400
            return JSONResponse({"success": False, "message": str(exc)}, status_code=status_code)

        return JSONResponse({
            "success": True,
            "recipe": normalized_recipe,
            "calculation": calculation
        })

    @app.post("/recipes/save")
    def save_recipe(payload: RecipePayload, scope: str = "current_upload", session_id: str | None = None):
        save_started_at = perf_counter()
        normalized_scope = normalize_analysis_scope(scope)
        try:
            analysis_load_started_at = perf_counter()
            analysis_bundle = load_recipe_analysis_bundle(scope=normalized_scope, session_id=session_id)
            frame = analysis_bundle["frame"]
            normalize_started_at = perf_counter()
            normalized_recipe = normalize_recipe_payload(payload.model_dump())
            normalize_ms = round((perf_counter() - normalize_started_at) * 1000, 1)
            calculation_started_at = perf_counter()
            calculation = calculate_recipe_cost(
                normalized_recipe,
                frame,
                pricing_lookup=analysis_bundle.get("pricing_lookup")
            )
            pricing_metrics = calculate_recipe_pricing_metrics(normalized_recipe, calculation)
            calculate_ms = round((perf_counter() - calculation_started_at) * 1000, 1)
        except ValueError as exc:
            status_code = 413 if str(exc) == get_upload_size_limit_message() else 400
            return JSONResponse({"success": False, "message": str(exc)}, status_code=status_code)

        now = datetime.now(timezone.utc).isoformat()
        store = load_recipe_store_for_scope(scope=normalized_scope, session_id=session_id)
        recipes = store.get("recipes", [])
        recipe_id = normalized_recipe["recipe_id"] or str(uuid.uuid4())
        existing_recipe = next((recipe for recipe in recipes if recipe.get("recipe_id") == recipe_id), None)

        saved_recipe = {
            "recipe_id": recipe_id,
            "name": normalized_recipe["name"],
            "yield_portions": normalized_recipe["yield_portions"],
            "pricing_mode": normalized_recipe["pricing_mode"],
            "ingredients": normalized_recipe["ingredients"],
            **pricing_metrics,
            "updated_at": now,
            "created_at": existing_recipe.get("created_at") if existing_recipe else now
        }

        if existing_recipe:
            recipes = [saved_recipe if recipe.get("recipe_id") == recipe_id else recipe for recipe in recipes]
        else:
            recipes.append(saved_recipe)

        store["recipes"] = recipes
        persist_started_at = perf_counter()
        save_recipe_store_for_scope(store, scope=normalized_scope, session_id=session_id)
        persist_ms = round((perf_counter() - persist_started_at) * 1000, 1)

        enrich_started_at = perf_counter()
        response_recipes = enrich_saved_recipes_with_costs(
            recipes,
            frame,
            pricing_lookup=analysis_bundle.get("pricing_lookup")
        )
        enrich_ms = round((perf_counter() - enrich_started_at) * 1000, 1)
        log_perf_details(
            "recipes.save.backend",
            analysis_load_ms=round((normalize_started_at - analysis_load_started_at) * 1000, 1),
            normalize_recipe_ms=normalize_ms,
            calculation_ms=calculate_ms,
            persist_ms=persist_ms,
            enrich_recipes_ms=enrich_ms,
            recipe_count=len(response_recipes),
            ingredient_count=len(normalized_recipe.get("ingredients", [])),
            total_ms=round((perf_counter() - save_started_at) * 1000, 1)
        )

        return JSONResponse({
            "success": True,
            "recipe": saved_recipe,
            "calculation": calculation,
            "recipes": response_recipes,
            "message": f"Recipe saved: {saved_recipe['name']}"
        })

    @app.post("/recipes/delete")
    def delete_recipe(payload: RecipeDeletePayload, scope: str = "current_upload", session_id: str | None = None):
        recipe_id = str(payload.recipe_id or "").strip()
        if not recipe_id:
            return JSONResponse({"success": False, "message": "Choose a saved recipe to delete."}, status_code=400)

        normalized_scope = normalize_analysis_scope(scope)
        store = load_recipe_store_for_scope(scope=normalized_scope, session_id=session_id)
        recipes = store.get("recipes", [])
        existing_recipe = next((recipe for recipe in recipes if recipe.get("recipe_id") == recipe_id), None)
        if not existing_recipe:
            return JSONResponse({"success": False, "message": "The selected recipe could not be found."}, status_code=404)

        updated_recipes = [recipe for recipe in recipes if recipe.get("recipe_id") != recipe_id]
        store["recipes"] = updated_recipes
        save_recipe_store_for_scope(store, scope=normalized_scope, session_id=session_id)

        try:
            frame = load_recipe_analysis_dataframe(scope=normalized_scope, session_id=session_id)
            response_recipes = enrich_saved_recipes_with_costs(updated_recipes, frame)
        except ValueError:
            response_recipes = updated_recipes

        return JSONResponse({
            "success": True,
            "deleted_recipe_id": recipe_id,
            "recipes": response_recipes,
            "message": f"Deleted recipe: {existing_recipe.get('name', 'Recipe')}"
        })

    @app.get("/recipes/{recipe_id}/export.csv")
    def export_recipe_csv(recipe_id: str, scope: str = "current_upload", session_id: str | None = None):
        recipe = get_saved_recipe_by_id(recipe_id, scope=scope, session_id=session_id)
        if not recipe:
            return JSONResponse({"success": False, "message": "The selected recipe could not be found."}, status_code=404)
        export_df = build_saved_recipe_export_dataframe(recipe, scope=scope, session_id=session_id)
        output = io.StringIO()
        export_df.to_csv(output, index=False)
        filename = build_saved_recipe_export_filename(recipe, "csv")
        return StreamingResponse(
            iter([output.getvalue().encode("utf-8-sig")]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    @app.get("/recipes/{recipe_id}/export.xlsx")
    def export_recipe_excel(recipe_id: str, scope: str = "current_upload", session_id: str | None = None):
        recipe = get_saved_recipe_by_id(recipe_id, scope=scope, session_id=session_id)
        if not recipe:
            return JSONResponse({"success": False, "message": "The selected recipe could not be found."}, status_code=404)
        output = build_saved_recipe_excel_stream(recipe, scope=scope, session_id=session_id)
        filename = build_saved_recipe_export_filename(recipe, "xlsx")
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    @app.get("/quote-compare/bootstrap")
    def quote_compare_bootstrap(session_id: str | None = None, include_comparisons: bool = False):
        request_started_at = perf_counter()
        store_load_started_at = request_started_at
        store = None
        comparisons: list[dict[str, Any]] = []
        active_session = None

        if include_comparisons:
            store = load_quote_comparisons_store()
        store_loaded_at = perf_counter()

        if include_comparisons and store is not None:
            comparisons = store.get("comparisons", [])
        comparisons_loaded_at = perf_counter()

        active_session = validate_quote_compare_active_session(
            load_quote_compare_active_session(session_id)
        )
        active_session_loaded_at = perf_counter()
        comparisons_bytes = get_json_payload_size_bytes(comparisons)
        active_session_bytes = get_json_payload_size_bytes(active_session)

        response_payload = {
            "success": True,
            "comparisons": comparisons,
            "active_session": active_session
        }
        response_json = json.dumps(response_payload, ensure_ascii=False, separators=(",", ":"))
        response_serialized_at = perf_counter()
        fast_path_reason = (
            "active_session_only"
            if active_session and not include_comparisons
            else ("comparisons_included" if include_comparisons else "no_active_session")
        )
        endpoint_selection = (
            "quote_compare.bootstrap.active_session"
            if active_session and not include_comparisons
            else ("quote_compare.bootstrap.comparisons" if include_comparisons else "quote_compare.bootstrap.default")
        )
        log_perf_details(
            "quote_compare.bootstrap.fast_path",
            has_session_id=bool(session_id),
            include_comparisons=include_comparisons,
            fast_path=bool(active_session and not include_comparisons),
            fast_path_reason=fast_path_reason,
            total_ms=round((response_serialized_at - request_started_at) * 1000, 1),
            response_bytes=len(response_json.encode("utf-8"))
        )
        log_perf_details(
            "bootstrap_overlap_reason",
            endpoint="quote_compare.bootstrap",
            reason=("active_session_contains_step3_state" if active_session and not include_comparisons else "response_contract_requires_bootstrap")
        )
        log_perf_details(
            "step3_restore_endpoint_selection",
            endpoint="quote_compare.bootstrap",
            selection=endpoint_selection
        )
        if not active_session and not comparisons:
            log_perf_details("bootstrap.empty_state_detected", endpoint="quote_compare.bootstrap", reason="no_active_session_or_comparisons")
        logger.info(
            "[Compare Prices bootstrap timing] session_id=%s include_comparisons=%s total_ms=%.1f store_load_ms=%.1f comparisons_ms=%.1f active_session_ms=%.1f response_serialize_ms=%.1f response_bytes=%s active_session_bytes=%s comparisons_bytes=%s active_session_comparison_bytes=%s active_session_evaluation_bytes=%s active_session_headers_bytes=%s active_session_field_reviews_bytes=%s comparisons=%s",
            bool(session_id),
            include_comparisons,
            (response_serialized_at - request_started_at) * 1000,
            (store_loaded_at - store_load_started_at) * 1000,
            (comparisons_loaded_at - store_loaded_at) * 1000,
            (active_session_loaded_at - comparisons_loaded_at) * 1000,
            (response_serialized_at - active_session_loaded_at) * 1000,
            len(response_json.encode("utf-8")),
            active_session_bytes,
            comparisons_bytes,
            get_json_payload_size_bytes((active_session or {}).get("comparison")),
            get_json_payload_size_bytes((active_session or {}).get("evaluation")),
            get_json_payload_size_bytes((active_session or {}).get("headers")),
            get_json_payload_size_bytes((active_session or {}).get("field_reviews")),
            len(comparisons)
        )
        return Response(content=response_json, media_type="application/json")

    @app.post("/quote-compare/demo-data")
    def quote_compare_demo_data():
        session_id = str(uuid.uuid4())
        demo_frame = build_demo_analysis_dataframe()
        normalized_comparison = normalize_quote_comparison_payload({
            "name": "Demo Data",
            "sourcing_need": "Cafe and restaurant purchasing demo",
            "bids": build_quote_bids_from_dataframe(demo_frame)
        })
        evaluation = calculate_quote_comparison(normalized_comparison)
        return JSONResponse({
            "success": True,
            "session_id": session_id,
            "headers": [str(column) for column in demo_frame.columns],
            "mapping": {
                "Product Name": "Product Name",
                "Supplier": "Supplier",
                "Unit": "Unit",
                "Quantity": "Quantity",
                "Unit Price": "Unit Price",
                "Date": "Date",
                "Currency": "Currency",
                "Delivery Time": "Delivery Time",
                "Payment Terms": "Payment Terms",
                "Valid Until": "Valid Until",
                "Notes": "Notes"
            },
            "comparison": normalized_comparison,
            "evaluation": evaluation,
            "message": "Loaded demo supplier offers."
        })

    @app.post("/quote-compare/upload/inspect")
    async def inspect_quote_compare_upload(file: UploadFile = File(...)):
        filename = file.filename or ""
        request_started_at = perf_counter()
        parse_started_at = request_started_at

        try:
            session_id = str(uuid.uuid4())
            df = read_uploaded_dataframe(file)
            parse_finished_at = perf_counter()
            cached_upload_path = cache_quote_compare_upload(file, session_id)
            cache_finished_at = perf_counter()
            columns = [str(column) for column in df.columns]
            required_detection = detect_column_mappings(
                columns,
                required_fields=QUOTE_COMPARE_REQUIRED_FIELDS,
                field_synonyms=QUOTE_COMPARE_FIELD_SYNONYMS
            )
            required_detection_finished_at = perf_counter()
            optional_detection = detect_column_mappings(
                columns,
                required_fields=QUOTE_COMPARE_OPTIONAL_FIELDS,
                field_synonyms=QUOTE_COMPARE_FIELD_SYNONYMS
            )
            optional_detection_finished_at = perf_counter()
        except ValueError as exc:
            return JSONResponse({"success": False, "message": str(exc)}, status_code=400)
        except Exception:
            logger.exception("Failed to inspect Compare Prices upload: %s", filename)
            return JSONResponse(
                {
                    "success": False,
                    "message": "The file could not be read. Please upload a valid CSV or Excel file (.csv, .xlsx, or .xls)."
                },
                status_code=400
            )

        optional_reviews = []
        used_headers = {value for value in required_detection["mapping"].values() if value}
        for review in optional_detection["field_reviews"]:
            detected_column = review.get("detected_column")
            if detected_column in used_headers:
                review = {
                    **review,
                    "detected_column": None,
                    "score": 0,
                    "match_quality": "missing"
                }
            elif detected_column:
                used_headers.add(detected_column)
            optional_reviews.append(review)
        optional_merge_finished_at = perf_counter()

        payload = {
            "session_id": session_id,
            "filename": filename,
            "required_fields": QUOTE_COMPARE_REQUIRED_FIELDS,
            "optional_fields": QUOTE_COMPARE_OPTIONAL_FIELDS,
            "message": "We detected likely supplier-offer fields from your upload.",
            "review_message": "Review required and optional column matches before moving into quote analysis.",
            "mapping": {
                **required_detection["mapping"],
                **{
                    review["field"]: review.get("detected_column")
                    for review in optional_reviews
                }
            },
            "field_reviews": required_detection["field_reviews"] + optional_reviews,
            "matched_fields": required_detection["matched_fields"],
            "missing_fields": required_detection["missing_fields"],
            "optional_columns": required_detection["optional_columns"],
            "headers": columns
        }

        save_quote_compare_active_session(
            payload["session_id"],
            {
                "session_id": payload["session_id"],
                "file_id": payload["session_id"],
                "step": "review",
                "filename": filename,
                "file_path": cached_upload_path,
                "headers": columns,
                "column_count": len(columns),
                "row_count": len(df.index),
                "mapping": payload["mapping"],
                "field_reviews": payload["field_reviews"],
                "required_fields": QUOTE_COMPARE_REQUIRED_FIELDS,
                "optional_fields": QUOTE_COMPARE_OPTIONAL_FIELDS,
                "matched_fields": payload["matched_fields"],
                "missing_fields": payload["missing_fields"],
                "optional_columns": payload["optional_columns"],
                "review_message": payload["review_message"],
                "message": payload["message"],
                "cached_upload_path": cached_upload_path
            }
        )
        session_saved_at = perf_counter()
        response_payload = {"success": True, **payload}
        response_json = json.dumps(response_payload, ensure_ascii=False, separators=(",", ":"))
        response_serialized_at = perf_counter()
        logger.info(
            "[Compare Prices inspect timing] filename=%s total_ms=%.1f parse_ms=%.1f cache_upload_ms=%.1f required_detect_ms=%.1f optional_detect_ms=%.1f optional_merge_ms=%.1f session_save_ms=%.1f response_serialize_ms=%.1f response_bytes=%s headers=%s",
            filename,
            (response_serialized_at - request_started_at) * 1000,
            (parse_finished_at - parse_started_at) * 1000,
            (cache_finished_at - parse_finished_at) * 1000,
            (required_detection_finished_at - cache_finished_at) * 1000,
            (optional_detection_finished_at - required_detection_finished_at) * 1000,
            (optional_merge_finished_at - optional_detection_finished_at) * 1000,
            (session_saved_at - optional_merge_finished_at) * 1000,
            (response_serialized_at - session_saved_at) * 1000,
            len(response_json.encode("utf-8")),
            len(columns)
        )

        return JSONResponse(response_payload)

    @app.post("/quote-compare/upload/confirm")
    async def confirm_quote_compare_upload(
        file: UploadFile | None = File(None),
        mappings: str = Form(...),
        session_id: str | None = Form(None)
    ):
        filename = file.filename if file is not None else ""
        request_started_at = perf_counter()

        try:
            parsed_mapping = json.loads(mappings)
            if not isinstance(parsed_mapping, dict):
                raise ValueError
        except ValueError:
            return JSONResponse(
                {
                    "success": False,
                    "message": "The selected quote field mappings could not be understood. Please review them and try again."
                },
                status_code=400
            )

        try:
            load_source_started_at = perf_counter()
            session_payload = validate_quote_compare_active_session(load_quote_compare_active_session(session_id))
            selected_source_columns = get_quote_compare_selected_source_columns(parsed_mapping)
            if file is None and not session_payload:
                logger.warning(
                    "[Compare Prices upload] confirm called without a file or restorable session | session_id=%s",
                    session_id
                )
                raise ValueError("The uploaded supplier file is missing. Please upload it again.")
            if file is not None and (file.filename or ""):
                df = read_uploaded_dataframe(
                    file,
                    selected_columns=selected_source_columns,
                    perf_label_prefix="confirm"
                )
            elif session_payload and session_payload.get("dataframe"):
                df = hydrate_dataframe_from_session(
                    session_payload["dataframe"].get("columns", []),
                    session_payload["dataframe"].get("records", [])
                )
                filename = session_payload.get("filename", filename)
            elif session_payload and session_payload.get("cached_upload_path"):
                filename = session_payload.get("filename", filename)
                df = read_cached_quote_compare_upload(
                    session_payload.get("cached_upload_path"),
                    filename=filename,
                    selected_columns=selected_source_columns,
                    perf_label_prefix="confirm"
                )
            else:
                raise ValueError("The uploaded supplier file is no longer available. Please upload it again.")
            log_perf("confirm.load_source_dataframe", load_source_started_at)
            mapping_started_at = perf_counter()
            full_mapping = {
                field_name: parsed_mapping.get(field_name)
                for field_name in [*QUOTE_COMPARE_REQUIRED_FIELDS, *QUOTE_COMPARE_OPTIONAL_FIELDS]
            }
            logger.info(
                "[Compare Prices upload] confirm mapping debug | selected_supplier_mapping=%s | dataframe_columns_before_rename=%s",
                full_mapping.get("Supplier"),
                list(df.columns)
            )
            mapped_df = apply_column_mapping(
                df,
                full_mapping,
                required_fields=QUOTE_COMPARE_REQUIRED_FIELDS
            )
            mapped_df = normalize_quote_compare_mapped_dataframe(
                mapped_df,
                selected_mapping=full_mapping,
                source_columns=list(df.columns)
            )
            logger.info(
                "[Compare Prices upload] confirm mapping debug | dataframe_columns_after_rename=%s",
                list(mapped_df.columns)
            )
            log_perf("confirm.mapping_normalization", mapping_started_at)
            import_started_at = perf_counter()
            import_result = build_quote_bid_import_result(mapped_df, text_already_normalized=True)
            if import_result["valid_row_count"] <= 0:
                raise ValueError("No valid supplier offer rows were found after filtering invalid or missing data.")
            log_perf("confirm.import_rows", import_started_at)
            comparison_started_at = perf_counter()
            normalized_comparison_started_at = perf_counter()
            normalized_comparison = build_pre_normalized_quote_comparison(
                upload_id=session_id,
                name=Path(filename).stem.replace("_", " ").strip() or "Uploaded quote comparison",
                sourcing_need="",
                source_type="upload",
                bids=import_result["bids"]
            )
            compare_substeps = {
                "build_pre_normalized_comparison_ms": round((perf_counter() - normalized_comparison_started_at) * 1000, 1)
            }
            evaluation_started_at = perf_counter()
            evaluation = calculate_quote_comparison(normalized_comparison)
            compare_substeps["calculate_quote_comparison_ms"] = round((perf_counter() - evaluation_started_at) * 1000, 1)
            log_perf_details("confirm.compare_calculation.substeps", **compare_substeps)
            log_perf("confirm.compare_calculation", comparison_started_at)
            persistence_started_at = perf_counter()
            persist_quote_compare_analysis_results(
                normalized_comparison,
                source_name=filename or normalized_comparison["name"] or "Compare Prices analysis",
                upload_id=session_id,
                comparison_is_normalized=True,
                perf_label_prefix="confirm"
            )
            log_perf("confirm.persist_analysis", persistence_started_at)
            if session_id:
                session_payload_started_at = perf_counter()
                compact_evaluation_started_at = perf_counter()
                compact_evaluation = compact_quote_compare_evaluation_for_session(evaluation) or {}
                compact_evaluation_ms = (perf_counter() - compact_evaluation_started_at) * 1000
                lightweight_session_payload = {
                    "session_id": session_id,
                    "file_id": (session_payload or {}).get("file_id") or session_id,
                    "step": "analyze",
                    "filename": filename,
                    "file_path": (session_payload or {}).get("file_path") or (session_payload or {}).get("cached_upload_path"),
                    "cached_upload_path": (session_payload or {}).get("cached_upload_path"),
                    "headers": (session_payload or {}).get("headers", []),
                    "column_count": (session_payload or {}).get("column_count", len(df.columns)),
                    "row_count": (session_payload or {}).get("row_count", len(df.index)),
                    "mapping": parsed_mapping,
                    "field_reviews": (session_payload or {}).get("field_reviews", []),
                    "required_fields": (session_payload or {}).get("required_fields", QUOTE_COMPARE_REQUIRED_FIELDS),
                    "optional_fields": (session_payload or {}).get("optional_fields", QUOTE_COMPARE_OPTIONAL_FIELDS),
                    "matched_fields": (session_payload or {}).get("matched_fields", 0),
                    "missing_fields": (session_payload or {}).get("missing_fields", []),
                    "optional_columns": (session_payload or {}).get("optional_columns", []),
                    "review_message": (session_payload or {}).get("review_message", ""),
                    "message": (session_payload or {}).get("message", ""),
                    "comparison": normalized_comparison,
                    "evaluation": compact_evaluation,
                    "_json_safe_payload": True
                }
                logger.info(
                    "[PERF] confirm.save_active_session.serialization: %s",
                    json.dumps(
                        {
                            "compact_evaluation_ms": round(compact_evaluation_ms, 1),
                            "comparison_bytes": get_json_payload_size_bytes(lightweight_session_payload.get("comparison")),
                            "evaluation_bytes": get_json_payload_size_bytes(lightweight_session_payload.get("evaluation"))
                        },
                        ensure_ascii=False,
                        separators=(",", ":")
                    )
                )
                save_quote_compare_active_session(
                    session_id,
                    lightweight_session_payload,
                    perf_label_prefix="confirm"
                )
                log_perf("confirm.save_active_session", session_payload_started_at)
        except ValueError as exc:
            return JSONResponse({"success": False, "message": str(exc)}, status_code=400)
        except Exception:
            logger.exception("Failed to confirm Compare Prices upload: %s", filename)
            return JSONResponse(
                {
                    "success": False,
                    "message": "The supplier offer file could not be imported. Please review the mappings and try again."
                },
                status_code=400
            )

        response_started_at = perf_counter()
        response_payload_started_at = perf_counter()
        compact_response_evaluation_started_at = perf_counter()
        compact_response_evaluation = compact_quote_compare_evaluation_for_session(evaluation) or {}
        response_payload = {
            "success": True,
            "session_id": session_id,
            "comparison": normalized_comparison,
            "evaluation": compact_response_evaluation,
            "skipped_rows": import_result["skipped_row_count"],
            "message": (
                f"Imported supplier offers from {filename}. "
                f"{import_result['skipped_row_count']} rows skipped due to invalid or missing data."
            )
        }
        response_substeps = {
            "compact_evaluation_ms": round((perf_counter() - compact_response_evaluation_started_at) * 1000, 1),
            "build_payload_ms": round((perf_counter() - response_payload_started_at) * 1000, 1),
            "comparison_bytes": get_json_payload_size_bytes(normalized_comparison),
            "evaluation_bytes": get_json_payload_size_bytes(compact_response_evaluation),
            "evaluation_compacted_bytes_before": get_json_payload_size_bytes(evaluation),
            "evaluation_compacted_bytes_after": get_json_payload_size_bytes(compact_response_evaluation)
        }
        serialization_started_at = perf_counter()
        response_json = json.dumps(response_payload, ensure_ascii=False, separators=(",", ":"))
        response_substeps["serialize_ms"] = round((perf_counter() - serialization_started_at) * 1000, 1)
        response_substeps["payload_bytes"] = len(response_json.encode("utf-8"))
        response_object_started_at = perf_counter()
        response = Response(content=response_json, media_type="application/json")
        response_substeps["response_object_ms"] = round((perf_counter() - response_object_started_at) * 1000, 1)
        response_substeps["total_ms"] = round((perf_counter() - response_started_at) * 1000, 1)
        log_perf_details("confirm.response_build.substeps", **response_substeps)
        log_perf("confirm.response_build", response_started_at)
        log_perf("confirm.total", request_started_at)
        return response

    @app.post("/quote-compare/evaluate")
    def evaluate_quote_comparison(payload: QuoteComparisonPayload):
        try:
            normalized_comparison = normalize_quote_comparison_payload(payload.model_dump())
            evaluation = calculate_quote_comparison(normalized_comparison)
            persist_quote_compare_analysis_results(
                normalized_comparison,
                source_name=normalized_comparison["name"] or "Compare Prices analysis",
                upload_id=normalized_comparison.get("upload_id")
            )
        except ValueError as exc:
            return JSONResponse({"success": False, "message": str(exc)}, status_code=400)

        return JSONResponse({
            "success": True,
            "comparison": normalized_comparison,
            "evaluation": evaluation
        })

    @app.post("/quote-compare/delete")
    def delete_quote_comparison(payload: QuoteComparisonDeletePayload):
        comparison_id = str(payload.comparison_id or "").strip()
        if not comparison_id:
            return JSONResponse({"success": False, "message": "Choose a comparison to delete."}, status_code=400)

        store = load_quote_comparisons_store()
        comparisons = store.get("comparisons", [])
        existing_comparison = next(
            (comparison for comparison in comparisons if comparison.get("comparison_id") == comparison_id),
            None
        )
        if not existing_comparison:
            return JSONResponse({"success": False, "message": "The selected comparison could not be found."}, status_code=404)

        updated_comparisons = [
            comparison for comparison in comparisons if comparison.get("comparison_id") != comparison_id
        ]
        store["comparisons"] = updated_comparisons
        save_quote_comparisons_store(store)

        return JSONResponse({
            "success": True,
            "deleted_comparison_id": comparison_id,
            "comparisons": updated_comparisons,
            "message": f"Deleted comparison: {existing_comparison.get('name', 'Quote Comparison')}"
        })

    return app

templates = build_templates()
app = create_app()


# ===== AUTH / LICENSE SECTION =====

from pydantic import BaseModel
import sqlite3
from fastapi import Request
from fastapi.responses import RedirectResponse, JSONResponse
from datetime import datetime


AUTH_DB_PATH = Path(os.getenv("AUTH_DB_PATH", str(PERSIST_DIR / "auth.db"))).expanduser()
DEFAULT_TEST_LICENSE_CODE = "TEST-123-ABC"
SEED_DEFAULT_TEST_LICENSE = env_flag("SEED_DEFAULT_TEST_LICENSE", not IS_PRODUCTION)
AUTH_COOKIE_NAME = os.getenv("AUTH_COOKIE_NAME", "ppa_auth_session").strip() or "ppa_auth_session"
AUTH_COOKIE_MAX_AGE_SECONDS = env_int("AUTH_COOKIE_MAX_AGE_SECONDS", 60 * 60 * 24 * 14)
AUTH_COOKIE_SAMESITE = os.getenv("AUTH_COOKIE_SAMESITE", "lax").strip().lower() or "lax"
DEMO_COOKIE_NAME = os.getenv("DEMO_COOKIE_NAME", "ppa_demo_session").strip() or "ppa_demo_session"
DEMO_COOKIE_MAX_AGE_SECONDS = env_int("DEMO_COOKIE_MAX_AGE_SECONDS", 60 * 60 * 8)
if AUTH_COOKIE_SAMESITE not in {"lax", "strict", "none"}:
    logger.warning("Invalid AUTH_COOKIE_SAMESITE value %r. Falling back to 'lax'.", AUTH_COOKIE_SAMESITE)
    AUTH_COOKIE_SAMESITE = "lax"
AUTH_COOKIE_SECURE = env_flag("AUTH_COOKIE_SECURE", IS_PRODUCTION)
PASSWORD_HASH_ITERATIONS = env_int("PASSWORD_HASH_ITERATIONS", 200_000)
RESEND_API_KEY = (os.getenv("RESEND_API_KEY") or "").strip()
RESEND_FROM_EMAIL = (os.getenv("RESEND_FROM_EMAIL") or "onboarding@resend.dev").strip()
ADMIN_EMAIL = (
    os.getenv("ADMIN_EMAIL")
    or os.getenv("ADMIN_USER_EMAIL")
    or ""
).strip().lower()
AUTH_SECRET_KEY_FROM_ENV = (
    os.getenv("AUTH_SECRET_KEY")
    or os.getenv("SECRET_KEY")
    or ""
).strip()
AUTH_SECRET_KEY = AUTH_SECRET_KEY_FROM_ENV or f"dev-only-{BASE_DIR.resolve()}"
AUTH_SHOW_RESET_LINK_IN_RESPONSE = env_flag("AUTH_SHOW_RESET_LINK_IN_RESPONSE", not IS_PRODUCTION)

if IS_PRODUCTION and not AUTH_SECRET_KEY_FROM_ENV:
    raise RuntimeError("Production mode requires AUTH_SECRET_KEY or SECRET_KEY to be set.")

if IS_PRODUCTION and AUTH_SECRET_KEY.startswith("dev-only-"):
    raise RuntimeError("Production mode cannot use the development fallback auth secret.")

if IS_PRODUCTION and AUTH_SHOW_RESET_LINK_IN_RESPONSE:
    logger.warning("Disabling reset-link exposure because APP_ENV is production.")
    AUTH_SHOW_RESET_LINK_IN_RESPONSE = False

if AUTH_COOKIE_SAMESITE == "none" and not AUTH_COOKIE_SECURE:
    logger.warning("AUTH_COOKIE_SAMESITE='none' requires secure cookies. Enabling AUTH_COOKIE_SECURE.")
    AUTH_COOKIE_SECURE = True

if AUTH_SECRET_KEY.startswith("dev-only-"):
    logger.warning(
        "Using development fallback auth secret. Set AUTH_SECRET_KEY or SECRET_KEY for persistent secure sessions."
    )


def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    derived_key = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt.encode("utf-8"),
        PASSWORD_HASH_ITERATIONS
    )
    password_digest = base64.urlsafe_b64encode(derived_key).decode("ascii")
    return f"pbkdf2_sha256${PASSWORD_HASH_ITERATIONS}${salt}${password_digest}"

def hash_reset_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def create_password_reset_token() -> str:
    return secrets.token_urlsafe(32)


def verify_password(password: str, stored_hash: str | None) -> tuple[bool, bool]:
    if not stored_hash:
        return False, False

    if stored_hash.startswith("pbkdf2_sha256$"):
        try:
            _, iteration_text, salt, password_digest = stored_hash.split("$", 3)
            iterations = max(int(iteration_text), 1)
            derived_key = hashlib.pbkdf2_hmac(
                "sha256",
                password.encode("utf-8"),
                salt.encode("utf-8"),
                iterations
            )
            computed_digest = base64.urlsafe_b64encode(derived_key).decode("ascii")
            return hmac.compare_digest(computed_digest, password_digest), False
        except (TypeError, ValueError):
            logger.warning("Stored password hash has an invalid PBKDF2 format.")
            return False, False

    legacy_sha256 = hashlib.sha256(password.encode("utf-8")).hexdigest()
    return hmac.compare_digest(legacy_sha256, stored_hash), True


def create_auth_cookie_value(user_id: int) -> str:
    expires_at = int(datetime.now(timezone.utc).timestamp()) + AUTH_COOKIE_MAX_AGE_SECONDS
    payload = f"{user_id}:{expires_at}"
    signature = hmac.new(
        AUTH_SECRET_KEY.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()
    token = f"{payload}:{signature}"
    return base64.urlsafe_b64encode(token.encode("utf-8")).decode("ascii")


def create_demo_cookie_value() -> str:
    expires_at = int(datetime.now(timezone.utc).timestamp()) + DEMO_COOKIE_MAX_AGE_SECONDS
    payload = f"demo:{expires_at}"
    signature = hmac.new(
        AUTH_SECRET_KEY.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()
    token = f"{payload}:{signature}"
    return base64.urlsafe_b64encode(token.encode("utf-8")).decode("ascii")


def get_user_by_id(user_id: int) -> sqlite3.Row | None:
    conn = sqlite3.connect(str(AUTH_DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, email, is_active, is_admin FROM users WHERE id = ?",
            (user_id,)
        )
        return cursor.fetchone()
    finally:
        conn.close()


def resolve_authenticated_user_id(request: Request) -> tuple[int | None, bool]:
    raw_cookie = request.cookies.get(AUTH_COOKIE_NAME)
    if not raw_cookie:
        return None, False

    try:
        decoded_token = base64.urlsafe_b64decode(raw_cookie.encode("ascii")).decode("utf-8")
        user_id_text, expires_at_text, signature = decoded_token.split(":", 2)
        payload = f"{user_id_text}:{expires_at_text}"
        expected_signature = hmac.new(
            AUTH_SECRET_KEY.encode("utf-8"),
            payload.encode("utf-8"),
            hashlib.sha256
        ).hexdigest()
        if not hmac.compare_digest(signature, expected_signature):
            return None, True

        expires_at = int(expires_at_text)
        if expires_at <= int(datetime.now(timezone.utc).timestamp()):
            return None, True

        user_id = int(user_id_text)
    except (ValueError, TypeError, UnicodeDecodeError, base64.binascii.Error):
        return None, True

    user_row = get_user_by_id(user_id)
    if not user_row or not bool(user_row["is_active"]):
        return None, True

    return int(user_row["id"]), False


def resolve_demo_access(request: Request) -> tuple[bool, bool]:
    raw_cookie = request.cookies.get(DEMO_COOKIE_NAME)
    if not raw_cookie:
        return False, False

    try:
        decoded_token = base64.urlsafe_b64decode(raw_cookie.encode("ascii")).decode("utf-8")
        mode_name, expires_at_text, signature = decoded_token.split(":", 2)
        payload = f"{mode_name}:{expires_at_text}"
        expected_signature = hmac.new(
            AUTH_SECRET_KEY.encode("utf-8"),
            payload.encode("utf-8"),
            hashlib.sha256
        ).hexdigest()
        if mode_name != "demo" or not hmac.compare_digest(signature, expected_signature):
            return False, True
        expires_at = int(expires_at_text)
        if expires_at <= int(datetime.now(timezone.utc).timestamp()):
            return False, True
    except (ValueError, TypeError, UnicodeDecodeError, base64.binascii.Error):
        return False, True

    return True, False


def clear_auth_cookie(response: JSONResponse | RedirectResponse | HTMLResponse) -> None:
    response.delete_cookie(
        key=AUTH_COOKIE_NAME,
        path="/",
        httponly=True,
        samesite=AUTH_COOKIE_SAMESITE,
        secure=AUTH_COOKIE_SECURE
    )


def clear_demo_cookie(response: JSONResponse | RedirectResponse | HTMLResponse) -> None:
    response.delete_cookie(
        key=DEMO_COOKIE_NAME,
        path="/",
        httponly=True,
        samesite=AUTH_COOKIE_SAMESITE,
        secure=AUTH_COOKIE_SECURE
    )


def get_request_auth_user(request: Request) -> dict[str, Any] | None:
    auth_user = getattr(request.state, "auth_user", None)
    return auth_user if isinstance(auth_user, dict) else None


def require_admin_user(
    request: Request,
    *,
    failure_status_code: int = 303
) -> RedirectResponse | JSONResponse | None:
    auth_user = get_request_auth_user(request)
    if auth_user and bool(auth_user.get("is_admin")):
        return None

    if getattr(request.state, "auth_user_id", None):
        if failure_status_code == 303:
            return RedirectResponse(url="/", status_code=303)
        return build_auth_json_response(
            success=False,
            detail="Admin access required.",
            status_code=failure_status_code
        )

    if failure_status_code == 303:
        return RedirectResponse(url="/login", status_code=303)
    return build_auth_json_response(
        success=False,
        detail="Authentication required.",
        status_code=401
    )


def ensure_auth_db() -> None:
    AUTH_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(AUTH_DB_PATH))
    try:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                is_admin INTEGER NOT NULL DEFAULT 0
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS license_codes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                is_used INTEGER NOT NULL DEFAULT 0,
                used_by_user_id INTEGER,
                created_at TEXT,
                activated_at TEXT,
                used_at TEXT,
                status TEXT NOT NULL DEFAULT 'active'
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS password_reset_tokens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                token_hash TEXT NOT NULL UNIQUE,
                expires_at TEXT NOT NULL,
                created_at TEXT NOT NULL,
                used_at TEXT
            )
        """)
        user_columns = {
            row[1]
            for row in cursor.execute("PRAGMA table_info(users)").fetchall()
        }
        if "is_admin" not in user_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN is_admin INTEGER NOT NULL DEFAULT 0")

        existing_columns = {
            row[1]
            for row in cursor.execute("PRAGMA table_info(license_codes)").fetchall()
        }
        if "created_at" not in existing_columns:
            cursor.execute("ALTER TABLE license_codes ADD COLUMN created_at TEXT")
        if "used_at" not in existing_columns:
            cursor.execute("ALTER TABLE license_codes ADD COLUMN used_at TEXT")

        reset_token_columns = {
            row[1]
            for row in cursor.execute("PRAGMA table_info(password_reset_tokens)").fetchall()
        }
        if "used_at" not in reset_token_columns:
            cursor.execute("ALTER TABLE password_reset_tokens ADD COLUMN used_at TEXT")

        if SEED_DEFAULT_TEST_LICENSE:
            seeded_created_at = datetime.now().isoformat()
            cursor.execute("""
                INSERT OR IGNORE INTO license_codes (code, is_used, status, created_at)
                VALUES (?, 0, 'active', ?)
            """, (DEFAULT_TEST_LICENSE_CODE, seeded_created_at))
            cursor.execute("""
                UPDATE license_codes
                SET created_at = COALESCE(created_at, ?),
                    used_at = COALESCE(used_at, activated_at)
                WHERE code = ?
            """, (seeded_created_at, DEFAULT_TEST_LICENSE_CODE))
        elif IS_PRODUCTION:
            cursor.execute(
                "UPDATE license_codes SET status = 'inactive' WHERE code = ? AND is_used = 0",
                (DEFAULT_TEST_LICENSE_CODE,)
            )
        if ADMIN_EMAIL:
            cursor.execute(
                "UPDATE users SET is_admin = 1 WHERE lower(email) = ?",
                (ADMIN_EMAIL,)
            )
        conn.commit()
        logger.info("Auth database ready: %s", AUTH_DB_PATH)
    finally:
        conn.close()


def list_license_codes() -> list[dict[str, Any]]:
    conn = sqlite3.connect(str(AUTH_DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute("""
            SELECT
                lc.code,
                lc.is_used,
                lc.status,
                lc.used_by_user_id,
                CASE
                    WHEN lc.used_by_user_id IS NULL THEN '-'
                    WHEN u.email IS NULL OR trim(u.email) = '' THEN 'unknown'
                    ELSE u.email
                END AS used_by_email,
                lc.created_at,
                COALESCE(used_at, activated_at) AS used_at
            FROM license_codes lc
            LEFT JOIN users u
                ON lc.used_by_user_id = u.id
            ORDER BY
                datetime(lc.created_at) DESC,
                lc.id DESC
        """).fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()


def create_unique_license_code() -> str:
    conn = sqlite3.connect(str(AUTH_DB_PATH))
    try:
        cursor = conn.cursor()
        while True:
            candidate = generate_license_code()
            existing = cursor.execute(
                "SELECT 1 FROM license_codes WHERE code = ?",
                (candidate,)
            ).fetchone()
            if existing:
                continue

            created_at = datetime.now().isoformat()
            cursor.execute("""
                INSERT INTO license_codes (code, is_used, status, created_at)
                VALUES (?, 0, 'active', ?)
            """, (candidate, created_at))
            conn.commit()
            return candidate
    finally:
        conn.close()


def build_password_reset_link(request: Request, token: str) -> str:
    base_url = str(request.base_url).rstrip("/")
    return f"{base_url}/reset-password?{urlencode({'token': token})}"


def send_password_reset_email(email: str, reset_link: str) -> None:
    print("RESEND_FROM_EMAIL =", RESEND_FROM_EMAIL)

    if not RESEND_API_KEY:
        raise RuntimeError("RESEND_API_KEY is not configured.")

    payload = json.dumps({
        "from": RESEND_FROM_EMAIL,
        "to": [email],
        "subject": "Reset your password",
        "text": f"Click the link below to reset your password:\n\n{reset_link}"
    }).encode("utf-8")
    request = UrlRequest(
        "https://api.resend.com/emails",
        data=payload,
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "Content-Type": "application/json",
            "User-Agent": "purchase-price-analyzer/1.0"
        },
        method="POST"
    )
    with urlopen(request, timeout=15) as response:
        status_code = getattr(response, "status", response.getcode())
        if int(status_code) >= 300:
            raise RuntimeError(f"Resend request failed with status {status_code}.")

def create_password_reset_request(email: str, request: Request) -> str | None:
    normalized_email = str(email or "").strip().lower()
    if not normalized_email:
        return None

    conn = sqlite3.connect(str(AUTH_DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        cursor = conn.cursor()
        user = cursor.execute(
            "SELECT id, email, is_active FROM users WHERE email = ?",
            (normalized_email,)
        ).fetchone()
        if not user or not bool(user["is_active"]):
            return None

        raw_token = create_password_reset_token()
        created_at = datetime.now(timezone.utc)
        expires_at = created_at.timestamp() + PASSWORD_RESET_TOKEN_TTL_SECONDS
        cursor.execute(
            "UPDATE password_reset_tokens SET used_at = COALESCE(used_at, ?) WHERE user_id = ? AND used_at IS NULL",
            (created_at.isoformat(), int(user["id"]))
        )
        cursor.execute("""
            INSERT INTO password_reset_tokens (user_id, token_hash, expires_at, created_at, used_at)
            VALUES (?, ?, ?, ?, NULL)
        """, (
            int(user["id"]),
            hash_reset_token(raw_token),
            datetime.fromtimestamp(expires_at, tz=timezone.utc).isoformat(),
            created_at.isoformat()
        ))
        conn.commit()
        reset_link = build_password_reset_link(request, raw_token)
        if AUTH_SHOW_RESET_LINK_IN_RESPONSE:
            logger.info("Password reset link for %s: %s", normalized_email, reset_link)
        else:
            logger.info("Password reset requested for %s", normalized_email)
        if IS_PRODUCTION:
            try:
                send_password_reset_email(normalized_email, reset_link)
                logger.info("Password reset email sent to %s via Resend", normalized_email)
            except Exception:
                logger.exception("Failed to send password reset email via Resend for %s", normalized_email)
        return reset_link
    finally:
        conn.close()


def get_password_reset_token_record(token: str) -> tuple[sqlite3.Row | None, str | None]:
    normalized_token = str(token or "").strip()
    if not normalized_token:
        return None, "Reset link is missing or invalid."

    conn = sqlite3.connect(str(AUTH_DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute("""
            SELECT
                prt.id,
                prt.user_id,
                prt.expires_at,
                prt.created_at,
                prt.used_at,
                u.email,
                u.is_active
            FROM password_reset_tokens prt
            JOIN users u
                ON prt.user_id = u.id
            WHERE prt.token_hash = ?
        """, (hash_reset_token(normalized_token),)).fetchone()
    finally:
        conn.close()

    if not row:
        return None, "This reset link is invalid."
    if row["used_at"]:
        return None, "This reset link has already been used."
    try:
        expires_at = datetime.fromisoformat(str(row["expires_at"]))
    except ValueError:
        return None, "This reset link is invalid."
    if expires_at <= datetime.now(timezone.utc):
        return None, "This reset link has expired."
    if not bool(row["is_active"]):
        return None, "This account is not active."
    return row, None


def consume_password_reset_token(token: str, new_password: str) -> tuple[bool, str]:
    token_row, error_message = get_password_reset_token_record(token)
    if token_row is None:
        return False, error_message or "Reset failed."

    updated_at = datetime.now(timezone.utc).isoformat()
    conn = sqlite3.connect(str(AUTH_DB_PATH))
    try:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE users SET password_hash = ? WHERE id = ?",
            (hash_password(new_password), int(token_row["user_id"]))
        )
        cursor.execute(
            "UPDATE password_reset_tokens SET used_at = ? WHERE user_id = ? AND used_at IS NULL",
            (updated_at, int(token_row["user_id"]))
        )
        conn.commit()
        return True, "Your password has been reset successfully."
    except Exception:
        conn.rollback()
        logger.exception("Password reset completion failed for user_id=%s", token_row["user_id"])
        return False, "Password reset failed."
    finally:
        conn.close()


@app.middleware("http")
async def check_auth(request: Request, call_next):
    path = request.url.path

    if path in UPLOAD_ROUTE_PATHS:
        raw_content_length = str(request.headers.get("content-length") or "").strip()
        if raw_content_length:
            try:
                content_length = int(raw_content_length)
            except ValueError:
                content_length = 0
            if content_length > (MAX_UPLOAD_SIZE_BYTES + 1024 * 1024):
                return build_upload_size_error_response()

    allowed_paths = [
        "/login",
        "/demo",
        "/activate",
        "/forgot-password",
        "/reset-password",
        "/api/login",
        "/api/logout",
        "/api/activate",
        "/api/forgot-password",
        "/api/reset-password",
        "/docs",
        "/openapi.json",
        "/redoc",
        "/static",
        "/favicon.ico"
    ]

    if any(path.startswith(p) for p in allowed_paths):
        auth_user_id, should_clear_cookie = resolve_authenticated_user_id(request)
        demo_mode, should_clear_demo_cookie = resolve_demo_access(request)
        if auth_user_id:
            demo_mode = False
        request.state.auth_user_id = auth_user_id
        request.state.demo_mode = demo_mode
        auth_user = dict(get_user_by_id(auth_user_id)) if auth_user_id else None
        request.state.auth_user = auth_user
        request.state.auth_is_admin = bool(auth_user and auth_user.get("is_admin"))
        storage_token = CURRENT_STORAGE_USER_ID.set(auth_user_id)
        try:
            response = await call_next(request)
            if should_clear_cookie:
                clear_auth_cookie(response)
            if should_clear_demo_cookie:
                clear_demo_cookie(response)
            return response
        finally:
            CURRENT_STORAGE_USER_ID.reset(storage_token)

    auth_user_id, should_clear_cookie = resolve_authenticated_user_id(request)
    demo_mode, should_clear_demo_cookie = resolve_demo_access(request)
    if auth_user_id:
        demo_mode = False
    request.state.auth_user_id = auth_user_id
    request.state.demo_mode = demo_mode
    auth_user = dict(get_user_by_id(auth_user_id)) if auth_user_id else None
    request.state.auth_user = auth_user
    request.state.auth_is_admin = bool(auth_user and auth_user.get("is_admin"))
    if not auth_user_id and not demo_mode:
        response = RedirectResponse(url="/login", status_code=303)
        if should_clear_cookie:
            clear_auth_cookie(response)
        if should_clear_demo_cookie:
            clear_demo_cookie(response)
        return response

    storage_token = CURRENT_STORAGE_USER_ID.set(auth_user_id)
    try:
        response = await call_next(request)
        if should_clear_cookie:
            clear_auth_cookie(response)
        if should_clear_demo_cookie:
            clear_demo_cookie(response)
        return response
    finally:
        CURRENT_STORAGE_USER_ID.reset(storage_token)


@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    redirect_response = redirect_if_authenticated(request)
    if redirect_response is not None:
        return redirect_response

    return safe_template_response(
        request,
        "login.html",
        {
            "request": request
        }
    )


@app.get("/demo")
async def enter_demo_mode():
    response = RedirectResponse(url="/?demo=1", status_code=303)
    response.set_cookie(
        key=DEMO_COOKIE_NAME,
        value=create_demo_cookie_value(),
        httponly=True,
        samesite=AUTH_COOKIE_SAMESITE,
        secure=AUTH_COOKIE_SECURE,
        path="/",
        max_age=DEMO_COOKIE_MAX_AGE_SECONDS
    )
    clear_auth_cookie(response)
    return response


@app.get("/activate", response_class=HTMLResponse)
async def activate_page(request: Request):
    redirect_response = redirect_if_authenticated(request)
    if redirect_response is not None:
        return redirect_response

    return safe_template_response(
        request,
        "activate.html",
        {
            "request": request
        }
    )


@app.get("/forgot-password", response_class=HTMLResponse)
async def forgot_password_page(request: Request):
    redirect_response = redirect_if_authenticated(request)
    if redirect_response is not None:
        return redirect_response

    return safe_template_response(
        request,
        "forgot_password.html",
        {
            "request": request,
            "reset_link_enabled": AUTH_SHOW_RESET_LINK_IN_RESPONSE
        }
    )


@app.get("/reset-password", response_class=HTMLResponse)
async def reset_password_page(request: Request, token: str | None = None):
    token_row, token_error = get_password_reset_token_record(token or "")
    return safe_template_response(
        request,
        "reset_password.html",
        {
            "request": request,
            "token": str(token or "").strip(),
            "token_valid": token_row is not None,
            "token_error": token_error,
        }
    )


class ActivatePayload(BaseModel):
    email: str
    password: str
    code: str


class ForgotPasswordPayload(BaseModel):
    email: str


class ResetPasswordPayload(BaseModel):
    token: str
    password: str
    confirm_password: str


def build_auth_json_response(
    *,
    success: bool,
    message: str | None = None,
    detail: str | None = None,
    status_code: int = 200
) -> JSONResponse:
    payload: dict[str, Any] = {"success": success}
    if message is not None:
        payload["message"] = message
    if detail is not None:
        payload["detail"] = detail
    return JSONResponse(content=payload, status_code=status_code)


def get_license_code_status(license_row: sqlite3.Row) -> str | None:
    for key in ("status", "code_status", "license_status"):
        if key in license_row.keys():
            value = str(license_row[key] or "").strip().lower()
            return value or None

    for key in ("is_active", "active"):
        if key in license_row.keys():
            return "active" if bool(license_row[key]) else "inactive"

    return None


@app.post("/api/activate")
async def activate(payload: ActivatePayload):
    email = str(payload.email or "").strip().lower()
    password = str(payload.password or "")
    code = str(payload.code or "").strip()
    conn: sqlite3.Connection | None = None

    if not code:
        return build_auth_json_response(
            success=False,
            detail="Access code is required.",
            status_code=400
        )
    if not email:
        return build_auth_json_response(
            success=False,
            detail="Email is required.",
            status_code=400
        )
    if not password:
        return build_auth_json_response(
            success=False,
            detail="Password is required.",
            status_code=400
        )

    try:
        conn = sqlite3.connect(str(AUTH_DB_PATH))
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM license_codes WHERE code = ?", (code,))
        license_row = cursor.fetchone()

        if not license_row:
            return build_auth_json_response(
                success=False,
                detail="Invalid code",
                status_code=400
            )

        if bool(license_row["is_used"]) if "is_used" in license_row.keys() else False:
            return build_auth_json_response(
                success=False,
                detail="Code already used",
                status_code=400
            )

        license_status = get_license_code_status(license_row)
        if license_status is not None and license_status != "active":
            return build_auth_json_response(
                success=False,
                detail="Code not active",
                status_code=400
            )

        existing_user = cursor.execute(
            "SELECT id FROM users WHERE email = ?",
            (email,)
        ).fetchone()
        if existing_user:
            return build_auth_json_response(
                success=False,
                detail="An account with this email already exists.",
                status_code=400
            )

        password_hash = hash_password(password)
        activated_at = datetime.now().isoformat()

        is_admin = 1 if ADMIN_EMAIL and email == ADMIN_EMAIL else 0
        cursor.execute("""
            INSERT INTO users (email, password_hash, created_at, is_active, is_admin)
            VALUES (?, ?, ?, ?, ?)
        """, (email, password_hash, activated_at, 1, is_admin))
        user_id = cursor.lastrowid

        cursor.execute("""
            UPDATE license_codes
            SET is_used = 1,
                used_by_user_id = ?,
                activated_at = ?,
                used_at = ?
            WHERE code = ?
        """, (user_id, activated_at, activated_at, code))

        if cursor.rowcount != 1:
            raise RuntimeError("License code update did not complete as expected.")

        conn.commit()
        return build_auth_json_response(
            success=True,
            message="Account activated"
        )
    except sqlite3.IntegrityError as exc:
        if conn is not None:
            conn.rollback()
        logger.exception("Activation integrity error for email=%s code=%s", email, code)
        return build_auth_json_response(
            success=False,
            detail="Account activation failed because the account or code already exists.",
            status_code=400
        )
    except Exception as exc:
        if conn is not None:
            conn.rollback()
        logger.exception("Activation failed for email=%s code=%s", email, code)
        return build_auth_json_response(
            success=False,
            detail=str(exc) or "Activation failed.",
            status_code=500
        )
    finally:
        if conn is not None:
            conn.close()


class LoginPayload(BaseModel):
    email: str
    password: str


@app.post("/api/login")
async def login(payload: LoginPayload, request: Request):
    email = str(payload.email or "").strip().lower()
    password = str(payload.password or "")
    retry_after = check_named_rate_limit(request, "login", identifier_suffix=email)
    if retry_after is not None:
        return build_rate_limited_json_response(
            "Too many login attempts. Please wait a minute and try again.",
            retry_after
        )

    conn = sqlite3.connect(str(AUTH_DB_PATH))
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, password_hash, is_active FROM users WHERE email = ?",
            (email,)
        )
        user = cursor.fetchone()

        if not user:
            return {"success": False, "message": "User not found"}

        user_id, stored_hash, is_active = user
        if not bool(is_active):
            return {"success": False, "message": "Account inactive"}

        password_valid, requires_upgrade = verify_password(password, stored_hash)
        if not password_valid:
            return {"success": False, "message": "Wrong password"}

        if requires_upgrade:
            cursor.execute(
                "UPDATE users SET password_hash = ? WHERE id = ?",
                (hash_password(password), user_id)
            )
            conn.commit()
    finally:
        conn.close()

    response = JSONResponse(
        content={"success": True, "message": "Login successful"}
    )
    response.set_cookie(
        key=AUTH_COOKIE_NAME,
        value=create_auth_cookie_value(int(user_id)),
        httponly=True,
        samesite=AUTH_COOKIE_SAMESITE,
        secure=AUTH_COOKIE_SECURE,
        path="/",
        max_age=AUTH_COOKIE_MAX_AGE_SECONDS
    )
    clear_demo_cookie(response)
    return response


@app.post("/api/forgot-password")
async def forgot_password(payload: ForgotPasswordPayload, request: Request):
    email = str(payload.email or "").strip().lower()
    reset_link = None
    retry_after = check_named_rate_limit(request, "forgot_password", identifier_suffix=email)
    if retry_after is not None:
        return build_rate_limited_json_response(
            "Too many password reset requests. Please wait a few minutes and try again.",
            retry_after
        )

    try:
        if email:
            reset_link = create_password_reset_request(email, request)
    except Exception:
        logger.exception("Password reset request failed for email=%s", email)

    response_payload: dict[str, Any] = {
        "success": True,
        "message": "If an account exists for this email, reset instructions have been sent."
    }
    if reset_link and AUTH_SHOW_RESET_LINK_IN_RESPONSE:
        response_payload["reset_link"] = reset_link
    return JSONResponse(content=response_payload, status_code=200)


@app.post("/api/reset-password")
async def reset_password(payload: ResetPasswordPayload, request: Request):
    token = str(payload.token or "").strip()
    password = str(payload.password or "")
    confirm_password = str(payload.confirm_password or "")
    retry_after = check_named_rate_limit(request, "reset_password", identifier_suffix=hash_reset_token(token) if token else "")
    if retry_after is not None:
        return build_rate_limited_json_response(
            "Too many password reset attempts. Please wait a few minutes and try again.",
            retry_after
        )

    if not token:
        return build_auth_json_response(
            success=False,
            detail="Reset link is missing or invalid.",
            status_code=400
        )
    if not password:
        return build_auth_json_response(
            success=False,
            detail="New password is required.",
            status_code=400
        )
    if password != confirm_password:
        return build_auth_json_response(
            success=False,
            detail="Passwords do not match.",
            status_code=400
        )

    success, message = consume_password_reset_token(token, password)
    return build_auth_json_response(
        success=success,
        message=message if success else None,
        detail=None if success else message,
        status_code=200 if success else 400
    )


@app.post("/api/logout")
async def auth_logout():
    response = JSONResponse(
        content={"success": True, "message": "Logout successful"}
    )
    clear_auth_cookie(response)
    clear_demo_cookie(response)
    return response


@app.get("/admin/licenses", response_class=HTMLResponse)
async def admin_licenses_page(request: Request):
    admin_gate = require_admin_user(request)
    if admin_gate is not None:
        return admin_gate

    success_message = str(request.query_params.get("success") or "").strip()
    error_message = str(request.query_params.get("error") or "").strip()
    return safe_template_response(
        request,
        "admin_licenses.html",
        {
            "request": request,
            "licenses": list_license_codes(),
            "generated_code": success_message,
            "error_message": error_message,
            "is_admin_only": True
        }
    )


@app.post("/admin/licenses/generate")
async def admin_generate_license(request: Request):
    admin_gate = require_admin_user(request, failure_status_code=403)
    if admin_gate is not None:
        return admin_gate
    retry_after = check_named_rate_limit(
        request,
        "admin_license_generate",
        identifier_suffix=str(getattr(request.state, "auth_user_id", "") or "admin")
    )
    if retry_after is not None:
        redirect_url = "/admin/licenses?" + urlencode({
            "error": "Too many license generation requests. Please wait a minute and try again."
        })
        return RedirectResponse(url=redirect_url, status_code=303)

    new_code = create_unique_license_code()
    redirect_url = "/admin/licenses?" + urlencode({"success": new_code})
    return RedirectResponse(url=redirect_url, status_code=303)







